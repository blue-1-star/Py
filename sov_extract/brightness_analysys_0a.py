import os
import rawpy
import openpyxl
from openpyxl.styles import Font
from PIL import Image, ImageStat, ImageDraw
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
from datetime import datetime
from openpyxl import load_workbook
import tempfile
# Функция обработки изображений (включая ORF)
def process_image(img_path):
    if img_path.lower().endswith('.orf'):
        with rawpy.imread(img_path) as raw:
            rgb = raw.postprocess()
        img = Image.fromarray(rgb)
    else:
        img = Image.open(img_path)
    return img

# Функция выделения области изображения
def crop_image(image, shape='square', size=100):
    """
    Вырезает заданную область (квадрат или окружность) из изображения.
    """
    width, height = image.size
    center_x, center_y = width // 2, height // 2

    if shape == 'square':
        left = max(center_x - size // 2, 0)
        upper = max(center_y - size // 2, 0)
        right = min(center_x + size // 2, width)
        lower = min(center_y + size // 2, height)
        return image.crop((left, upper, right, lower))

    elif shape == 'circle':
        mask = Image.new("L", image.size, 0)
        draw = ImageDraw.Draw(mask)
        draw.ellipse(
            (center_x - size // 2, center_y - size // 2, 
             center_x + size // 2, center_y + size // 2),
            fill=255
        )
        result = Image.composite(image, Image.new("RGB", image.size, (0, 0, 0)), mask)
        bbox = mask.getbbox()
        return result.crop(bbox)

# Функции расчёта яркости
def calculate_brightness_pil(image_path, lower_threshold=0, upper_threshold=255):
    # img = process_image(image_path)  # Завантажуємо та обробляємо зображення
    img = process_image(image_path) 
    # image = Image.open(image_path)
    grayscale_img = img.convert('L')
    pixel_values = np.array(grayscale_img)
    filtered_pixels = pixel_values[(pixel_values >= lower_threshold) & (pixel_values <= upper_threshold)]
    mean_brightness = np.mean(filtered_pixels) if len(filtered_pixels) > 0 else 0
    return {
        'mean_brightness': mean_brightness,
        'total_pixels': pixel_values.size
    }

def calculate_brightness_color(image_path, lower_threshold=0, upper_threshold=255):
    img = process_image(image_path) 
    # img = Image.open(image_path)
    pixel_values = np.array(img)
    brightness_values = 0.299 * pixel_values[:, :, 0] + 0.587 * pixel_values[:, :, 1] + 0.114 * pixel_values[:, :, 2]
    filtered_pixels = brightness_values[(brightness_values >= lower_threshold) & (brightness_values <= upper_threshold)]
    mean_brightness = np.mean(filtered_pixels) if len(filtered_pixels) > 0 else 0
    return {
        'mean_brightness': mean_brightness,
        'total_pixels': brightness_values.size
    }

# def calculate_brightness_with_area(image_path, shape='square', size=100, lower_threshold=0, upper_threshold=255):
#     img = process_image(image_path) 
#     # img = Image.open(image_path)
#     cropped_img = crop_image(img, shape=shape, size=size)
#     return calculate_brightness_pil(cropped_img, lower_threshold, upper_threshold)


def calculate_brightness_with_area(image_path, shape='square', size=100, lower_threshold=0, upper_threshold=255):
    img = process_image(image_path) 
    cropped_img = crop_image(img, shape=shape, size=size)
    
    with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as temp_file:
        cropped_img.save(temp_file.name)
        temp_path = temp_file.name
    
    try:
        return calculate_brightness_pil(temp_path, lower_threshold, upper_threshold)
    finally:
        os.remove(temp_path)

def calculate_color_with_area(image_path, shape='square', size=100, lower_threshold=0, upper_threshold=255):
    """Calculates detailed color analysis for the selected area of the image."""
    img = process_image(image_path)
    cropped_img = crop_image(img, shape=shape, size=size)

    # Convert cropped image to numpy array
    pixel_values = np.array(cropped_img)

    # Calculate brightness values
    brightness_values = 0.299 * pixel_values[:, :, 0] + 0.587 * pixel_values[:, :, 1] + 0.114 * pixel_values[:, :, 2]

    # Filter pixels based on brightness thresholds
    mask = (brightness_values >= lower_threshold) & (brightness_values <= upper_threshold)
    filtered_pixels = pixel_values[mask]

    # Calculate average color if there are filtered pixels
    if len(filtered_pixels) > 0:
        avg_r = np.average(filtered_pixels[:, 0], weights=(0.299 * filtered_pixels[:, 0]))
        avg_g = np.average(filtered_pixels[:, 1], weights=(0.587 * filtered_pixels[:, 1]))
        avg_b = np.average(filtered_pixels[:, 2], weights=(0.114 * filtered_pixels[:, 2]))
        avg_r = int(avg_r)
        avg_g = int(avg_g)
        avg_b = int(avg_b)
        # return int(avg_r), int(avg_g), int(avg_b)
        # avg_color_hex = f'#{avg_color[0]:02x}{avg_color[1]:02x}{avg_color[2]:02x}'
        avg_color_hex = f'#{avg_r:02x}{avg_g:02x}{avg_b:02x}'
        return avg_color_hex

    



# Функция для создания DataFrame с результатами
def calculate_brightness_dataframe(image_dir, lower_threshold, size):
    image_files = [f for f in os.listdir(image_dir) if f.lower().endswith(('.orf', '.jpg', '.jpeg' ))]
    # image_files = [f for f in os.listdir(image_dir) if f.lower().endswith(('.orf', '.jpg', '.jpeg','.png' ))]
    results = []
    
    for img_file in image_files:
        img_path = os.path.join(image_dir, img_file)
        img = process_image(img_path)
        brightness_pil = calculate_brightness_pil(img_path, lower_threshold)
        brightness_color = calculate_brightness_color(img_path, lower_threshold)
        # brightness_square = calculate_brightness_with_area(img_path, shape='square', size, lower_threshold=lower_threshold)
        brightness_square = calculate_brightness_with_area(img_path, 'square', size, lower_threshold=lower_threshold)
        brightness_circle = calculate_brightness_with_area(img_path, 'circle', size, lower_threshold=lower_threshold)
        avg_color_circle = calculate_color_with_area(img_path, 'circle', size, lower_threshold=lower_threshold)
        avg_col_square = calculate_color_with_area(img_path, 'square', size, lower_threshold=lower_threshold)

        file_format = "ORF" if img_file.lower().endswith('.orf') else "JPEG"

        results.append({
            "Filename": img_file,
            "Format": file_format,
            "Brightness_PIL": brightness_pil['mean_brightness'],
            "Brightness_Color": brightness_color['mean_brightness'],
            "Brightness_Square": brightness_square['mean_brightness'],
            "Brightness_Circle": brightness_circle['mean_brightness'],
            "col_cir": avg_color_circle,
            "col_sq": avg_col_square,
        })

    df = pd.DataFrame(results)

    # Добавление рангов
    df['Rank_in_Group_PIL'] = df.groupby('Format')['Brightness_PIL'].rank(ascending=False).astype(int)
    df['Rank_Global_PIL'] = df['Brightness_PIL'].rank(ascending=False).astype(int)
    df['Rank_in_Group_Color'] = df.groupby('Format')['Brightness_Color'].rank(ascending=False).astype(int)
    df['Rank_Global_Color'] = df['Brightness_Color'].rank(ascending=False).astype(int)

    return df

# Функция сохранения в Excel
# def save_brightness_excel(df, output_file):
#     with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
#         df.sort_values(['Filename'], inplace=True)
#         df.to_excel(writer, index=False, sheet_name='Original Order')
#         df.sort_values(['Format', 'Rank_in_Group_PIL'], inplace=True)
#         df.to_excel(writer, index=False, sheet_name='Sorted by Brightness_PIL')
#         df.sort_values(['Format', 'Rank_in_Group_Color'], inplace=True)
#         df.to_excel(writer, index=False, sheet_name='Sorted by Brightness_Color')
def auto_adjust_column_width(output_file):
    workbook = load_workbook(output_file)

    # Перебираємо всі аркуші
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        
        for col in worksheet.columns:
            max_length = 0
            col_letter = col[0].column_letter  # Отримуємо букву стовпця
            
            # Перебираємо всі комірки в стовпці
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            # Встановлюємо ширину стовпця (додаємо запас у 2 символи)
            adjusted_width = max_length + 2
            worksheet.column_dimensions[col_letter].width = adjusted_width

    # Зберігаємо відкоригований файл
    workbook.save(output_file)

def save_brightness_excel(df, output_file, lower_threshold):
    # Створюємо Excel-файл з трьома листами
    summary = { 
        'date'  : datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'files' : len(df),
        'lower_threshold' : lower_threshold
    }
    summary_df = pd.DataFrame([summary])

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Лист 1: Оригінальний порядок
        # df(['Filename'], inplace=True)
        df['Format'] = pd.Categorical(df['Format'], categories=['ORF', 'JPEG'], ordered=True)
        df = df.sort_values(
        by=['Format', 'Filename'], 
        key=lambda col: col.str.extract(r'(\d+)')[0].astype(int) if col.name == 'Filename' else col)        
        df1 = df.drop(columns=['Format'])
        rename_map = col_ren_bright_excel()
        df1.rename(columns=rename_map, inplace=True) 
        # df1 = df1.rename(columns={'Brightness_PIL': 'Bright_PIL'})
        sheet_name = 'Original Order'
        df1.to_excel(writer, index=False, sheet_name=sheet_name)
        format_dataframe_columns_to_excel(writer,sheet_name, df1, ['Bright_PIL', 'Bright_Col', 'Bright_Sq', 'Bright_Sc'],\
        [1, 1, 1, 1, 1])
        summary_df.to_excel(writer, sheet_name='Original Order', index=False, startrow=len(df) + 2)

        # Лист 2: Сортовані по яскравості всередині груп (PIL)
        df.sort_values(['Format', 'Rank_in_Group_PIL'], inplace=True)
        df.to_excel(writer, index=False, sheet_name='Sorted by Brightness_PIL')
        summary_df.to_excel(writer, sheet_name='Sorted by Brightness_PIL', index=False, startrow=len(df) + 2)

        # Лист 3: Сортовані по яскравості всередині груп (Color)
        df.sort_values(['Format', 'Rank_in_Group_Color'], inplace=True)
        df.to_excel(writer, index=False, sheet_name='Sorted by Brightness_Color')
        summary_df.to_excel(writer, sheet_name='Sorted by Brightness_Color', index=False, startrow=len(df) + 2)

        # Лист 4: Глобальне сортування по яскравості (PIL)
        df.sort_values(['Rank_Global_PIL'], inplace=True)
        df.to_excel(writer, index=False, sheet_name='Global Sorted by PIL')
        summary_df.to_excel(writer, sheet_name='Global Sorted by PIL', index=False, startrow=len(df) + 2)
        
        # Лист 5: Глобальне сортування по яскравості (Color)
        df.sort_values(['Rank_Global_Color'], inplace=True)
        df.to_excel(writer, index=False, sheet_name='Global Sorted by Color')
        summary_df.to_excel(writer, sheet_name='Global Sorted by Color', index=False, startrow=len(df) + 2)

    # print(f"Excel-файл '{output_file}' створено успішно!")
# графика
import calc_brightness_plt_1, calc_brightness_plt_1a
def col_ren_bright_excel():
    return {
        'Brightness_PIL': 'Bright_PIL',
        'Brightness_Color': 'Bright_Col',
        'Brightness_Square': 'Bright_Sq',
        'Brightness_Circle': 'Bright_Sc',        
        'avg_color_circle': 'Col_Sc',
        'avg_color_square': 'Col_sq',        
        # Добавьте другие столбцы
    }



def format_dataframe_columns_to_excel(writer, sheet_name, df, columns, decimal_places):
    # Применение стиля к указанным столбцам
    sheet = writer.sheets[sheet_name]
    for column, decimals in zip(columns, decimal_places):
        if column in df.columns and df[column].dtype in ['float64', 'int64']:
            number_format = f'0.{"0" * decimals}'
            col_idx = df.columns.get_loc(column) + 1  # Индекс столбца (Excel 1-based)
            for row_idx in range(2, len(df) + 2):  # Excel строки начинаются с 1, учитываем заголовок
                cell = sheet.cell(row=row_idx, column=col_idx)
                cell.number_format = number_format
        else:
            print(f"Столбец {column} не является числовым или отсутствует в DataFrame и не будет отформатирован.")

# Пример использования функции
# data = {
#     'A': [1.12345, 2.54321, 3.6789],
#     'B': [4.12345, 5.54321, 6.6789]
# }
# df = pd.DataFrame(data)
# format_dataframe_column_to_excel(df, 'A', 2, 'пример.xlsx')


# Основной блок выполнения
if __name__ == "__main__":
    image_dir = r"G:\My\sov\extract"  # Ваш путь к папке с изображениями
    current_date = datetime.now().strftime("%d_%m")
    output_dir = os.path.join(os.path.dirname(__file__), 'Data')
    output_file = os.path.join(output_dir, f"brightness_analysis_{current_date}.xlsx")
    lower_threshold = 0
    size = 2000

    # df = calculate_brightness_dataframe(image_dir, lower_threshold, size)
    cache_file = os.path.join(image_dir, "brightness_data.csv")

    if os.path.exists(cache_file):
        df = pd.read_csv(cache_file)
        print("Data loaded from cache.")
    else:
        df = calculate_brightness_dataframe(image_dir, lower_threshold, size)
        df.to_csv(cache_file, index=False)
        print("Brightness calculations completed and cached.")


    save_brightness_excel(df, output_file, lower_threshold)
    auto_adjust_column_width(output_file)
    print(f"Анализ яркости завершён. Результаты сохранены в {output_file}")

