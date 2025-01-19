import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime, timedelta
import os
from brightness_analysys_0a import auto_adjust_column_width, format_dataframe_columns_to_excel 
# from coloring_columns_hex_2 import insert_empty_columns 

def insert_empty_columns(df, columns_list):
    """
    Вставляет пустой столбец после каждого столбца из списка columns_list в датафрейм df.

    :param df: pd.DataFrame, исходный датафрейм
    :param columns_list: list, список имен столбцов, после которых нужно вставить пустые столбцы
    :return: pd.DataFrame, новый датафрейм с добавленными пустыми столбцами
    """
    df = df.copy()  # Создаем копию датафрейма, чтобы избежать изменений оригинала
    
    for col in columns_list:
        if col in df.columns:
            empty_col_name = f"{col}_empty"
            col_index = df.columns.get_loc(col) + 1  # Индекс столбца, после которого добавляем пустой
            df.insert(col_index, empty_col_name, None)  # Вставляем пустой столбец

    return df

def add_color_columns_to_excel(input_file, output_file, columns_with_hex):
    # Читаем Excel файл в DataFrame
    df = pd.read_excel(input_file)
    
    # Загружаем файл для работы с openpyxl
    wb = load_workbook(input_file)
    ws = wb.active

    # Проходим по указанным столбцам с HEX-кодами
    for col in columns_with_hex:
        if col not in df.columns:
            print(f"Столбец '{col}' не найден в таблице. Пропуск.")
            continue

        # Определяем индекс текущего столбца
        col_index = df.columns.get_loc(col) + 1  # Индекс столбца в openpyxl (нумерация с 1)
        new_col_index = col_index + 1  # Новый столбец будет добавлен сразу после текущего

        # Вставляем новый столбец в Excel
        ws.insert_cols(new_col_index)

        # Копируем данные из исходного столбца в новый
        for row in range(2, ws.max_row + 1):  # Пропускаем заголовок
            ws.cell(row=row, column=new_col_index).value = ws.cell(row=row, column=col_index).value

        # Закрашиваем ячейки нового столбца
        for i, hex_code in enumerate(df[col]):
            excel_row = i + 2  # Нумерация строк в Excel (с учётом заголовков)
            if isinstance(hex_code, str) and hex_code.startswith('#') and len(hex_code) == 7:
                try:
                    cell = ws.cell(row=excel_row, column=new_col_index)
                    cell.fill = PatternFill(start_color=hex_code[1:], end_color=hex_code[1:], fill_type="solid")
                except Exception as e:
                    print(f"Ошибка обработки цвета '{hex_code}' в строке {excel_row}: {e}")
            else:
                print(f"Некорректный HEX-код '{hex_code}' в строке {excel_row}. Пропуск.")
        
        # Добавляем название нового столбца
        ws.cell(row=1, column=new_col_index).value = f"{col}_Color"
    
    # Сохраняем изменения в новый файл
    wb.save(output_file)
    print(f"Файл с добавленными цветами сохранен как '{output_file}'.")


def fill_color_columns_to_excel1(input_file, output_file, columns_with_hex):
    # Читаем Excel файл в DataFrame
    df = pd.read_excel(input_file)
    
    # Загружаем файл для работы с openpyxl
    wb = load_workbook(input_file)
    ws = wb.active

    # Проходим по указанным столбцам с HEX-кодами
    for col in columns_with_hex:
        if col not in df.columns:
            print(f"Столбец '{col}' не найден в таблице. Пропуск.")
            continue

        # Определяем индекс текущего столбца
        col_index = df.columns.get_loc(col) + 1  # Индекс столбца в openpyxl (нумерация с 1)
        new_col_index = col_index + 1  #  столбец  для закрашивания
        # Вставляем новый столбец в Excel
        # ws.insert_cols(new_col_index)

        # Копируем данные из исходного столбца в новый
        # for row in range(2, ws.max_row + 1):  # Пропускаем заголовок
        #     ws.cell(row=row, column=new_col_index).value = ws.cell(row=row, column=col_index).value

        # Закрашиваем ячейки нового столбца
        for i, hex_code in enumerate(df[col]):
            excel_row = i + 2  # Нумерация строк в Excel (с учётом заголовков)
            if isinstance(hex_code, str) and hex_code.startswith('#') and len(hex_code) == 7:
                try:
                    cell = ws.cell(row=excel_row, column=new_col_index)
                    cell.fill = PatternFill(start_color=hex_code[1:], end_color=hex_code[1:], fill_type="solid")
                except Exception as e:
                    print(f"Ошибка обработки цвета '{hex_code}' в строке {excel_row}: {e}")
            else:
                print(f"Некорректный HEX-код '{hex_code}' в строке {excel_row}. Пропуск.")
        
        # Добавляем название нового столбца
        # ws.cell(row=1, column=new_col_index).value = f"{col}_Color"
    
    # Сохраняем изменения в новый файл
    wb.save(output_file)
    print(f"Файл с добавленными цветами сохранен как '{output_file}'.")

# def fill_color_columns_to_excel(output_file, df, columns_with_hex):
#     """
#     Заполняет цветами ячейки в Excel на основе HEX-кодов, содержащихся в указанных столбцах.

#     :param output_file: str, путь к выходному Excel-файлу
#     :param df: pd.DataFrame, датафрейм с данными
#     :param columns_with_hex: list, список столбцов с HEX-кодами цветов
#     """
#     # Сохраняем DataFrame в Excel
#     df.to_excel(output_file, index=False, engine='openpyxl')

#     # Открываем Excel-файл для модификации
#     wb = load_workbook(output_file)
#     ws = wb.active

#     for col in columns_with_hex:
#         if col in df.columns:
#             col_index = df.columns.get_loc(col) + 1  # Индекс столбца с HEX-кодами (1-based для Excel)
#             empty_col_index = col_index + 1  # Индекс пустого столбца для раскрашивания

#             for row_index, hex_code in enumerate(df[col], start=2):  # Начинаем с 2, т.к. 1 строка - заголовки
#                 if pd.notna(hex_code):
#                     try:
#                         # Устанавливаем цвет ячейки в пустом столбце
#                         fill = PatternFill(start_color=hex_code.lstrip('#'), end_color=hex_code.lstrip('#'), fill_type="solid")
#                         ws.cell(row=row_index, column=empty_col_index).fill = fill
#                     except Exception as e:
#                         print(f"Ошибка при применении цвета {hex_code} в строке {row_index}: {e}")

#     # Сохраняем изменения в Excel-файле
#     wb.save(output_file)
#     print(f"Цвета успешно применены и сохранены в файл: {output_file}")
"""
Чтобы интегрировать объект ExcelWriter в функцию fill_color_columns_to_excel, необходимо перестроить её логику.
 Вместо сохранения и повторного открытия файла с использованием load_workbook,
 можно сразу работать с объектом ExcelWriter для внесения изменений в лист Excel.
"""    
# модифицированный fill_color_columns_to_excel  с объектом Excel - writer

def fill_color_columns_to_excel(writer, df, columns_with_hex, sheet_name):
    """
    Раскрашивает столбцы с указанными цветами и сохраняет в указанный лист через writer.
    
    writer: объект ExcelWriter
    df: DataFrame для записи в Excel
    columns_with_hex: словарь {имя столбца: HEX-цвет}, которые нужно закрасить
    columns_with_hex: список столбцов, содержащих HEX-коды для закраск
    sheet_name: название листа для записи
    """
    # Сохраняем DataFrame на указанный лист через writer
    df.to_excel(writer, index=False, sheet_name=sheet_name)

    # Получаем объект листа Excel через writer
    ws = writer.sheets[sheet_name]

    # Применяем стили к нужным столбцам
    # for col_name, hex_color in columns_with_hex.items():
    #     # Найдём индекс столбца
    #     if col_name in df.columns:
    #         col_index = df.columns.get_loc(col_name) + 1  # Excel использует 1-индексацию
    #         fill = PatternFill(start_color=hex_color.lstrip("#"), end_color=hex_color.lstrip("#"), fill_type="solid")

    #         # Применяем заливку к каждой ячейке столбца
    #         for row in range(2, len(df) + 2):  # С учётом заголовка
    #             ws.cell(row=row, column=col_index).fill = fill
    for col in columns_with_hex:
        if col in df.columns:
            col_index = df.columns.get_loc(col) + 1  # Индекс столбца с HEX-кодами (1-based для Excel)
            empty_col_index = col_index + 1  # Индекс пустого столбца для раскрашивания

            for row_index, hex_code in enumerate(df[col], start=2):  # Начинаем с 2, т.к. 1 строка - заголовки
                if pd.notna(hex_code):
                    try:
                        # Устанавливаем цвет ячейки в пустом столбце
                        fill = PatternFill(start_color=hex_code.lstrip('#'), end_color=hex_code.lstrip('#'), fill_type="solid")
                        ws.cell(row=row_index, column=empty_col_index).fill = fill
                    except Exception as e:
                        print(f"Ошибка при применении цвета {hex_code} в строке {row_index}: {e}")

# def form_float(df,columns_list,number):
    
#     df[columns_list]= df[columns_list].round(number)
#     return df

def form_float(df, columns_list, number):
    """
    Форматирует вещественные числа в указанных столбцах DataFrame с заданной точностью.

    df: DataFrame
    columns_list: список имен столбцов, которые нужно форматировать
    number: количество знаков после запятой
    """
    df[columns_list] = df[columns_list].astype(float)
    for col in columns_list:
        if col in df.columns:  # Проверяем, что столбец существует
            if pd.api.types.is_numeric_dtype(df[col]):  # Проверяем, что столбец числовой
                df[col] = df[col].round(number)
            else:
                print(f"Столбец '{col}' не числовой и был пропущен.")
        else:
            print(f"Столбец '{col}' отсутствует в DataFrame.")
    return df


# Пример использования
columns_with_hex = ['col_cir', 'col_sq']
columns_list = ['Bright_PIL', 'Bright_Col', 'Bright_Sq', 'Bright_Sc']
decimal_places = [1,1,1,1]
current_date = datetime.now().strftime("%d_%m")
yesterday_date = datetime.now() - timedelta(days=1)
prev_d = yesterday_date.strftime("%d_%m")
output_dir = os.path.join(os.path.dirname(__file__), 'Data')
input_file = os.path.join(output_dir, f"brightness_analysis_{current_date}.xlsx")
# input_file = os.path.join(output_dir, f"brightness_analysis_{prev_d}.xlsx")
# input_file = os.path.join(output_dir, f"brightness_analysis_17_01.xlsx")
out_file_emp = os.path.join(output_dir, f"bright_analysis_{current_date}_emp.xlsx")
output_file = os.path.join(output_dir, f"bright_analysis_{current_date}_color.xlsx")
# df = pd.read_excel(input_file)
# df = insert_empty_columns(df,columns_with_hex)
with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
    # Обрабатываем каждый лист
    # sheet_name = "Original Order"
    all_sheets = pd.read_excel(input_file, sheet_name=None)
    for sheet_name, df in all_sheets.items():
        if sheet_name == "Original Order":
            pd.options.display.float_format = "{:.1f}".format
            df = insert_empty_columns(df,columns_with_hex)
             # Сначала записываем DataFrame в файл
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            format_dataframe_columns_to_excel(writer,sheet_name, df, columns_list, decimal_places)
            fill_color_columns_to_excel(writer, df, columns_with_hex, sheet_name)
            # print(df1.columns)
            # print(df1)
            # df1 = form_float(df1, columns_list, 1)
        else:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

pd.reset_option("display.float_format")
print(f"Файл успешно сохранён: {output_file}")

# df.to_excel(out_file_emp, index = False)
# print(f"Файл c добавленными столбцами успешно сохранён: {out_file_emp}")

# fill_color_columns_to_excel(output_file, df, columns_with_hex)
# fill_color_columns_to_excel(
#     input_file=input_file,
#     output_file=output_file,
#     columns_with_hex=columns_with_hex
# )
"""
Если параметр sheet_name не указан в функции pd.read_excel, он по умолчанию принимает значение 0.
Это означает, что pandas загрузит только первый лист из файла Excel.

Поведение sheet_name в зависимости от значения:
Если sheet_name=None:

Функция загружает все листы из файла.
Результат возвращается в виде словаря, где ключи — это имена листов, а значения — соответствующие DataFrame
 (данные каждого листа).

all_sheets = pd.read_excel(input_file, sheet_name=None)
# all_sheets -> {'Sheet1': DataFrame1, 'Sheet2': DataFrame2, ...}
Если sheet_name не указан (или sheet_name=0):

Загружается только первый лист (по порядку в книге Excel).
Возвращается один DataFrame.

first_sheet = pd.read_excel(input_file)
# first_sheet -> DataFrame из первого листа
Если sheet_name задан как строка:

Загружается лист с указанным именем.
Возвращается один DataFrame.

specific_sheet = pd.read_excel(input_file, sheet_name="Sheet2")
# specific_sheet -> DataFrame с данными из листа "Sheet2"
Если sheet_name задан как список:

Загружаются только листы из списка.
Результат возвращается как словарь.

some_sheets = pd.read_excel(input_file, sheet_name=["Sheet1", "Sheet3"])
# some_sheets -> {'Sheet1': DataFrame1, 'Sheet3': DataFrame3}
Итог
Если параметр sheet_name=None не указан, загрузится только первый лист, а не весь файл.
Чтобы гарантированно работать со всеми листами, нужно явно указывать sheet_name=None.

"""