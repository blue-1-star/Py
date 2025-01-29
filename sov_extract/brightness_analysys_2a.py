import os
import rawpy
import openpyxl
from openpyxl.styles import Font
from PIL import Image, ImageStat, ImageDraw
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
from datetime import datetime
from openpyxl import load_workbook
import tempfile
import brightness_analysys_0a
import calc_brightness_plt_1, calc_brightness_plt_1a
from calc_brightness_plt_1a import shorten_filename, shorten_filename_list
from brightness_analysys_1b import get_image_files_with_metadata, get_elements_by_index, extract_number_from_filename
import re

# Функция обработки изображений (включая ORF)
def process_image(img_path):
    if img_path.lower().endswith('.orf'):
        with rawpy.imread(img_path) as raw:
            rgb = raw.postprocess()
        img = Image.fromarray(rgb)
    else:
        img = Image.open(img_path)
    return img

# Функция выделения области изображения
def crop_image(image, shape='square', size=100):
    """
    Вырезает заданную область (квадрат или окружность) из изображения.
    """
    width, height = image.size
    center_x, center_y = width // 2, height // 2

    if shape == 'square':
        left = max(center_x - size // 2, 0)
        upper = max(center_y - size // 2, 0)
        right = min(center_x + size // 2, width)
        lower = min(center_y + size // 2, height)
        return image.crop((left, upper, right, lower))

    elif shape == 'circle':
        mask = Image.new("L", image.size, 0)
        draw = ImageDraw.Draw(mask)
        draw.ellipse(
            (center_x - size // 2, center_y - size // 2, 
             center_x + size // 2, center_y + size // 2),
            fill=255
        )
        result = Image.composite(image, Image.new("RGB", image.size, (0, 0, 0)), mask)
        bbox = mask.getbbox()
        return result.crop(bbox)

# Функции расчёта яркости
def calculate_brightness_pil(image_path, lower_threshold=0, upper_threshold=255):
    # img = process_image(image_path)  # Завантажуємо та обробляємо зображення
    img = process_image(image_path) 
    # image = Image.open(image_path)
    grayscale_img = img.convert('L')
    pixel_values = np.array(grayscale_img)
    filtered_pixels = pixel_values[(pixel_values >= lower_threshold) & (pixel_values <= upper_threshold)]
    mean_brightness = np.mean(filtered_pixels) if len(filtered_pixels) > 0 else 0
    return {
        'mean_brightness': mean_brightness,
        'used_pixels' : len(filtered_pixels),
        'total_pixels': pixel_values.size
    }

def calculate_brightness_color(image_path, lower_threshold=0, upper_threshold=255):
    img = process_image(image_path) 
    # img = Image.open(image_path)
    pixel_values = np.array(img)
    brightness_values = 0.299 * pixel_values[:, :, 0] + 0.587 * pixel_values[:, :, 1] + 0.114 * pixel_values[:, :, 2]
    filtered_pixels = brightness_values[(brightness_values >= lower_threshold) & (brightness_values <= upper_threshold)]
    mean_brightness = np.mean(filtered_pixels) if len(filtered_pixels) > 0 else 0
    return {
        'mean_brightness': mean_brightness,
        'total_pixels': brightness_values.size
    }

# def calculate_brightness_with_area(image_path, shape='square', size=100, lower_threshold=0, upper_threshold=255):
#     img = process_image(image_path) 
#     # img = Image.open(image_path)
#     cropped_img = crop_image(img, shape=shape, size=size)
#     return calculate_brightness_pil(cropped_img, lower_threshold, upper_threshold)


def calculate_brightness_with_area(image_path, shape='square', size=100, lower_threshold=0, upper_threshold=255):
    img = process_image(image_path) 
    cropped_img = crop_image(img, shape=shape, size=size)
    
    with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as temp_file:
        cropped_img.save(temp_file.name)
        temp_path = temp_file.name
    
    try:
        return calculate_brightness_pil(temp_path, lower_threshold, upper_threshold)
    finally:
        os.remove(temp_path)

def calculate_color_with_area(image_path, shape='square', size=100, lower_threshold=0, upper_threshold=255):
    """Calculates detailed color analysis for the selected area of the image."""
    img = process_image(image_path)
    cropped_img = crop_image(img, shape=shape, size=size)

    # Convert cropped image to numpy array
    pixel_values = np.array(cropped_img)

    # Calculate brightness values
    brightness_values = 0.299 * pixel_values[:, :, 0] + 0.587 * pixel_values[:, :, 1] + 0.114 * pixel_values[:, :, 2]

    # Filter pixels based on brightness thresholds
    mask = (brightness_values >= lower_threshold) & (brightness_values <= upper_threshold)
    filtered_pixels = pixel_values[mask]

    # Calculate average color if there are filtered pixels
    if len(filtered_pixels) > 0:
        avg_r = np.average(filtered_pixels[:, 0], weights=(0.299 * filtered_pixels[:, 0]))
        avg_g = np.average(filtered_pixels[:, 1], weights=(0.587 * filtered_pixels[:, 1]))
        avg_b = np.average(filtered_pixels[:, 2], weights=(0.114 * filtered_pixels[:, 2]))
        avg_r = int(avg_r)
        avg_g = int(avg_g)
        avg_b = int(avg_b)
        # return int(avg_r), int(avg_g), int(avg_b)
        # avg_color_hex = f'#{avg_color[0]:02x}{avg_color[1]:02x}{avg_color[2]:02x}'
        avg_color_hex = f'#{avg_r:02x}{avg_g:02x}{avg_b:02x}'
        return avg_color_hex

    
def group_and_extract(df):
    """
    Функция, которая принимает датафрейм, заменяет содержимое столбца 'Filename' на число, 
    извлеченное из имени файла, и группирует по столбцам ['Filename', 'Substrate', 'Camera'].

    Args:
        df (pd.DataFrame): Исходный датафрейм с колонками 'Filename', 'Substrate', 'Camera'.

    Returns:
        pd.DataFrame: Сгруппированный датафрейм.
    """
    # Извлечение чисел из столбца 'Filename'
    df['Filename'] = df['Filename'].apply(lambda x: int(re.search(r'\d+', x).group()) if re.search(r'\d+', x) else None)

    # Группировка по столбцам
    grouped_df = df.groupby(['Filename', 'Substrate', 'Camera']).sum().reset_index()

    return grouped_df

def sort_dataframe(df):
    """
    Функция сортирует датафрейм по указанным столбцам: сначала по ['Filename', 'Substrate', 'Camera'] 
    в заданном порядке, а затем внутри подгрупп сортирует по Bright_P.

    Args:
        df (pd.DataFrame): Исходный датафрейм.

    Returns:
        pd.DataFrame: Отсортированный датафрейм.
    """
    # Сортировка по столбцам ['Filename', 'Substrate', 'Camera'], затем по 'Bright_P'
    # sorted_df = df.sort_values(by=['Substrate','Camera',  'Filename',  'Bright_P'], 
    #                            ascending=[True, True, True, False])
    # sorted_df = df.sort_values(by=['Substrate', 'Filename'], ascending=[True, True])
    sorted_df = df.sort_values(by=['Filename','Substrate', 'Camera'], ascending=[True, True, True])
    return sorted_df


def sort_and_rank(df):
    """
    Сортирует данные по столбцам 'Substrate', 'Camera', 'Filename', 'Bright_P', а затем добавляет
    столбец 'rank', где вычисляются ранги Bright_P по убыванию внутри каждой группы ['Substrate', 'Camera'].
    """
    # Сортировка по заданным столбцам
    sorted_df = df.sort_values(by=['Substrate', 'Camera', 'Filename', 'Bright_P'], 
                               ascending=[True, True, True, False])

    # Добавление рангов по убыванию Bright_P внутри подгрупп ['Substrate', 'Camera']
    sorted_df['rank'] = sorted_df.groupby(['Substrate', 'Camera'])['Bright_P'] \
                                 .rank(method='dense', ascending=False).astype(int)
     # Перемещение столбца 'rank' сразу после 'Bright_P'
    bright_p_index = sorted_df.columns.get_loc('Bright_P')  # Индекс столбца 'Bright_P'
    rank_column = sorted_df.pop('rank')  # Удаляем столбец 'rank' временно
    sorted_df.insert(bright_p_index + 1, 'rank', rank_column)  # Вставляем 'rank' после 'Bright_P'                                 
    return sorted_df

# Функция для создания DataFrame с результатами
def calculate_brightness_dataframe(image_dir, lower_threshold, size):
    image_files = [f for f in os.listdir(image_dir) if f.lower().endswith(('.orf', '.jpg', '.jpeg' ))]
    # image_files = [f for f in os.listdir(image_dir) if f.lower().endswith(('.orf', '.jpg', '.jpeg','.png' ))]
    # image_files_tupl_meta = get_image_files_with_metadata(image_dir)
    # image_files = get_elements_by_index(image_files_tupl_meta,0)  # filename 
    # subst= get_elements_by_index(image_files_tupl_meta,1)    #  substrat
    # cam = get_elements_by_index(image_files_tupl_meta,2)  # camera

    results = []
    low_list, upper_list = [20,40], [235,215]
    for idx, img_file in enumerate(image_files):

        img_path = os.path.join(image_dir, img_file)
        img = process_image(img_path)
        brightness_pil = calculate_brightness_pil(img_path, lower_threshold)
        # brightness_color = calculate_brightness_color(img_path, lower_threshold)
        # brightness_square = calculate_brightness_with_area(img_path, shape='square', size, lower_threshold=lower_threshold)
        # brightness_square = calculate_brightness_with_area(img_path, 'square', size, lower_threshold=lower_threshold)
        # brightness_circle = calculate_brightness_with_area(img_path, 'circle', size, lower_threshold=lower_threshold)
        # avg_color_circle = calculate_color_with_area(img_path, 'circle', size, lower_threshold=lower_threshold)
        # avg_col_square = calculate_color_with_area(img_path, 'square', size, lower_threshold=lower_threshold)

        # Основной словарь результатов
        result_row = {
            "Filename": img_file,
            "Substrate": subst[idx],
            "Camera": cam[idx],
            "Bright_P": brightness_pil['mean_brightness'],
            "used pixels %": brightness_pil['used_pixels'] / brightness_pil['total_pixels'],
        }
        # Если списки не пустые, выполняем дополнительные расчеты
        if low_list and upper_list:
            for lower_threshold, upper_threshold in zip(low_list, upper_list):
                # Выполняем расчет с заданными порогами
                brightness_pil = calculate_brightness_pil(img_path, lower_threshold, upper_threshold)

                # Добавляем результаты в словарь с нужными именами столбцов
                l_column = f"L{lower_threshold}"
                u_column = f"U{upper_threshold}"
                result_row[l_column] = brightness_pil['mean_brightness']
                result_row[f"us_pix{l_column}"] = brightness_pil['used_pixels'] / brightness_pil['total_pixels']
                result_row[u_column] = brightness_pil['mean_brightness']
                result_row[f"us_pix{u_column}"] = brightness_pil['used_pixels'] / brightness_pil['total_pixels']

            results.append(result_row)

        # results.append({
        #     # "Filename": extract_number_from_filename(img_file),
        #     "Filename": img_file,
        #     "Substrate": subst[idx],
        #     "Camera": cam[idx],
        #     # "Format": file_format,
        #     "Bright_P": brightness_pil['mean_brightness'],
        #     "used pixels %" : brightness_pil['used_pixels']/brightness_pil['total_pixels'],
        #     # "Brightness_Color": brightness_color['mean_brightness'],
        #     # "Brightness_Circle": brightness_circle['mean_brightness'],
        #     # "col_cir": avg_color_circle,
        #     # "col_sq": avg_col_square,

        # })


    df = pd.DataFrame(results)

    # df['Rank_in_Group_PIL'] = df.groupby('Format')['Bright_P'].rank(ascending=False).astype(int)
    # df = df.sort_values(
    # by=['Substrate','Camera', 'Filename'], 
    # key=lambda col: col.str.extract(r'(\d+)')[0].astype(int) if col.name == 'Filename' else col)   
    # Добавление рангов
    # df['Rank_in_Group_PIL'] = df.groupby('Format')['Brightness_PIL'].rank(ascending=False).astype(int)
    # df['Rank_Global_PIL'] = df['Brightness_PIL'].rank(ascending=False).astype(int)
    # df['Rank_in_Group_Color'] = df.groupby('Format')['Brightness_Color'].rank(ascending=False).astype(int)
    # df['Rank_Global_Color'] = df['Brightness_Color'].rank(ascending=False).astype(int)
    # df['Format'] = pd.Categorical(df['Format'], categories=['ORF', 'JPEG'], ordered=True)
    # df = df.sort_values(
    # by=['Format', 'Filename'], 
    # key=lambda col: col.str.extract(r'(\d+)')[0].astype(int) if col.name == 'Filename' else col)    
    return df

# Функция сохранения в Excel
# def save_brightness_excel(df, output_file):
#     with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
#         df.sort_values(['Filename'], inplace=True)
#         df.to_excel(writer, index=False, sheet_name='Original Order')
#         df.sort_values(['Format', 'Rank_in_Group_PIL'], inplace=True)
#         df.to_excel(writer, index=False, sheet_name='Sorted by Brightness_PIL')
#         df.sort_values(['Format', 'Rank_in_Group_Color'], inplace=True)
#         df.to_excel(writer, index=False, sheet_name='Sorted by Brightness_Color')
def auto_adjust_column_width(output_file):
    workbook = load_workbook(output_file)

    # Перебираємо всі аркуші
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        
        for col in worksheet.columns:
            max_length = 0
            col_letter = col[0].column_letter  # Отримуємо букву стовпця
            
            # Перебираємо всі комірки в стовпці
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            # Встановлюємо ширину стовпця (додаємо запас у 2 символи)
            adjusted_width = max_length + 2
            worksheet.column_dimensions[col_letter].width = adjusted_width

    # Зберігаємо відкоригований файл
    workbook.save(output_file)

def save_brightness_excel(df, output_file, lower_threshold):
    # Створюємо Excel-файл з трьома листами
    summary = { 
        'date'  : datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'files' : len(df),
        'lower_threshold' : lower_threshold
    }
    summary_df = pd.DataFrame([summary])

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Лист 1: Оригінальний порядок
        # df(['Filename'], inplace=True)
        df['Format'] = pd.Categorical(df['Format'], categories=['ORF', 'JPEG'], ordered=True)
        df1 = df.sort_values(
        by=['Format', 'Filename'], 
        key=lambda col: col.str.extract(r'(\d+)')[0].astype(int) if col.name == 'Filename' else col)        
        df1 = df1.drop(columns=['Format'])
        rename_map = col_ren_bright_excel()
        df1.rename(columns=rename_map, inplace=True) 
        # df1 = df1.rename(columns={'Brightness_PIL': 'Bright_PIL'})
        sheet_name = 'Original Order'
        df1.to_excel(writer, index=False, sheet_name=sheet_name)
        format_dataframe_columns_to_excel(writer,sheet_name, df1, ['Bright_PIL', 'Bright_Col', 'Bright_Sq', 'Bright_Sc'],\
        [1, 1, 1, 1, 1])
        summary_df.to_excel(writer, sheet_name='Original Order', index=False, startrow=len(df) + 2)

        # Лист 2: Сортовані по яскравості всередині груп (PIL)
        df.sort_values(['Format', 'Rank_in_Group_PIL'], inplace=True)
        df.to_excel(writer, index=False, sheet_name='Sorted by Brightness_PIL')
        summary_df.to_excel(writer, sheet_name='Sorted by Brightness_PIL', index=False, startrow=len(df) + 2)

        # Лист 3: Сортовані по яскравості всередині груп (Color)
        df.sort_values(['Format', 'Rank_in_Group_Color'], inplace=True)
        df.to_excel(writer, index=False, sheet_name='Sorted by Brightness_Color')
        summary_df.to_excel(writer, sheet_name='Sorted by Brightness_Color', index=False, startrow=len(df) + 2)

        # Лист 4: Глобальне сортування по яскравості (PIL)
        df.sort_values(['Rank_Global_PIL'], inplace=True)
        df.to_excel(writer, index=False, sheet_name='Global Sorted by PIL')
        summary_df.to_excel(writer, sheet_name='Global Sorted by PIL', index=False, startrow=len(df) + 2)
        
        # Лист 5: Глобальне сортування по яскравості (Color)
        df.sort_values(['Rank_Global_Color'], inplace=True)
        df.to_excel(writer, index=False, sheet_name='Global Sorted by Color')
        summary_df.to_excel(writer, sheet_name='Global Sorted by Color', index=False, startrow=len(df) + 2)

    # print(f"Excel-файл '{output_file}' створено успішно!")
# графика

def col_ren_bright_excel():
    return {
        'Brightness_PIL': 'Bright_PIL',
        'Brightness_Color': 'Bright_Col',
        'Brightness_Square': 'Bright_Sq',
        'Brightness_Circle': 'Bright_Sc',        
        'avg_color_circle': 'Col_Sc',
        'avg_color_square': 'Col_sq',        
        # Добавьте другие столбцы
    }



def format_dataframe_columns_to_excel(writer, target_sheet_name, df, columns, decimal_places):
    """
    Применяет форматирование столбцов в указанном листе Excel.
    
    writer: объект ExcelWriter
    target_sheet_name: имя листа
    df: DataFrame
    columns: список столбцов для форматирования
    decimal_places: количество знаков после запятой
    """
    # Получаем лист Excel
    sheet = writer.sheets[target_sheet_name]
    for column, decimals in zip(columns, decimal_places):
        if column in df.columns and df[column].dtype in ['float64', 'int64']:
            number_format = f'0.{"0" * decimals}'
            col_idx = df.columns.get_loc(column) + 1  # Индекс столбца (Excel 1-based)
            for row_idx in range(2, len(df) + 2):  # Excel строки начинаются с 1, учитываем заголовок
                cell = sheet.cell(row=row_idx, column=col_idx)
                cell.number_format = number_format
        else:
            print(f"Столбец {column} не является числовым или отсутствует в DataFrame и не будет отформатирован.")

def create_brightness_plots(df, output_dir=os.path.join(os.path.dirname(__file__), 'Data')):
    # Prepare data
    df = sort_dataframe(df)
    # df['Format'] = pd.Categorical(df['Format'], categories=['ORF', 'JPEG'], ordered=True)
    # df = df.sort_values(
    # by=['Format', 'Filename'], 
    # key=lambda col: col.str.extract(r'(\d+)')[0].astype(int) if col.name == 'Filename' else col)        
    subs_data = df[df['Substrate'] == 'A']
    cam_data = df[df['Camera'] == 'Kam']

    # Ensure output directory exists
    os.makedirs(output_dir, exist_ok=True)

    # Graph 1: Natural order of filenames
    fig, axes = plt.subplots(1, 2, figsize=(12, 6), sharey=True)

    for ax, data, title in zip(axes, [subs_data, cam_data], ['Substrate', 'Camera']):
        x_labels = data['Filename']
        x = range(len(x_labels))
        brightness_values = [
            data['Bright_P'],
            data['Brightness_Color'],
            data['Brightness_Square'],
            data['Brightness_Circle']
        ]
        # Установка ширины столбцов и вычисление позиций меток
        bar_width = 0.2  # Ширина одного столбца
        group_width = bar_width * 4  # Общая ширина группы из 4 столбцов
        x_ticks = [pos + group_width / 2 - bar_width / 2 for pos in x]  # Смещение к середине группы
        # Рисуем график
        for i, brightness in enumerate(brightness_values):
            ax.bar([pos + bar_width * i for pos in x], brightness, width=bar_width, label=brightness.name)
        
        ax.set_xticks(x_ticks)  # Устанавливаем середину группы как положение для меток
        ax.set_xticklabels(x_labels, ha='center')  # Метки в центре и повернуты
        # for i, brightness in enumerate(brightness_values):
        #     ax.bar([pos + 0.2 * i for pos in x], brightness, width=0.2, label=brightness.name)



        # ax.set_xticks([pos + 0.3 for pos in x])
        # ax.set_xticklabels(x_labels, rotation=45, ha='right')
        # ax.set_xticklabels(x_labels,  ha='right')
        ax.set_title(title)
        ax.legend(loc='upper left')

    plt.tight_layout()
    output_path = os.path.join(output_dir, f'brightness_graph_{datetime.now().strftime("%d_%m")}.pdf')
    plt.savefig(output_path)
    plt.show()

    # Graph 2: Sorted by mean brightness
    df['Mean_Brightness'] = (
        df['Brightness_PIL'] + df['Brightness_Color'] + df['Brightness_Square'] + df['Brightness_Circle']
    ) / 4
    sorted_df = df.sort_values(by='Mean_Brightness', ascending=False)

    fig, ax = plt.subplots(figsize=(10, 6))
    x_labels = shorten_filename_list(sorted_df['Filename'])
    x = range(len(x_labels))
    brightness_values = [
        sorted_df['Brightness_PIL'],
        sorted_df['Brightness_Color'],
        sorted_df['Brightness_Square'],
        sorted_df['Brightness_Circle']
    ]
    colors = ['#ff9999', '#66b3ff', '#99ff99', '#ffcc99']
    bar_width = 0.2  # Ширина одного столбца
    group_width = bar_width * 4  # Общая ширина группы из 4 столбцов
    x_ticks = [pos + group_width / 2 - bar_width / 2 for pos in x]  # Смещение к середине группы
    # Рисуем график
    for i, brightness in enumerate(brightness_values):
        ax.bar([pos + bar_width * i for pos in x], brightness, width=bar_width, label=brightness.name)
    ax.set_xticks(x_ticks)  # Устанавливаем середину группы как положение для меток
    ax.set_xticklabels(x_labels, ha='center')  # Метки в центре и повернуты

    # for i, (brightness, color) in enumerate(zip(brightness_values, colors)):
    #     ax.bar([pos + 0.2 * i for pos in x], brightness, width=0.2, color=color, label=brightness.name)

    # ax.set_xticks([pos + 0.3 for pos in x])
    # ax.set_xticklabels(x_labels,  ha='right')
    ax.set_title('Sorted by Mean Brightness')
    ax.legend()

    output_path = os.path.join(output_dir, f'brightness_sorted_graph_{datetime.now().strftime("%d_%m")}.pdf')
    plt.tight_layout()
    plt.savefig(output_path)
    plt.show()

    # Graph 3: Mean brightness for each format
    mean_brightness_orf = orf_data[['Brightness_PIL', 'Brightness_Color', 'Brightness_Square', 'Brightness_Circle']].mean()
    mean_brightness_jpg = jpg_data[['Brightness_PIL', 'Brightness_Color', 'Brightness_Square', 'Brightness_Circle']].mean()

    fig, ax = plt.subplots(figsize=(8, 6))
    x = range(4)
    bar_width = 0.4

    ax.bar([pos for pos in x], mean_brightness_orf, width=bar_width, label='ORF', color='#4caf50')
    ax.bar([pos + bar_width for pos in x], mean_brightness_jpg, width=bar_width, label='JPEG', color='#2196f3')

    ax.set_xticks([pos + bar_width / 2 for pos in x])
    ax.set_xticklabels(mean_brightness_orf.index, rotation=45, ha='right')
    ax.set_title('Mean Brightness by Format')
    ax.legend()

    output_path = os.path.join(output_dir, f'mean_brightness_graph_{datetime.now().strftime("%d_%m")}.pdf')
    plt.tight_layout()
    plt.savefig(output_path)
    plt.show()
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages

def plot_bright(df, output_pdf_name="bright_barcharts.pdf"):
    """
    Создает общий график, где для каждого значения 'Filename' отображаются 4 столбца
    (группы по Substrate и Camera) с высотами из 'Bright_P'.

    Параметры:
    - df: DataFrame, содержащий столбцы ['Filename', 'Substrate', 'Camera', 'Bright_P'].
    - output_pdf_name: Имя выходного PDF-файла.
    """
    # Уникальные значения Filename
    filenames = sorted(df['Filename'].unique())
    substrates = ['A', 'F']
    cameras = ['K', 'C']
    combinations = [f"{s}-{c}" for s in substrates for c in cameras]

    # Подготовка данных для построения
    plot_data = []
    for filename in filenames:
        subset = df[df['Filename'] == filename]
        values = []
        for substrate in substrates:
            for camera in cameras:
                value = subset[(subset['Substrate'] == substrate) & (subset['Camera'] == camera)]['Bright_P']
                values.append(value.values[0] if not value.empty else 0)  # Значение или 0
        plot_data.append(values)

    # Создание столбчатой диаграммы
    x = range(len(filenames))  # Индексы для Filename
    width = 0.2  # Ширина столбиков

    plt.figure(figsize=(12, 8))
    for i, comb in enumerate(combinations):
        bars = [data[i] for data in plot_data]  # Значения для текущей комбинации (A-K, A-C, ...)
        plt.bar([pos + i * width for pos in x], bars, width, label=comb)

    # Настройка осей и легенды
    plt.xticks([pos + 1.5 * width for pos in x], filenames, fontsize=10)
    plt.xlabel("Filename", fontsize=12)
    plt.ylabel("Bright_P", fontsize=12)
    plt.title("Яркость (Bright_P) по Filename с учетом Substrate и Camera", fontsize=14)
    plt.legend(title="Substrate-Camera", fontsize=10)
    plt.grid(axis='y', linestyle='--', alpha=0.7)

    # Сохранение и вывод
    plt.tight_layout()
    plt.savefig(output_pdf_name)
    plt.show()
    pdf_pages.close()

def plot_bright_fixed(df, output_pdf_name="bright_barcharts_fixed.pdf"):
    """
    Создает общий график, где для каждого значения 'Filename' отображаются 4 столбца
    (группы по Substrate и Camera) с высотами из 'Bright_P'.
    Добавлен отладочный вывод для проверки данных.
    
    Параметры:
    - df: DataFrame, содержащий столбцы ['Filename', 'Substrate', 'Camera', 'Bright_P'].
    - output_pdf_name: Имя выходного PDF-файла.
    """
    # Уникальные значения и комбинации
    filenames = sorted(df['Filename'].unique())
    substrates = ['A', 'F']
    cameras = ['Kam', 'Sm']
    combinations = [f"{s}-{c}" for s in substrates for c in cameras]

    # Проверка и подготовка данных
    plot_data = []
    for filename in filenames:
        subset = df[df['Filename'] == filename]
        values = []
        for substrate in substrates:
            for camera in cameras:
                value = subset[(subset['Substrate'] == substrate) & (subset['Camera'] == camera)]['Bright_P']
                if not value.empty:
                    values.append(value.values[0])  # Добавляем значение Bright_P
                else:
                    values.append(0)  # Если данных нет, добавляем 0
        plot_data.append(values)
        print(f"Filename {filename}: {values}")  # Отладочный вывод для проверки

    # Если данные пустые, выводим предупреждение
    if not plot_data:
        print("Данные для построения графика отсутствуют!")
        return

    # Построение столбчатой диаграммы
    x = range(len(filenames))  # Индексы для Filename
    width = 0.2  # Ширина столбиков

    plt.figure(figsize=(12, 8))
    for i, comb in enumerate(combinations):
        bars = [data[i] for data in plot_data]  # Значения для текущей комбинации
        plt.bar([pos + i * width for pos in x], bars, width, label=comb)

    # Настройка осей и легенды
    plt.xticks([pos + 1.5 * width for pos in x], filenames, fontsize=10)
    plt.xlabel("Filename", fontsize=12)
    # plt.ylabel("Bright_P", fontsize=12)
    plt.ylabel("L40", fontsize=12)
    plt.title("Яркость (Bright_P) по Filename с учетом Substrate и Camera", fontsize=14)
    plt.legend(title="Substrate-Camera", fontsize=10)
    plt.grid(axis='y', linestyle='--', alpha=0.7)

    # Сохранение и вывод
    plt.tight_layout()
    plt.savefig(output_pdf_name)
    plt.show()

def plot_bright_debug(df, output_pdf_name="bright_barcharts_debug.pdf"):
    """
    Создает общий график, где для каждого значения 'Filename' отображаются 4 столбца
    (группы по Substrate и Camera) с высотами из 'Bright_P'. Добавлены отладочные выводы.
    
    Параметры:
    - df: DataFrame, содержащий столбцы ['Filename', 'Substrate', 'Camera', 'Bright_P'].
    - output_pdf_name: Имя выходного PDF-файла.
    """
    # Уникальные значения и комбинации
    filenames = sorted(df['Filename'].unique())
    substrates = ['A', 'F']
    cameras = ['K', 'C']
    combinations = [f"{s}-{c}" for s in substrates for c in cameras]

    # Проверка и подготовка данных
    plot_data = []
    for filename in filenames:
        subset = df[df['Filename'] == filename]
        print(f"\nFilename {filename}:")
        print(subset)  # Отладочный вывод текущего подмножества данных

        values = []
        for substrate in substrates:
            for camera in cameras:
                condition = (subset['Substrate'] == substrate) & (subset['Camera'] == camera)
                value = subset[condition]['Bright_P']
                
                # Отладочный вывод
                print(f"  Substrate={substrate}, Camera={camera}, Bright_P: {value.values if not value.empty else 'Empty'}")

                if not value.empty:
                    values.append(value.values[0])  # Добавляем значение Bright_P
                else:
                    values.append(0)  # Если данных нет, добавляем 0

        plot_data.append(values)
        print(f"  Values for Filename {filename}: {values}")  # Проверка собранных значений

    # Если данные пустые, выводим предупреждение
    if not plot_data:
        print("Данные для построения графика отсутствуют!")
        return

    # Построение столбчатой диаграммы
    x = range(len(filenames))  # Индексы для Filename
    width = 0.2  # Ширина столбиков

    plt.figure(figsize=(12, 8))
    for i, comb in enumerate(combinations):
        bars = [data[i] for data in plot_data]  # Значения для текущей комбинации
        plt.bar([pos + i * width for pos in x], bars, width, label=comb)

    # Настройка осей и легенды
    plt.xticks([pos + 1.5 * width for pos in x], filenames, fontsize=10)
    plt.xlabel("Filename", fontsize=12)
    plt.ylabel("Bright_P", fontsize=12)
    plt.title("Яркость (Bright_P) по Filename с учетом Substrate и Camera", fontsize=14)
    plt.legend(title="Substrate-Camera", fontsize=10)
    plt.grid(axis='y', linestyle='--', alpha=0.7)

    # Сохранение и вывод
    plt.tight_layout()
    plt.savefig(output_pdf_name)
    plt.show()


# Вызов функции для отладки
# plot_bright_debug(df)




# Пример использования функции
# Создание примерного датафрейма
data = {
    'Filename': [1, 1, 1, 1, 2, 2, 2, 2],
    'Substrate': ['A', 'A', 'F', 'F', 'A', 'A', 'F', 'F'],
    'Camera': ['K', 'C', 'K', 'C', 'K', 'C', 'K', 'C'],
    'Bright_P': [50, 60, 70, 80, 55, 65, 75, 85]
}
# df = pd.DataFrame(data)

# Вызов функции для примера
# plot_bright(df)



# Основной блок выполнения
if __name__ == "__main__":
    # image_dir = r"G:\My\sov\extract\photo"  # Ваш путь к папке с изображениями
    image_dir = r"G:\My\sov\extract\ORF\AF"  # Ваш путь к папке с изображениями
    current_date = datetime.now().strftime("%d_%m")
    output_dir = os.path.join(os.path.dirname(__file__), 'Data')
    output_file = os.path.join(output_dir, f"brightness_analys_6_{current_date}.xlsx")
    output_pdf = os.path.join(output_dir, f"bright_analys_5_{current_date}.pdf")
    lower_threshold = 0
    size = 200
    # df = calculate_brightness_dataframe(image_dir, lower_threshold, size)
    cache_file = os.path.join(image_dir, "brightness_data_5.csv")

    if os.path.exists(cache_file):
        df = pd.read_csv(cache_file)
        print("Data loaded from cache.")
    else:
        df = calculate_brightness_dataframe(image_dir, lower_threshold, size)
        df.to_csv(cache_file, index=False)
        print("Brightness calculations completed and cached.")


    # save_brightness_excel(df, output_file, lower_threshold)
    df = df.round(1)  # Округляем все числа до 1 знака
    df = group_and_extract(df)
    # df = sort_dataframe(df)
    df = sort_and_rank(df)
    df.to_excel(output_file) 
    auto_adjust_column_width(output_file)
    print(f"Анализ яркости завершён. Результаты сохранены в {output_file}")
    plot_bright_fixed(df,output_pdf_name=output_pdf)
    # plot_bright_debug(df,output_pdf_name=output_pdf)
    
    # plot_bright(df, output_pdf_name=output_pdf)
    # create_brightness_plots(df, output_dir=os.path.join(os.path.dirname(__file__), 'Data'))
