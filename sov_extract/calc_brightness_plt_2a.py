import os
import rawpy
import openpyxl
from openpyxl.styles import Font
from PIL import Image, ImageStat
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook

# Функція обробки зображень (включно з ORF)
def process_image(img_path):
    if img_path.lower().endswith('.orf'):  # Обробка ORF-файлів через rawpy
        with rawpy.imread(img_path) as raw:
            rgb = raw.postprocess()
        img = Image.fromarray(rgb)
    else:
        img = Image.open(img_path)
    return img
from PIL import Image, ImageStat
import numpy as np
import os
import pandas as pd
from datetime import datetime

def calculate_brightness_pil(image_path, lower_threshold=0, upper_threshold=255):
    # img = Image.open(image_path)
    img = process_image(image_path)  # Завантажуємо та обробляємо зображення
    grayscale_img = img.convert('L')
    # pixel_values = np.array(grayscale_img)
    pixel_values = np.array(grayscale_img.getdata()).reshape(grayscale_img.size[::-1])
    filtered_pixels = pixel_values[(pixel_values >= lower_threshold) & (pixel_values <= upper_threshold)]
    mean_brightness = np.mean(filtered_pixels) if len(filtered_pixels) > 0 else 0
    dark_pixels_removed = np.sum(pixel_values < lower_threshold)
    bright_pixels_removed = np.sum(pixel_values > upper_threshold)
    return {
        'mean_brightness': mean_brightness,
        'dark_pixels_removed': dark_pixels_removed,
        'bright_pixels_removed': bright_pixels_removed
    }

def calculate_brightness_color(image_path, lower_threshold=0, upper_threshold=255):
    # img = Image.open(image_path)
    img = process_image(image_path)  # Завантажуємо та обробляємо зображення
    pixel_values = np.array(img)
    brightness_values = 0.299 * pixel_values[:, :, 0] + 0.587 * pixel_values[:, :, 1] + 0.114 * pixel_values[:, :, 2]
    # Фільтруємо пікселі за яскравістю (залишаємо в межах порогів)
    mask = (brightness_values >= lower_threshold) & (brightness_values <= upper_threshold)
    filtered_pixels = pixel_values[mask]
    # filtered_pixels = brightness_values[(brightness_values >= lower_threshold) & (brightness_values <= upper_threshold)]
    # Якщо є пікселі в межах порогів, обчислюємо середній колір
    if len(filtered_pixels) > 0:
        # avg_r = np.average(filtered_pixels[:, 0], weights=(0.299))
        # avg_g = np.average(filtered_pixels[:, 1], weights=(0.587))
        # avg_b = np.average(filtered_pixels[:, 2], weights=(0.114))
        avg_r = np.average(filtered_pixels[:, 0], weights=(0.299 * filtered_pixels[:, 0]))
        avg_g = np.average(filtered_pixels[:, 1], weights=(0.587 * filtered_pixels[:, 1]))
        avg_b = np.average(filtered_pixels[:, 2], weights=(0.114 * filtered_pixels[:, 2]))
        avg_color = (int(avg_r), int(avg_g), int(avg_b))
    else:
        avg_color = (0, 0, 0)  # У випадку, якщо всі пікселі відкинуті
    # mean_brightness = np.mean(filtered_pixels) if len(filtered_pixels) > 0 else 0
    mean_brightness = np.mean(brightness_values[mask]) if np.any(mask) else 0
    dark_pixels_removed = np.sum(brightness_values < lower_threshold)
    bright_pixels_removed = np.sum(brightness_values > upper_threshold)
    # Обчислюємо середній колір зображення (RGB)
    # avg_r = np.mean(pixel_values[:, :, 0])
    # avg_g = np.mean(pixel_values[:, :, 1])
    # avg_b = np.mean(pixel_values[:, :, 2])
    # avg_color = (int(avg_r), int(avg_g), int(avg_b))  # округлюємо до цілих
    return {
        'mean_brightness': mean_brightness,
        'dark_pixels_removed': dark_pixels_removed,
        'bright_pixels_removed': bright_pixels_removed,
        'avg_color': avg_color  # Додаємо середній колір
    }

def calculate_brightness_dataframe(image_dir, lower_threshold=0, upper_threshold=255):
    image_files = [f for f in os.listdir(image_dir) if f.lower().endswith(('.orf', '.jpg', '.jpeg'))]
    results = []
    for img_file in image_files:
        img_path = os.path.join(image_dir, img_file)
        brightness_pil = calculate_brightness_pil(img_path, lower_threshold, upper_threshold)
        brightness_color = calculate_brightness_color(img_path, lower_threshold, upper_threshold)
        avg_color = brightness_color.get('avg_color', (0, 0, 0))
        avg_color_hex = '#{:02x}{:02x}{:02x}'.format(avg_color[0], avg_color[1], avg_color[2])

        # avg_color = brightness_color['avg_color']  # Дістаємо середній колір
        # Формуємо HEX-код кольору
        # avg_color_hex = f'#{avg_color[0]:02x}{avg_color[1]:02x}{avg_color[2]:02x}'

        file_format = "ORF" if img_file.lower().endswith('.orf') else "JPEG"
        
        results.append({
            "Filename": img_file,
            "Format": file_format,
            "Brightness_PIL": brightness_pil['mean_brightness'],
            "Brightness_Color": brightness_color['mean_brightness'],
            "Removed_pix_pil": brightness_pil['dark_pixels_removed'],
            "Removed_pix_col": brightness_color['dark_pixels_removed'],
            "avg_color_hex": avg_color_hex,
            "Lower_Threshold": lower_threshold,
            "Upper_Threshold": upper_threshold,
        })
        
    return pd.DataFrame(results)

def save_brightness_excel(df, output_file):
    summary = {
        'date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'files': len(df),
        'lower_threshold': lower_threshold
    }
    summary_df = pd.DataFrame([summary])
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.sort_values(['Filename'], inplace=True)
        df.to_excel(writer, index=False, sheet_name='Original Order')
        summary_df.to_excel(writer, sheet_name='Original Order', index=False, startrow=len(df) + 2)
        df.sort_values(['Format', 'Brightness_PIL'], inplace=True)
        df.to_excel(writer, index=False, sheet_name='Sorted by Brightness_PIL')
        summary_df.to_excel(writer, sheet_name='Sorted by Brightness_PIL', index=False, startrow=len(df) + 2)
        df.sort_values(['Format', 'Brightness_Color'], inplace=True)
        df.to_excel(writer, index=False, sheet_name='Sorted by Brightness_Color')
        summary_df.to_excel(writer, sheet_name='Sorted by Brightness_Color', index=False, startrow=len(df) + 2)

     # Додаємо лист з діаграмою
        workbook = writer.book
        sheet = workbook.create_sheet('Chart')

        # Запис даних для діаграми
        for r_idx, row in enumerate(df[['Filename', 'Brightness_PIL', 'avg_color_hex']].values, 2):
        # for r_idx, row in enumerate(df.loc[:, ['Filename', 'Brightness_PIL', 'avg_color_hex']].values, 2):
            for c_idx, value in enumerate(row, 1):
                sheet.cell(row=r_idx, column=c_idx, value=value)

        # Створюємо стовпчикову діаграму
        chart = BarChart()
        chart.type = 'col'
        chart.style = 10
        chart.title = 'Brightness Analysis'
        chart.y_axis.title = 'Brightness (PIL)'
        chart.x_axis.title = 'Images'

        # Діапазон даних для діаграми
        data = Reference(sheet, min_col=2, min_row=1, max_row=len(df) + 1)
        categories = Reference(sheet, min_col=1, min_row=2, max_row=len(df) + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        sheet.add_chart(chart, 'E5')

        # Наносимо кольори для кожного стовпчика
        for idx, color in enumerate(df['avg_color_hex'], 0):
            col = chart.series[0]
            point = col.graphicalProperties.solidFill
            point = PatternFill(start_color=color, end_color=color, fill_type='solid')
            col.dPt.append({'idx': idx, 'spPr': {'solidFill': {'srgbClr': color[1:]}}})

    print(f"Excel-файл '{output_file}' створено успішно з діаграмою!")
   
    # print(f"Excel-файл '{output_file}' створено успішно!")
    auto_adjust_column_width(output_file)
def auto_adjust_column_width(output_file):
    workbook = load_workbook(output_file)

    # Перебираємо всі аркуші
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        
        for col in worksheet.columns:
            max_length = 0
            col_letter = col[0].column_letter  # Отримуємо букву стовпця
            
            # Перебираємо всі комірки в стовпці
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            # Встановлюємо ширину стовпця (додаємо запас у 2 символи)
            adjusted_width = max_length + 2
            worksheet.column_dimensions[col_letter].width = adjusted_width

    # Зберігаємо відкоригований файл
    workbook.save(output_file)


# Функція для скорочення імен файлів
def shorten_filename(filename):
    if 'Treatment' in filename:
        return filename.split(' ')[-1].split('.')[0]  # Витягуємо число після Treatment
    elif 'T_' in filename:
        return filename.split('_')[-1].split('.')[0]  # Витягуємо число після T_
    return filename

def shorten_filename_list(filenames):
    return [shorten_filename(f) for f in filenames]

# Функція для створення стовпчастого графіка
def plot_brightness(data, title, ax):
    values = data['Brightness_PIL']
    # files = data['Filename']
    files = shorten_filename_list(data['Filename'])  # Застосовуємо функцію до списку
    # colors = plt.cm.Blues(np.linspace(0.4, 0.8, len(values)))
    colors = data['Avg_Color_Hex']
    ax.bar(files, values, color=colors)
    ax.set_title(title)
    ax.set_ylabel('Brightness (PIL)')
    ax.set_xlabel('Files')
    ax.set_xticks(range(len(files)))
    # ax.set_xticklabels([shorten_filename(f) for f in files], rotation=45)
    ax.set_xticklabels([shorten_filename(f) for f in files])
    # ax.tick_params(axis='x', rotation=45)

# Функція для побудови графіків
def create_brightness_plots(df, output_dir = os.path.join(os.path.dirname(__file__),'Data')):
    orf_data = df[df['Format'] == 'ORF']
    jpg_data = df[df['Format'] == 'JPEG']

    # Перші два графіки
    fig, axs = plt.subplots(1, 2, figsize=(12, 6), constrained_layout=True)
    plot_brightness(orf_data, 'Brightness for ORF Files', axs[0])
    plot_brightness(jpg_data, 'Brightness for JPG Files', axs[1])
    output_path = os.path.join(output_dir, f'brightness_graph_{datetime.now().strftime("%d_%m")}.pdf')
    plt.savefig(output_path,  format="pdf", dpi=300, bbox_inches='tight')  # Рекомендується встановити високий dpi для якості
    plt.show()
    plt.close()
    # Третій графік: порівняння середньої яскравості
    mean_brightness = df.groupby('Format')['Brightness_PIL'].mean()
    formats = mean_brightness.index
    means = mean_brightness.values

    fig, ax = plt.subplots(figsize=(6, 6))
    colors = ['#1f77b4', '#ff7f0e']  # Кольори для ORF і JPG
    ax.bar(formats, means, color=colors)
    ax.set_title('Average Brightness: ORF vs JPG')
    ax.set_ylabel('Average Brightness (PIL)')
    ax.set_xlabel('File Format')
    # output_path = os.path.join(image_dir, "comparision_orf.pdf")
    output_path = os.path.join(output_dir, f'comparision_orf_{datetime.now().strftime("%d_%m")}.pdf')
    plt.savefig(output_path,  format="pdf", dpi=300, bbox_inches='tight')  # Рекомендується встановити високий dpi для якості
    plt.show()
    plt.close()
# Виконання коду
if __name__ == "__main__":
    image_dir = r"G:\My\sov\extract"  # Твій шлях до папки зображень
    output_file = os.path.join(image_dir, "brightness_comparison_comb2.xlsx")
    # Поточна дата у форматі DD_MM
    current_date = datetime.now().strftime("%d_%m")
    # Формуємо шлях до каталогу Data
    # Перевірка чи існує каталог Data, якщо ні - створюємо його
    output_dir = os.path.join(os.path.dirname(__file__), 'Data')
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    output_file = os.path.join(output_dir, f"brightness_comparison_{current_date}.xlsx")
    print(f"Файл буде збережено за адресою: {output_file}")
    lower_threshold = 10
    # Перевірка, чи існує збережений DataFrame
    cache_file = os.path.join(image_dir, "brightness_data.csv")

    if os.path.exists(cache_file):
        df = pd.read_csv(cache_file)
        print("Data loaded from cache.")
    else:
        df = calculate_brightness_dataframe(image_dir, lower_threshold)
        print(df.columns)  # Перевіряємо список стовпців
        print(df.head())   # Перевіряємо перші кілька рядків
        df.to_csv(cache_file, index=False)
        print("Brightness calculations completed and cached.")

    # Збереження у файл Excel
    save_brightness_excel(df, output_file)

    # Створення графіків
    create_brightness_plots(df)
