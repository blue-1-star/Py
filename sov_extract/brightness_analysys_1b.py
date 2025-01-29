import os
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import re


def get_image_files_with_meta_se(image_dir):
    """
    Формирует список файлов в каталоге и возвращает их с метаданными:
    имя файла, субстрат и номер опыта.

    Аргументы:
        image_dir (str): Путь к каталогу с изображениями.

    Возвращает:
        list: Список кортежей (file_name, substrate,  experiment_number)
    """
    supported_extensions = ['.jpg', '.png', '.jpeg', '.bmp','.orf']  # Добавьте сюда нужные расширения
    result = []
    for filename in os.listdir(image_dir):
        if filename.lower().endswith(tuple(supported_extensions)):
            match = re.match(r"([A|F])(\d+)", filename)
            if match:
                substrate, experiment_number = match.groups()
                result.append((filename, substrate, experiment_number))
    return result

# Пример использования:
# image_directory = "путь_к_вашему_каталогу"
# files_with_meta = get_image_files_with_meta_se(image_directory)

# for file_info in files_with_meta:
#     print(file_info)


def get_image_files_with_metadata(image_dir):
    """
    Формирует список файлов в каталоге и возвращает их с метаданными:
    имя файла, субстрат, камера и номер опыта.

    Аргументы:
        image_dir (str): Путь к каталогу с изображениями.

    Возвращает:
        list: Список кортежей (file_name, substrate, camera, experiment_number), где:
              - file_name: Имя файла.
              - substrate: Субстрат ('A' или 'F').
              - camera: Камера ('Kam' или 'Sm').
              - experiment_number: Номер опыта (строка).
    """
    image_files = []

    for file_name in os.listdir(image_dir):
        if file_name.endswith(".png"):
            # Убираем расширение для анализа
            base_name = file_name[:-4]
            
            # Ищем субстрат (A или F в начале имени)
            substrate = None
            if base_name[0].lower() == 'a':
                substrate = 'A'
            elif base_name[0].lower() == 'f':
                substrate = 'F'

            # Ищем номер опыта (первое число в имени файла)
            experiment_number_match = re.search(r'\d+', base_name)
            experiment_number = experiment_number_match.group(0) if experiment_number_match else None

            # Ищем камеру ('k' или 's', рядом с разделителем или в конце)
            camera = None
            if 'k' in base_name.lower():
                camera = 'Kam'
            elif 's' in base_name.lower():
                camera = 'Sm'

            # Добавляем в список только файлы с валидными параметрами
            if substrate and camera and experiment_number:
                image_files.append((file_name, substrate, camera, experiment_number))
            # else:
            #     print(f"Файл {file_name} не соответствует ожидаемому формату (субстрат: {substrate}, камера: {camera}, номер опыта: {experiment_number}).")
    
    return image_files
def get_image_files_with_metadata_se1(image_dir):
    """
    Формирует список файлов в каталоге и возвращает их с метаданными:
    имя файла, субстрат, камера и номер опыта.

    Аргументы:
        image_dir (str): Путь к каталогу с изображениями.

    Возвращает:
        list: Список кортежей (file_name, substrate, camera, experiment_number), где:
              - file_name: Имя файла.
              - substrate: Субстрат ('A' или 'F').
              - experiment_number: Номер опыта (строка).
    """
    image_files = []

    for file_name in os.listdir(image_dir):
        if file_name.endswith(".png"):
            # Убираем расширение для анализа
            base_name = file_name[:-4]
            
            # Ищем субстрат (A или F в начале имени)
            substrate = None
            if base_name[0].lower() == 'a':
                substrate = 'A'
            elif base_name[0].lower() == 'f':
                substrate = 'F'

            # Ищем номер опыта (первое число в имени файла)
            experiment_number_match = re.search(r'\d+', base_name)
            experiment_number = experiment_number_match.group(0) if experiment_number_match else None

            # Ищем камеру ('k' или 's', рядом с разделителем или в конце)
            camera = None
            if 'k' in base_name.lower():
                camera = 'Kam'
            elif 's' in base_name.lower():
                camera = 'Sm'

            # Добавляем в список только файлы с валидными параметрами
            if substrate and camera and experiment_number:
                image_files.append((file_name, substrate, camera, experiment_number))
            # else:
            #     print(f"Файл {file_name} не соответствует ожидаемому формату (субстрат: {substrate}, камера: {camera}, номер опыта: {experiment_number}).")
    
    return image_files

def get_elements_by_index(input_list, i):
    """
    Возвращает список элементов из входного списка кортежей, которые находятся на позиции i.

    :param input_list: Входной список кортежей
    :param i: Индекс элемента в кортеже, который нужно извлечь
    :return: Список элементов, находящихся на позиции i в каждом кортеже
    """
    if not input_list:
        return []

    # Проверка на допустимость индекса i
    if not all(isinstance(item, tuple) for item in input_list):
        raise ValueError("Все элементы входного списка должны быть кортежами.")
    if not all(len(item) > i for item in input_list):
        raise IndexError("Индекс i выходит за пределы длины одного или нескольких кортежей.")

    return [item[i] for item in input_list]

def extract_number_from_filename(filename):
    """
    Извлекает первую числовую строку из имени файла.

    :param filename: Имя файла (строка)
    :return: Числовая строка или None, если числа не найдены
    """
    match = re.search(r'\d+', filename)
    return match.group(0) if match else None

# Функция для сохранения яркости и цветности в Excel
def save_brightness_excel(df_a, df_f, output_file, metadata):
    """
    Сохраняет информацию о яркости и цветности в Excel файл с двумя листами (Sheet1 для A и Sheet2 для F).

    Аргументы:
        df_a (pd.DataFrame): Данные для субстрата A.
        df_f (pd.DataFrame): Данные для субстрата F.
        output_file (str): Путь к выходному Excel файлу.
        metadata (dict): Метаданные для секции Summary (общее описание).
    """
    # Создаем Excel файл
    wb = Workbook()

    # Добавление листа для субстрата A
    ws_a = wb.active
    ws_a.title = "A"

    # Заполнение заголовков для листа A
    headers_a = ["Bright_Kam", "Bright_Sm", "Hex_color_Kam", "Color_Cell_Kam", "Hex_color_Sm", "Color_Cell_Sm"]
    ws_a.append(headers_a)

    # Заполнение данных для листа A
    for _, row in df_a.iterrows():
        row_data = [
            row["Bright_Kam"],
            row["Bright_Sm"],
            row["Hex_color_Kam"],
            row["Hex_color_Kam"],  # Дублируем для цвета ячейки
            row["Hex_color_Sm"],
            row["Hex_color_Sm"]   # Дублируем для цвета ячейки
        ]
        ws_a.append(row_data)

    # Закрашивание ячеек для Color_Cell_Kam и Color_Cell_Sm
    for row in ws_a.iter_rows(min_row=2, max_row=ws_a.max_row, min_col=4, max_col=4):
        for cell in row:
            cell.fill = PatternFill(start_color=cell.value, end_color=cell.value, fill_type="solid")

    for row in ws_a.iter_rows(min_row=2, max_row=ws_a.max_row, min_col=6, max_col=6):
        for cell in row:
            cell.fill = PatternFill(start_color=cell.value, end_color=cell.value, fill_type="solid")

    # Добавление секции Summary для A
    ws_a.append([])
    ws_a.append(["Summary"])
    for key, value in metadata.items():
        ws_a.append([key, value])

    # Добавление листа для субстрата F
    ws_f = wb.create_sheet(title="F")

    # Заполнение заголовков для листа F
    headers_f = ["Bright_Kam", "Bright_Sm", "Hex_color_Kam", "Color_Cell_Kam", "Hex_color_Sm", "Color_Cell_Sm"]
    ws_f.append(headers_f)

    # Заполнение данных для листа F
    for _, row in df_f.iterrows():
        row_data = [
            row["Bright_Kam"],
            row["Bright_Sm"],
            row["Hex_color_Kam"],
            row["Hex_color_Kam"],  # Дублируем для цвета ячейки
            row["Hex_color_Sm"],
            row["Hex_color_Sm"]   # Дублируем для цвета ячейки
        ]
        ws_f.append(row_data)

    # Закрашивание ячеек для Color_Cell_Kam и Color_Cell_Sm
    for row in ws_f.iter_rows(min_row=2, max_row=ws_f.max_row, min_col=4, max_col=4):
        for cell in row:
            cell.fill = PatternFill(start_color=cell.value, end_color=cell.value, fill_type="solid")

    for row in ws_f.iter_rows(min_row=2, max_row=ws_f.max_row, min_col=6, max_col=6):
        for cell in row:
            cell.fill = PatternFill(start_color=cell.value, end_color=cell.value, fill_type="solid")

    # Добавление секции Summary для F
    ws_f.append([])
    ws_f.append(["Summary"])
    for key, value in metadata.items():
        ws_f.append([key, value])

    # Сохраняем файл
    wb.save(output_file)
def calculate_brightness_dataframe(image_dir, lower_threshold, size):
    # image_files = [f for f in os.listdir(image_dir) if f.lower().endswith(('.orf', '.jpg', '.jpeg' ))]
    # image_files = [f for f in os.listdir(image_dir) if f.lower().endswith(('.orf', '.jpg', '.jpeg','.png' ))]
    # image_files_tupl_meta = get_image_files_with_metadata(image_dir)
    image_files = get_elements_by_index(image_files_tupl_meta,0)  # filename 
    subst= get_elements_by_index(image_files_tupl_meta,1)    #  substrat
    cam = get_elements_by_index(image_files_tupl_meta,2)  # camera

    results = []
    low_list, upper_list = [20,40], [235,215]
    for idx, img_file in enumerate(image_files):

        img_path = os.path.join(image_dir, img_file)
        img = process_image(img_path)
        brightness_pil = calculate_brightness_pil(img_path, lower_threshold)
        # brightness_color = calculate_brightness_color(img_path, lower_threshold)
        # brightness_square = calculate_brightness_with_area(img_path, shape='square', size, lower_threshold=lower_threshold)
        # brightness_square = calculate_brightness_with_area(img_path, 'square', size, lower_threshold=lower_threshold)
        # brightness_circle = calculate_brightness_with_area(img_path, 'circle', size, lower_threshold=lower_threshold)
        # avg_color_circle = calculate_color_with_area(img_path, 'circle', size, lower_threshold=lower_threshold)
        # avg_col_square = calculate_color_with_area(img_path, 'square', size, lower_threshold=lower_threshold)

        # Основной словарь результатов
        result_row = {
            "Filename": img_file,
            "Substrate": subst[idx],
            "Camera": cam[idx],
            "Bright_P": brightness_pil['mean_brightness'],
            "used pixels %": brightness_pil['used_pixels'] / brightness_pil['total_pixels'],
        }
        # Если списки не пустые, выполняем дополнительные расчеты
        if low_list and upper_list:
            for lower_threshold, upper_threshold in zip(low_list, upper_list):
                # Выполняем расчет с заданными порогами
                brightness_pil = calculate_brightness_pil(img_path, lower_threshold, upper_threshold)

                # Добавляем результаты в словарь с нужными именами столбцов
                l_column = f"L{lower_threshold}"
                u_column = f"U{upper_threshold}"
                result_row[l_column] = brightness_pil['mean_brightness']
                result_row[f"us_pix{l_column}"] = brightness_pil['used_pixels'] / brightness_pil['total_pixels']
                result_row[u_column] = brightness_pil['mean_brightness']
                result_row[f"us_pix{u_column}"] = brightness_pil['used_pixels'] / brightness_pil['total_pixels']

            results.append(result_row)

        # results.append({
        #     # "Filename": extract_number_from_filename(img_file),
        #     "Filename": img_file,
        #     "Substrate": subst[idx],
        #     "Camera": cam[idx],
        #     # "Format": file_format,
        #     "Bright_P": brightness_pil['mean_brightness'],
        #     "used pixels %" : brightness_pil['used_pixels']/brightness_pil['total_pixels'],
        #     # "Brightness_Color": brightness_color['mean_brightness'],
        #     # "Brightness_Circle": brightness_circle['mean_brightness'],
        #     # "col_cir": avg_color_circle,
        #     # "col_sq": avg_col_square,

        # })




if __name__ == "__main__":
    image_dir = r"G:\My\sov\extract\photo" # Ваш путь к папке с изображениями
    current_date = datetime.now().strftime("%d_%m")
    output_dir = os.path.join(os.path.dirname(__file__), 'Data')
    output_file = os.path.join(output_dir, f"brightness_analysis_{current_date}.xlsx")
    lower_threshold = 0
    size = 2000

    # df = calculate_brightness_dataframe(image_dir, lower_threshold, size)
    cache_file = os.path.join(image_dir, "brightness_data_4.csv")
    # image_files_tuple = get_image_files_with_metadata(image_dir)
    # image_files = get_elements_by_index(image_files_tuple,0)
    # print(image_files_tuple)  # Проверка работы функции
    # print(image_files)  #  список файлов
    # if os.path.exists(cache_file):
    #     df = pd.read_csv(cache_file)
    #     print("Data loaded from cache.")
    # else:
    #     df = calculate_brightness_dataframe(image_dir, lower_threshold, size)
    #     df.to_csv(cache_file, index=False)
    #     print("Brightness calculations completed and cached.")


    # save_brightness_excel(df, output_file, lower_threshold)