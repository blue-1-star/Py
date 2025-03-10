import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime
import os
from brightness_analysys_0a import auto_adjust_column_width 
def add_color_columns_to_excel(input_file, output_file, columns_with_hex):
    # Читаем Excel файл в DataFrame
    df = pd.read_excel(input_file)
    
    # Загружаем файл для работы с openpyxl
    wb = load_workbook(input_file)
    ws = wb.active
    # Сортируем столбцы для обработки с конца, чтобы избежать конфликтов индексов при вставке
    sorted_columns_with_hex = sorted(columns_with_hex, key=lambda col: df.columns.get_loc(col), reverse=True)
    # Проходим по указанным столбцам с HEX-кодами
    for col in columns_with_hex:
        if col not in df.columns:
            print(f"Столбец '{col}' не найден в таблице. Пропуск.")
            continue

        # Индекс текущего столбца
        col_index = df.columns.get_loc(col) + 1  # Индекс столбца для openpyxl (нумерация с 1)
        # new_col_index = ws.max_column + 1  # Новый столбец будет добавлен в конец
        new_col_index = col_index + 1  # Новый столбец будет добавлен сразу после текущего

        # Добавляем новый столбец в Excel
        ws.cell(row=1, column=new_col_index).value = f"{col}_Color"  # Название нового столбца

        # Применяем цвета из столбца к новому столбцу
        for i, hex_code in enumerate(df[col]):
            excel_row = i + 2  # Нумерация строк в Excel (с учётом заголовков)
            if isinstance(hex_code, str) and hex_code.startswith('#') and len(hex_code) == 7:
                try:
                    # Закрашиваем ячейку
                    cell = ws.cell(row=excel_row, column=new_col_index)
                    cell.fill = PatternFill(start_color=hex_code[1:], end_color=hex_code[1:], fill_type="solid")
                except Exception as e:
                    print(f"Ошибка обработки цвета '{hex_code}' в строке {excel_row}: {e}")
            else:
                print(f"Некорректный HEX-код '{hex_code}' в строке {excel_row}. Пропуск.")
    
    # Сохраняем изменения в новый файл
    wb.save(output_file)
    print(f"Файл с добавленными цветами сохранен как '{output_file}'.")
# Пример использования
columns_with_hex = ['avg_color_circle', 'avg_color_square']
current_date = datetime.now().strftime("%d_%m")
output_dir = os.path.join(os.path.dirname(__file__), 'Data')
input_file = os.path.join(output_dir, f"brightness_analysis_{current_date}.xlsx")
output_file = os.path.join(output_dir, f"bright_analysis_{current_date}_color.xlsx")

add_color_columns_to_excel(
    input_file=input_file,
    output_file=output_file,
    columns_with_hex=columns_with_hex
)
auto_adjust_column_width(output_file)
