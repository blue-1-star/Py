import os
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill

import os
import re

def get_image_files_with_metadata(image_dir):
    """
    Формирует список файлов в каталоге и возвращает их с метаданными:
    имя файла, субстрат, камера и номер опыта.

    Аргументы:
        image_dir (str): Путь к каталогу с изображениями.

    Возвращает:
        list: Список кортежей (file_name, substrate, camera, experiment_number), где:
              - file_name: Имя файла.
              - substrate: Субстрат ('A' или 'F').
              - camera: Камера ('Kam' или 'Sm').
              - experiment_number: Номер опыта (строка).
    """
    image_files = []

    for file_name in os.listdir(image_dir):
        if file_name.endswith(".png"):
            # Убираем расширение для анализа
            base_name = file_name[:-4]
            
            # Ищем субстрат (A или F в начале имени)
            substrate = None
            if base_name[0].lower() == 'a':
                substrate = 'A'
            elif base_name[0].lower() == 'f':
                substrate = 'F'

            # Ищем номер опыта (первое число в имени файла)
            experiment_number_match = re.search(r'\d+', base_name)
            experiment_number = experiment_number_match.group(0) if experiment_number_match else None

            # Ищем камеру ('k' или 's', рядом с разделителем или в конце)
            camera = None
            if 'k' in base_name.lower():
                camera = 'Kam'
            elif 's' in base_name.lower():
                camera = 'Sm'

            # Добавляем в список только файлы с валидными параметрами
            if substrate and camera and experiment_number:
                image_files.append((file_name, substrate, camera, experiment_number))
            else:
                print(f"Файл {file_name} не соответствует ожидаемому формату (субстрат: {substrate}, камера: {camera}, номер опыта: {experiment_number}).")
    
    return image_files



# Функция для сохранения яркости и цветности в Excel
def save_brightness_excel(df_a, df_f, output_file, metadata):
    """
    Сохраняет информацию о яркости и цветности в Excel файл с двумя листами (Sheet1 для A и Sheet2 для F).

    Аргументы:
        df_a (pd.DataFrame): Данные для субстрата A.
        df_f (pd.DataFrame): Данные для субстрата F.
        output_file (str): Путь к выходному Excel файлу.
        metadata (dict): Метаданные для секции Summary (общее описание).
    """
    # Создаем Excel файл
    wb = Workbook()

    # Добавление листа для субстрата A
    ws_a = wb.active
    ws_a.title = "A"

    # Заполнение заголовков для листа A
    headers_a = ["Bright_Kam", "Bright_Sm", "Hex_color_Kam", "Color_Cell_Kam", "Hex_color_Sm", "Color_Cell_Sm"]
    ws_a.append(headers_a)

    # Заполнение данных для листа A
    for _, row in df_a.iterrows():
        row_data = [
            row["Bright_Kam"],
            row["Bright_Sm"],
            row["Hex_color_Kam"],
            row["Hex_color_Kam"],  # Дублируем для цвета ячейки
            row["Hex_color_Sm"],
            row["Hex_color_Sm"]   # Дублируем для цвета ячейки
        ]
        ws_a.append(row_data)

    # Закрашивание ячеек для Color_Cell_Kam и Color_Cell_Sm
    for row in ws_a.iter_rows(min_row=2, max_row=ws_a.max_row, min_col=4, max_col=4):
        for cell in row:
            cell.fill = PatternFill(start_color=cell.value, end_color=cell.value, fill_type="solid")

    for row in ws_a.iter_rows(min_row=2, max_row=ws_a.max_row, min_col=6, max_col=6):
        for cell in row:
            cell.fill = PatternFill(start_color=cell.value, end_color=cell.value, fill_type="solid")

    # Добавление секции Summary для A
    ws_a.append([])
    ws_a.append(["Summary"])
    for key, value in metadata.items():
        ws_a.append([key, value])

    # Добавление листа для субстрата F
    ws_f = wb.create_sheet(title="F")

    # Заполнение заголовков для листа F
    headers_f = ["Bright_Kam", "Bright_Sm", "Hex_color_Kam", "Color_Cell_Kam", "Hex_color_Sm", "Color_Cell_Sm"]
    ws_f.append(headers_f)

    # Заполнение данных для листа F
    for _, row in df_f.iterrows():
        row_data = [
            row["Bright_Kam"],
            row["Bright_Sm"],
            row["Hex_color_Kam"],
            row["Hex_color_Kam"],  # Дублируем для цвета ячейки
            row["Hex_color_Sm"],
            row["Hex_color_Sm"]   # Дублируем для цвета ячейки
        ]
        ws_f.append(row_data)

    # Закрашивание ячеек для Color_Cell_Kam и Color_Cell_Sm
    for row in ws_f.iter_rows(min_row=2, max_row=ws_f.max_row, min_col=4, max_col=4):
        for cell in row:
            cell.fill = PatternFill(start_color=cell.value, end_color=cell.value, fill_type="solid")

    for row in ws_f.iter_rows(min_row=2, max_row=ws_f.max_row, min_col=6, max_col=6):
        for cell in row:
            cell.fill = PatternFill(start_color=cell.value, end_color=cell.value, fill_type="solid")

    # Добавление секции Summary для F
    ws_f.append([])
    ws_f.append(["Summary"])
    for key, value in metadata.items():
        ws_f.append([key, value])

    # Сохраняем файл
    wb.save(output_file)

# Пример вызова
# if __name__ == "__main__":
#     # Пример данных
#     data_a = {
#         "Bright_Kam": [120, 130],
#         "Bright_Sm": [110, 125],
#         "Hex_color_Kam": ["FF5733", "33FF57"],
#         "Hex_color_Sm": ["3357FF", "FF33A1"]
#     }
#     df_a = pd.DataFrame(data_a)

#     data_f = {
#         "Bright_Kam": [115, 140],
#         "Bright_Sm": [105, 135],
#         "Hex_color_Kam": ["57FF33", "FF3357"],
#         "Hex_color_Sm": ["FF5733", "5733FF"]
#     }
#     df_f = pd.DataFrame(data_f)

#     metadata = {
#         "Субстрат": "A и F",
#         "Способ вычисления": "квадрат, 100x100, [0, 255]",
#         "Количество пикселей": 10000,
#         "Общее количество пикселей": 1000000,
#         "Дата и время": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
#     }

#     image_dir = "./images"
#     image_files = get_image_files_with_metadata(image_dir)
#     print(image_files)  # Проверка работы функции

#     save_brightness_excel(df_a, df_f, "output.xlsx", metadata)


if __name__ == "__main__":
    image_dir = r"G:\My\sov\extract\photo" # Ваш путь к папке с изображениями
    current_date = datetime.now().strftime("%d_%m")
    output_dir = os.path.join(os.path.dirname(__file__), 'Data')
    output_file = os.path.join(output_dir, f"brightness_analysis_{current_date}.xlsx")
    lower_threshold = 0
    size = 2000

    # df = calculate_brightness_dataframe(image_dir, lower_threshold, size)
    cache_file = os.path.join(image_dir, "brightness_data_4.csv")
    image_files = get_image_files_with_metadata(image_dir)
    print(image_files)  # Проверка работы функции
    # if os.path.exists(cache_file):
    #     df = pd.read_csv(cache_file)
    #     print("Data loaded from cache.")
    # else:
    #     df = calculate_brightness_dataframe(image_dir, lower_threshold, size)
    #     df.to_csv(cache_file, index=False)
    #     print("Brightness calculations completed and cached.")


    # save_brightness_excel(df, output_file, lower_threshold)