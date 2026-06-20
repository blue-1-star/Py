from pathlib import Path
import sys
import sqlite3
import argparse
from datetime import datetime
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

OSBB_ROOT = Path(__file__).resolve().parent
PY_ROOT = OSBB_ROOT.parent

if str(PY_ROOT) not in sys.path:
    sys.path.insert(0, str(PY_ROOT))

from config import paths, USE_TEST_DB


DEFAULT_PERIOD_CODE = "2026-05_2026-06"


def get_db_file():
    return paths.OSBB_TEST_DB_FILE if USE_TEST_DB else paths.OSBB_DB_FILE


def now_ts():
    return datetime.now().strftime("%Y-%m-%d_%H-%M-%S")


def money(value):
    value = float(value or 0)
    return round(value, 2)


def table_exists(cur, table_name):
    cur.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name=?",
        (table_name,),
    )
    return cur.fetchone() is not None


def table_columns(cur, table_name):
    cur.execute(f"PRAGMA table_info({table_name})")
    return [row[1] for row in cur.fetchall()]


def pick_table(cur, *names):
    for name in names:
        if table_exists(cur, name):
            return name
    return None


def apt_sort_key(apt):
    s = "" if apt is None else str(apt)
    return (0, int(s)) if s.isdigit() else (1, s)


def service_label(code):
    return {
        "PARKING_DAY": "Day",
        "PARKING_NIGHT": "Night",
        "BARRIER_CALL": "Шлагбаум",
        "IMPROVEMENT": "Благоустройство",
        None: "",
        "": "",
    }.get(code, code)


def load_charges(cur, period_code):
    charges_table = pick_table(cur, "charges", "service_charges")
    if not charges_table:
        return []

    cols = table_columns(cur, charges_table)
    status_col = "charge_status" if "charge_status" in cols else ("status" if "status" in cols else None)

    sql = f"""
        SELECT
            c.id AS charge_id,
            c.period_code,
            c.apartment_number,
            c.vehicle_id,
            c.service_code,
            c.amount,
            c.quantity,
            c.unit_price,
            COALESCE(v.license_plate_normalized, v.license_plate, '') AS plate,
            COALESCE(v.car_model_normalized, v.car_model, '') AS model,
            COALESCE(v.parking_time, '') AS parking_time
        FROM {charges_table} c
        LEFT JOIN vehicles v ON v.id = c.vehicle_id
        WHERE c.period_code = ?
    """

    if status_col:
        sql += f" AND COALESCE(c.{status_col}, '') <> 'cancelled'"

    sql += " ORDER BY c.apartment_number, c.vehicle_id, c.id"

    cur.execute(sql, (period_code,))
    return [dict(row) for row in cur.fetchall()]


def load_payments(cur, period_code):
    payments_table = pick_table(cur, "payments", "service_payments")
    if not payments_table:
        return []

    cols = table_columns(cur, payments_table)

    def expr(col, default="NULL"):
        return f"p.{col}" if col in cols else default

    cur.execute(f"""
        SELECT
            p.id AS payment_id,
            {expr('payment_date')} AS payment_date,
            p.period_code,
            {expr('apartment_number')} AS apartment_number,
            {expr('vehicle_id')} AS vehicle_id,
            {expr('service_code')} AS service_code,
            p.amount,
            {expr('payment_method')} AS payment_method,
            {expr('source')} AS source,
            {expr('source_ref')} AS source_ref,
            {expr('comment')} AS comment,
            COALESCE(v.license_plate_normalized, v.license_plate, '') AS plate,
            COALESCE(v.car_model_normalized, v.car_model, '') AS model
        FROM {payments_table} p
        LEFT JOIN vehicles v ON v.id = {expr('vehicle_id')}
        WHERE p.period_code = ?
        ORDER BY {expr('apartment_number')}, {expr('payment_date')}, p.id
    """, (period_code,))

    return [dict(row) for row in cur.fetchall()]


def load_allocations(cur):
    allocations_table = pick_table(cur, "payment_allocations", "service_payment_allocations")
    if not allocations_table:
        return []

    cur.execute(f"""
        SELECT
            id AS allocation_id,
            payment_id,
            charge_id,
            amount
        FROM {allocations_table}
        ORDER BY id
    """)
    return [dict(row) for row in cur.fetchall()]


def load_vehicles_without_tariff(cur):
    if not table_exists(cur, "vehicles"):
        return []

    cur.execute("""
        SELECT
            v.id AS vehicle_id,
            a.apartment_number,
            COALESCE(v.license_plate_normalized, v.license_plate, '') AS plate,
            COALESCE(v.car_model_normalized, v.car_model, '') AS model,
            COALESCE(v.parking_time, '') AS parking_time
        FROM vehicles v
        JOIN apartments a ON a.id = v.apartment_id
        WHERE v.parking_time IS NULL OR TRIM(CAST(v.parking_time AS TEXT)) = ''
        ORDER BY a.apartment_number, v.id
    """)
    return [dict(row) for row in cur.fetchall()]


def build_statement(conn, period_code):
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    charges = load_charges(cur, period_code)
    payments = load_payments(cur, period_code)
    allocations = load_allocations(cur)
    no_tariff = load_vehicles_without_tariff(cur)

    allocated_by_charge = defaultdict(float)
    allocated_by_payment = defaultdict(float)

    for a in allocations:
        allocated_by_charge[a["charge_id"]] += float(a["amount"] or 0)
        allocated_by_payment[a["payment_id"]] += float(a["amount"] or 0)

    by_apt = {}

    def apt_item(apt):
        apt = "" if apt is None else str(apt)
        if apt not in by_apt:
            by_apt[apt] = {
                "apt": apt,
                "charged": 0.0,
                "paid": 0.0,
                "allocated": 0.0,
                "unallocated": 0.0,
                "charges": [],
                "payments": [],
            }
        return by_apt[apt]

    for ch in charges:
        apt = str(ch["apartment_number"] or "")
        allocated = allocated_by_charge[ch["charge_id"]]
        amount = float(ch["amount"] or 0)

        row = dict(ch)
        row["allocated"] = allocated
        row["charge_balance"] = amount - allocated

        item = apt_item(apt)
        item["charged"] += amount
        item["allocated"] += allocated
        item["charges"].append(row)

    for p in payments:
        apt = str(p["apartment_number"] or "")
        amount = float(p["amount"] or 0)
        allocated = allocated_by_payment[p["payment_id"]]

        row = dict(p)
        row["allocated"] = allocated
        row["unallocated"] = amount - allocated

        item = apt_item(apt)
        item["paid"] += amount
        item["unallocated"] += amount - allocated
        item["payments"].append(row)

    for item in by_apt.values():
        item["balance"] = item["charged"] - item["paid"]

    return {
        "charges": charges,
        "payments": payments,
        "allocations": allocations,
        "by_apt": by_apt,
        "no_tariff": no_tariff,
    }


def add_table(ws, name):
    if ws.max_row < 2 or ws.max_column < 1:
        return
    ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    tab = Table(displayName=name, ref=ref)
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)


def autosize(ws):
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 10
        for cell in ws[letter]:
            if cell.value is not None:
                max_len = max(max_len, min(len(str(cell.value)), 60))
        ws.column_dimensions[letter].width = max_len + 2


def style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    border = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    autosize(ws)


def write_xlsx(filename, data, period_code):
    wb = Workbook()

    # Summary
    ws = wb.active
    ws.title = "Ведомость"
    ws.append([
        "Квартира",
        "Начислено",
        "Оплачено",
        "Распределено",
        "Нераспределено",
        "Баланс",
        "Статус",
        "Авто/начисления",
        "Оплаты",
    ])

    for apt, item in sorted(data["by_apt"].items(), key=lambda x: apt_sort_key(x[0])):
        balance = money(item["balance"])
        if balance > 0:
            status = "ДОЛГ"
        elif balance < 0:
            status = "ПЕРЕПЛАТА"
        else:
            status = "OK"

        charges_text = []
        for ch in item["charges"]:
            charges_text.append(
                f"{ch.get('plate') or '-'} {ch.get('model') or ''} "
                f"{service_label(ch.get('service_code'))}: {money(ch.get('amount'))}"
            )

        payments_text = []
        for p in item["payments"]:
            payments_text.append(
                f"{p.get('payment_date') or '-'} {money(p.get('amount'))} "
                f"{p.get('plate') or ''} {p.get('source') or ''}"
            )

        ws.append([
            apt or "НЕ ОПРЕДЕЛЕНО",
            money(item["charged"]),
            money(item["paid"]),
            money(item["allocated"]),
            money(item["unallocated"]),
            balance,
            status,
            "\n".join(charges_text),
            "\n".join(payments_text),
        ])

    add_table(ws, "Statement")
    style_sheet(ws)

    # Debtors
    ws = wb.create_sheet("Должники")
    ws.append(["Квартира", "Долг", "Начислено", "Оплачено", "Авто/начисления"])

    debtors = [
        item for item in data["by_apt"].values()
        if item["balance"] > 0.01
    ]
    for item in sorted(debtors, key=lambda x: (-x["balance"], apt_sort_key(x["apt"]))):
        ws.append([
            item["apt"],
            money(item["balance"]),
            money(item["charged"]),
            money(item["paid"]),
            "\n".join(
                f"{ch.get('plate') or '-'} {service_label(ch.get('service_code'))}: {money(ch.get('amount'))}"
                for ch in item["charges"]
            ),
        ])
    add_table(ws, "Debtors")
    style_sheet(ws)

    # Overpayments
    ws = wb.create_sheet("Переплаты")
    ws.append(["Квартира", "Переплата", "Оплачено", "Начислено", "Нераспределено"])
    over = [
        item for item in data["by_apt"].values()
        if item["balance"] < -0.01 or item["unallocated"] > 0.01
    ]
    for item in sorted(over, key=lambda x: (x["balance"], apt_sort_key(x["apt"]))):
        ws.append([
            item["apt"] or "НЕ ОПРЕДЕЛЕНО",
            money(-item["balance"]) if item["balance"] < 0 else 0,
            money(item["paid"]),
            money(item["charged"]),
            money(item["unallocated"]),
        ])
    add_table(ws, "Overpayments")
    style_sheet(ws)

    # Charges details
    ws = wb.create_sheet("Начисления")
    ws.append([
        "charge_id", "Квартира", "vehicle_id", "Номер", "Марка", "Тариф",
        "Услуга", "Начислено", "Распределено", "Остаток"
    ])
    for ch in data["charges"]:
        amount = float(ch["amount"] or 0)
        allocated = 0
        # already not in original rows, compute via statement details
        for item in data["by_apt"].values():
            for ch2 in item["charges"]:
                if ch2["charge_id"] == ch["charge_id"]:
                    allocated = ch2["allocated"]
        ws.append([
            ch["charge_id"],
            ch["apartment_number"],
            ch["vehicle_id"],
            ch["plate"],
            ch["model"],
            ch["parking_time"],
            service_label(ch["service_code"]),
            money(amount),
            money(allocated),
            money(amount - allocated),
        ])
    add_table(ws, "Charges")
    style_sheet(ws)

    # Payments details
    ws = wb.create_sheet("Оплаты")
    ws.append([
        "payment_id", "Дата", "Квартира", "vehicle_id", "Номер", "Марка",
        "Услуга", "Оплачено", "Распределено", "Нераспределено",
        "Источник", "Комментарий"
    ])
    for p in data["payments"]:
        allocated = 0
        unallocated = 0
        for item in data["by_apt"].values():
            for p2 in item["payments"]:
                if p2["payment_id"] == p["payment_id"]:
                    allocated = p2["allocated"]
                    unallocated = p2["unallocated"]
        ws.append([
            p["payment_id"],
            p["payment_date"],
            p["apartment_number"],
            p["vehicle_id"],
            p["plate"],
            p["model"],
            service_label(p["service_code"]),
            money(p["amount"]),
            money(allocated),
            money(unallocated),
            p["source"],
            p["comment"],
        ])
    add_table(ws, "Payments")
    style_sheet(ws)

    # Unallocated payments
    ws = wb.create_sheet("Нераспределенные оплаты")
    ws.append([
        "payment_id", "Дата", "Квартира", "Номер", "Оплачено", "Распределено", "Остаток", "Источник", "Комментарий"
    ])
    for item in sorted(data["by_apt"].values(), key=lambda x: apt_sort_key(x["apt"])):
        for p in item["payments"]:
            if p["unallocated"] > 0.01:
                ws.append([
                    p["payment_id"],
                    p["payment_date"],
                    p["apartment_number"] or "НЕ ОПРЕДЕЛЕНО",
                    p["plate"],
                    money(p["amount"]),
                    money(p["allocated"]),
                    money(p["unallocated"]),
                    p["source"],
                    p["comment"],
                ])
    add_table(ws, "UnallocatedPayments")
    style_sheet(ws)

    # Vehicles without tariff
    ws = wb.create_sheet("Авто без тарифа")
    ws.append(["vehicle_id", "Квартира", "Номер", "Марка", "Тариф"])
    for row in data["no_tariff"]:
        ws.append([
            row["vehicle_id"],
            row["apartment_number"],
            row["plate"],
            row["model"],
            row["parking_time"],
        ])
    add_table(ws, "NoTariff")
    style_sheet(ws)

    # Metadata
    ws = wb.create_sheet("Сводка")
    ws.append(["Параметр", "Значение"])
    ws.append(["Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws.append(["DB", str(get_db_file())])
    ws.append(["MODE", "TEST/WORK" if USE_TEST_DB else "PROD"])
    ws.append(["Period", period_code])
    ws.append(["Apartments in statement", len(data["by_apt"])])
    ws.append(["Charges total", money(sum(float(ch["amount"] or 0) for ch in data["charges"]))])
    ws.append(["Payments total", money(sum(float(p["amount"] or 0) for p in data["payments"]))])
    ws.append(["Vehicles without tariff", len(data["no_tariff"])])
    style_sheet(ws)

    wb.save(filename)


def main():
    parser = argparse.ArgumentParser(
        description="Excel billing statement: charges, payments, balances, work queues."
    )
    parser.add_argument("--period", default=DEFAULT_PERIOD_CODE)

    args = parser.parse_args()

    db_file = get_db_file()
    conn = sqlite3.connect(db_file)
    conn.row_factory = sqlite3.Row

    data = build_statement(conn, args.period)
    conn.close()

    report_dir = paths.OSBB_EXPORTS_DIR / "billing"
    report_dir.mkdir(parents=True, exist_ok=True)

    xlsx_file = report_dir / f"billing_statement_{args.period}_{now_ts()}.xlsx"
    write_xlsx(xlsx_file, data, args.period)

    total_charges = sum(float(ch["amount"] or 0) for ch in data["charges"])
    total_payments = sum(float(p["amount"] or 0) for p in data["payments"])

    print("=" * 70)
    print("BILLING STATEMENT EXCEL")
    print("=" * 70)
    print("DB:", db_file)
    print("MODE:", "TEST/WORK" if USE_TEST_DB else "PROD")
    print("Period:", args.period)
    print("Apartments:", len(data["by_apt"]))
    print("Charges total:", money(total_charges))
    print("Payments total:", money(total_payments))
    print("Vehicles without tariff:", len(data["no_tariff"]))
    print("")
    print("XLSX:")
    print(xlsx_file)
    print("=" * 70)


if __name__ == "__main__":
    main()
