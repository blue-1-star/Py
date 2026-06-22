from pathlib import Path
import sys
import sqlite3
import argparse
from datetime import datetime
from uuid import uuid4

OSBB_ROOT = Path(__file__).resolve().parent
PY_ROOT = OSBB_ROOT.parent

if str(PY_ROOT) not in sys.path:
    sys.path.insert(0, str(PY_ROOT))

from config import paths, USE_TEST_DB
from audit_logger import audit_log


# ---------------------------------------------------------------------
# БАЗОВЫЕ ПРАВИЛА
# ---------------------------------------------------------------------
# base_service_code = тип услуги:
#   PARKING_DAY, PARKING_NIGHT, BARRIER_PHONE, IMPROVEMENT ...
#
# service_item_code = конкретная статья:
#   05_26_Day, 05_26_Night, 01_Imp, 01_ParkEquip ...
#
# Для новых ежемесячных оплат период всегда ОДИН месяц: YYYY-MM.
# Двухмесячный период не является базовым периодом.
# ---------------------------------------------------------------------

SERVICE_META = {
    "PARKING_DAY": {
        "name": "Парковка Day",
        "type": "MONTHLY",
        "category": "PARKING",
        "monthly": 1,
        "fundraising": 0,
        "commercial": 0,
        "access": 0,
        "cash": 1,
    },
    "PARKING_NIGHT": {
        "name": "Парковка Night",
        "type": "MONTHLY",
        "category": "PARKING",
        "monthly": 1,
        "fundraising": 0,
        "commercial": 0,
        "access": 0,
        "cash": 1,
    },
    "BARRIER_PHONE": {
        "name": "Телефонный доступ к шлагбауму",
        "type": "MONTHLY",
        "category": "ACCESS",
        "monthly": 1,
        "fundraising": 0,
        "commercial": 0,
        "access": 1,
        "cash": 1,
    },
    "BARRIER_PHONE_CONNECT": {
        "name": "Подключение / повторное подключение телефонного доступа",
        "type": "ONE_TIME",
        "category": "ACCESS",
        "monthly": 0,
        "fundraising": 0,
        "commercial": 0,
        "access": 1,
        "cash": 1,
    },
    "IMPROVEMENT": {
        "name": "Благоустройство",
        "type": "FUNDRAISING",
        "category": "IMPROVEMENT",
        "monthly": 0,
        "fundraising": 1,
        "commercial": 0,
        "access": 0,
        "cash": 1,
    },
    "PARKING_EQUIPMENT": {
        "name": "Оборудование парковки",
        "type": "FUNDRAISING",
        "category": "EQUIPMENT",
        "monthly": 0,
        "fundraising": 1,
        "commercial": 0,
        "access": 0,
        "cash": 1,
    },
    "BARRIER_REPAIR": {
        "name": "Ремонт / сбор на шлагбаум",
        "type": "FUNDRAISING",
        "category": "BARRIER",
        "monthly": 0,
        "fundraising": 1,
        "commercial": 0,
        "access": 0,
        "cash": 1,
    },
    "GUEST_PARKING": {
        "name": "Гостевая парковка",
        "type": "COMMERCIAL",
        "category": "PARKING",
        "monthly": 0,
        "fundraising": 0,
        "commercial": 1,
        "access": 0,
        "cash": 1,
    },
    "PARK_PLACE_RENT": {
        "name": "Аренда паркоместа",
        "type": "COMMERCIAL",
        "category": "PARKING_ASSET",
        "monthly": 0,
        "fundraising": 0,
        "commercial": 1,
        "access": 0,
        "cash": 1,
    },
    "PARK_PLACE_SALE": {
        "name": "Продажа паркоместа",
        "type": "COMMERCIAL",
        "category": "PARKING_ASSET",
        "monthly": 0,
        "fundraising": 0,
        "commercial": 1,
        "access": 0,
        "cash": 1,
    },
    "HISTORICAL_UNCLASSIFIED": {
        "name": "Историческая нераспределённая операция",
        "type": "HISTORICAL",
        "category": "HISTORICAL",
        "monthly": 0,
        "fundraising": 0,
        "commercial": 0,
        "access": 0,
        "cash": 1,
    },
    "CASH_RECONCILIATION": {
        "name": "Сверка остатка кассы",
        "type": "RECONCILIATION",
        "category": "CASHIER",
        "monthly": 0,
        "fundraising": 0,
        "commercial": 0,
        "access": 0,
        "cash": 1,
    },
    "CASH_TRANSFER": {
        "name": "Передача между кассами",
        "type": "INTERNAL",
        "category": "CASHIER",
        "monthly": 0,
        "fundraising": 0,
        "commercial": 0,
        "access": 0,
        "cash": 0,
    },
}

SERVICE_ALIASES = {
    "P": "PARKING",
    "П": "PARKING",
    "ПАРКОВКА": "PARKING",
    "PARKING": "PARKING",
    "PARKING_DAY": "PARKING_DAY",
    "DAY": "PARKING_DAY",
    "ДЕНЬ": "PARKING_DAY",
    "PARKING_NIGHT": "PARKING_NIGHT",
    "NIGHT": "PARKING_NIGHT",
    "НОЧЬ": "PARKING_NIGHT",
    "СУТКИ": "PARKING_NIGHT",
    "Ш": "BARRIER_REPAIR",
    "ШЛАГБАУМ": "BARRIER_REPAIR",
    "BARRIER_REPAIR": "BARRIER_REPAIR",
    "РЕМОНТ_ШЛАГБАУМА": "BARRIER_REPAIR",
    "BARRIER_PHONE": "BARRIER_PHONE",
    "ТЕЛЕФОН": "BARRIER_PHONE",
    "ТЕЛЕФОННЫЙ_ДОСТУП": "BARRIER_PHONE",
    "BARRIER_PHONE_CONNECT": "BARRIER_PHONE_CONNECT",
    "ПОДКЛЮЧЕНИЕ_ТЕЛЕФОНА": "BARRIER_PHONE_CONNECT",
    "ПОДКЛЮЧЕНИЕ_ШЛАГБАУМА": "BARRIER_PHONE_CONNECT",
    "ПОВТОРНОЕ_ПОДКЛЮЧЕНИЕ": "BARRIER_PHONE_CONNECT",
    "Б": "IMPROVEMENT",
    "БЛАГОУСТРОЙСТВО": "IMPROVEMENT",
    "БЛАГОУСТРІЙ": "IMPROVEMENT",
    "IMPROVEMENT": "IMPROVEMENT",
    "PARKING_EQUIPMENT": "PARKING_EQUIPMENT",
    "ОБОРУДОВАНИЕ_ПАРКОВКИ": "PARKING_EQUIPMENT",
    "GUEST_PARKING": "GUEST_PARKING",
    "ГОСТЕВАЯ_ПАРКОВКА": "GUEST_PARKING",
    "PARK_PLACE_RENT": "PARK_PLACE_RENT",
    "PARK_PLACE_SALE": "PARK_PLACE_SALE",
}


def get_db_file():
    return paths.OSBB_TEST_DB_FILE if USE_TEST_DB else paths.OSBB_DB_FILE


def now_db():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def today():
    return datetime.now().strftime("%Y-%m-%d")


def money(value):
    value = float(value or 0)
    return int(value) if value.is_integer() else round(value, 2)


def norm_text(value):
    return "" if value is None else str(value).strip()


def norm_apartment(value):
    text = norm_text(value)
    return text[:-2] if text.endswith(".0") else text


def normalize_plate(value):
    text = norm_text(value).upper()
    cyr_to_lat = str.maketrans({
        "А": "A", "В": "B", "Е": "E", "І": "I", "К": "K",
        "М": "M", "Н": "H", "О": "O", "Р": "P", "С": "C",
        "Т": "T", "Х": "X", "У": "Y",
    })
    text = text.translate(cyr_to_lat)
    text = "".join(ch for ch in text if ch.isalnum())
    if len(text) == 4 and text.startswith("O") and text[1:].isdigit():
        text = "0" + text[1:]
    return text


def normalize_period(value, operation_date=None, required=False):
    """
    Возвращает базовый ежемесячный период YYYY-MM.
    Допускает YYYY-MM, YYYY/MM, MM_YY, MM-YY.
    """
    raw = norm_text(value)
    if not raw:
        raw = norm_text(operation_date)

    if not raw:
        if required:
            raise RuntimeError("Не указан период. Для ежемесячной услуги нужен один месяц, например 2026-06.")
        return ""

    raw = raw.replace("/", "-").replace(".", "-").replace("_", "-")

    # YYYY-MM-DD -> YYYY-MM
    if len(raw) >= 7 and raw[:4].isdigit() and raw[4] == "-" and raw[5:7].isdigit():
        return raw[:7]

    # MM-YY -> YYYY-MM
    parts = raw.split("-")
    if len(parts) == 2 and len(parts[0]) == 2 and len(parts[1]) == 2 and all(p.isdigit() for p in parts):
        return f"20{parts[1]}-{parts[0]}"

    raise RuntimeError(
        f"Неверный период: {value!r}. Используйте один месяц: 2026-06 или 06_26. "
        "Диапазон 2026-05_2026-06 здесь не допускается."
    )


def monthly_item_code(period_code, base_service_code):
    period = normalize_period(period_code, required=True)
    year = period[2:4]
    month = period[5:7]

    suffix_map = {
        "PARKING_DAY": "Day",
        "PARKING_NIGHT": "Night",
        "BARRIER_PHONE": "BarrierPhone",
    }

    if base_service_code not in suffix_map:
        raise RuntimeError(f"Неизвестная ежемесячная услуга: {base_service_code}")

    return f"{month}_{year}_{suffix_map[base_service_code]}"


def get_conn():
    conn = sqlite3.connect(get_db_file())
    conn.row_factory = sqlite3.Row
    return conn


def table_exists(cur, table_name):
    cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table_name,))
    return cur.fetchone() is not None


def table_columns(cur, table_name):
    cur.execute(f"PRAGMA table_info({table_name})")
    return [row[1] for row in cur.fetchall()]


def table_info(cur, table_name):
    cur.execute(f"PRAGMA table_info({table_name})")
    return cur.fetchall()


def required_defaults(cur, table_name):
    """
    Заполняет обязательные старые поля, которые могут остаться от ранней
    версии структуры. Это особенно важно для service_catalog.
    """
    defaults = {}
    for row in table_info(cur, table_name):
        name = row[1]
        col_type = row[2] or ""
        notnull = row[3]
        dflt_value = row[4]
        pk = row[5]

        if pk or not notnull or dflt_value is not None:
            continue

        name_upper = name.upper()
        type_upper = col_type.upper()

        if name == "service_group":
            defaults[name] = "GENERAL"
        elif name == "unit":
            defaults[name] = "service"
        elif name in {"service_name", "name", "title"}:
            defaults[name] = "Без названия"
        elif name in {"service_code", "code"}:
            defaults[name] = "UNKNOWN"
        elif "ACTIVE" in name_upper:
            defaults[name] = 1
        elif any(word in name_upper for word in ("AMOUNT", "PRICE", "SUM", "BALANCE")):
            defaults[name] = 0
        elif name_upper.endswith("_AT") or "DATE" in name_upper or "TIME" in name_upper:
            defaults[name] = now_db()
        elif "INT" in type_upper:
            defaults[name] = 0
        elif any(word in type_upper for word in ("REAL", "NUM", "DEC")):
            defaults[name] = 0
        else:
            defaults[name] = ""

    return defaults


def insert_dynamic(cur, table_name, values):
    cols = table_columns(cur, table_name)
    safe_values = dict(values)

    for col, default in required_defaults(cur, table_name).items():
        safe_values.setdefault(col, default)

    insert_cols = [name for name in safe_values if name in cols]
    if not insert_cols:
        raise RuntimeError(f"Нет подходящих колонок для вставки в {table_name}")

    placeholders = ",".join("?" for _ in insert_cols)
    cur.execute(
        f"INSERT INTO {table_name} ({', '.join(insert_cols)}) VALUES ({placeholders})",
        tuple(safe_values[name] for name in insert_cols),
    )
    return cur.lastrowid


def update_dynamic(cur, table_name, row_id, values):
    cols = table_columns(cur, table_name)
    sets, params = [], []

    for name, value in values.items():
        if name in cols:
            sets.append(f"{name} = ?")
            params.append(value)

    if not sets:
        return False

    params.append(row_id)
    cur.execute(f"UPDATE {table_name} SET {', '.join(sets)} WHERE id = ?", tuple(params))
    return cur.rowcount > 0


def update_dynamic_by(cur, table_name, where_column, where_value, values):
    cols = table_columns(cur, table_name)
    sets, params = [], []

    for name, value in values.items():
        if name in cols and name != where_column:
            sets.append(f"{name} = ?")
            params.append(value)

    if not sets:
        return False

    params.append(where_value)
    cur.execute(
        f"UPDATE {table_name} SET {', '.join(sets)} WHERE {where_column} = ?",
        tuple(params),
    )
    return cur.rowcount > 0


def service_group_for(service_type):
    if service_type == "MONTHLY":
        return "MONTHLY"
    if service_type == "FUNDRAISING":
        return "FUNDRAISING"
    if service_type == "COMMERCIAL":
        return "COMMERCIAL"
    if service_type == "RECONCILIATION":
        return "CASHIER"
    if service_type == "HISTORICAL":
        return "HISTORICAL"
    return "GENERAL"


def get_service_meta(base_service_code):
    return SERVICE_META.get(base_service_code, {
        "name": base_service_code,
        "type": "GENERAL",
        "category": "GENERAL",
        "monthly": 0,
        "fundraising": 0,
        "commercial": 0,
        "access": 0,
        "cash": 1,
    })


def ensure_service_catalog(cur, base_service_code):
    if not table_exists(cur, "service_catalog"):
        raise RuntimeError("Не найдена таблица service_catalog. Сначала выполните migrate_service_items.py.")

    meta = get_service_meta(base_service_code)

    cur.execute("SELECT id FROM service_catalog WHERE service_code = ?", (base_service_code,))
    row = cur.fetchone()

    values = {
        "service_code": base_service_code,
        "service_name": meta["name"],
        "service_type": meta["type"],
        "service_group": service_group_for(meta["type"]),
        "category": meta["category"],
        "is_monthly": int(meta["monthly"]),
        "is_fundraising": int(meta["fundraising"]),
        "is_commercial": int(meta["commercial"]),
        "is_access_control": int(meta["access"]),
        "is_cash_collectable": int(meta["cash"]),
        "is_active": 1,
        "unit": "service",
        "comment": "",
        "created_at": now_db(),
        "updated_at": now_db(),
    }

    if row:
        values.pop("service_code", None)
        values.pop("created_at", None)
        update_dynamic_by(cur, "service_catalog", "service_code", base_service_code, values)
        return False

    insert_dynamic(cur, "service_catalog", values)
    return True


def get_service_type(cur, base_service_code):
    if table_exists(cur, "service_catalog"):
        cols = table_columns(cur, "service_catalog")
        if "service_type" in cols:
            cur.execute(
                "SELECT COALESCE(service_type, '') FROM service_catalog WHERE service_code = ?",
                (base_service_code,),
            )
            row = cur.fetchone()
            if row and row[0]:
                return row[0]
    return get_service_meta(base_service_code)["type"]


def get_service_item(cur, item_code):
    if not table_exists(cur, "service_items"):
        raise RuntimeError("Не найдена таблица service_items. Сначала выполните migrate_service_items.py.")

    cur.execute("SELECT * FROM service_items WHERE service_item_code = ?", (item_code,))
    row = cur.fetchone()
    return dict(row) if row else None


def ensure_service_item(
    cur,
    item_code,
    base_service_code,
    item_name,
    service_type,
    period_code="",
    sequence_no=None,
    amount_default=None,
    date_from="",
    date_to="",
    description="",
    comment="",
):
    """
    Создаёт статью либо бережно обновляет её.
    При автоматическом использовании статьи нельзя затирать уже заданные
    сумму, описание или сроки пустыми значениями.
    """
    ensure_service_catalog(cur, base_service_code)

    existing = get_service_item(cur, item_code)
    values = {
        "service_item_code": item_code,
        "service_code": base_service_code,
        "service_item_name": item_name,
        "service_type": service_type,
        "period_code": period_code or None,
        "sequence_no": sequence_no,
        "amount_default": amount_default,
        "currency": "UAH",
        "date_from": date_from or None,
        "date_to": date_to or None,
        "status": "active",
        "is_active": 1,
        "description": description,
        "comment": comment,
        "created_at": now_db(),
        "updated_at": now_db(),
    }

    if existing:
        # Постоянные поля можно синхронизировать.
        update_values = {
            "service_code": base_service_code,
            "service_item_name": item_name,
            "service_type": service_type,
            "currency": "UAH",
            "status": "active",
            "is_active": 1,
            "updated_at": now_db(),
        }

        # Не затираем заполненные данные пустыми/None при автоматическом вызове.
        if period_code:
            update_values["period_code"] = period_code
        if sequence_no is not None:
            update_values["sequence_no"] = sequence_no
        if amount_default is not None:
            update_values["amount_default"] = amount_default
        if date_from:
            update_values["date_from"] = date_from
        if date_to:
            update_values["date_to"] = date_to
        if description:
            update_values["description"] = description
        if comment:
            update_values["comment"] = comment

        update_dynamic_by(cur, "service_items", "service_item_code", item_code, update_values)
        return False

    insert_dynamic(cur, "service_items", values)
    return True


def resolve_base_service(raw_service, vehicle=None):
    raw = norm_text(raw_service).upper()
    service = SERVICE_ALIASES.get(raw, raw)

    if service != "PARKING":
        return service

    parking_time = norm_text((vehicle or {}).get("parking_time"))

    if parking_time == "Day":
        return "PARKING_DAY"
    if parking_time == "Night":
        return "PARKING_NIGHT"

    raise RuntimeError(
        "Для PARKING не удалось определить Day/Night. "
        "Укажите --service PARKING_DAY или PARKING_NIGHT, "
        "либо заполните parking_time у автомобиля."
    )


def resolve_service_item(cur, base_service_code, period_code, explicit_item="", amount=None, date_from="", date_to="", comment=""):
    service_type = get_service_type(cur, base_service_code)
    explicit_item = norm_text(explicit_item)

    if explicit_item:
        existing = get_service_item(cur, explicit_item)
        if not existing:
            raise RuntimeError(
                f"Статья {explicit_item!r} не найдена. "
                "Сначала создайте её: cashier_journal.py service-item add ..."
            )
        if existing.get("service_code") != base_service_code:
            raise RuntimeError(
                f"Статья {explicit_item!r} относится к {existing.get('service_code')}, "
                f"а указана услуга {base_service_code}."
            )
        return explicit_item, service_type, existing.get("period_code") or period_code

    if service_type == "MONTHLY":
        month = normalize_period(period_code, required=True)
        item_code = monthly_item_code(month, base_service_code)
        item_name = f"{get_service_meta(base_service_code)['name']} — {month}"
        ensure_service_item(
            cur=cur,
            item_code=item_code,
            base_service_code=base_service_code,
            item_name=item_name,
            service_type=service_type,
            period_code=month,
            amount_default=None,
            date_from=f"{month}-01",
            comment="Автоматически создана ежемесячная статья",
        )
        return item_code, service_type, month

    raise RuntimeError(
        f"Для услуги {base_service_code} нужна конкретная статья. "
        "Например: --item 01_Imp, --item 01_ParkEquip или --item 01_BarrierRepair."
    )


def ensure_cashbox(cur, cashbox_code):
    code = norm_text(cashbox_code).upper().replace("О", "O")
    cur.execute("SELECT * FROM cashboxes WHERE cashbox_code = ?", (code,))
    row = cur.fetchone()
    if row:
        return dict(row)

    if code == "CENTRAL":
        insert_dynamic(cur, "cashboxes", {
            "cashbox_code": "CENTRAL",
            "cashbox_name": "Центральная касса",
            "currency": "UAH",
            "initial_balance": 0,
            "current_balance": 0,
            "is_active": 1,
            "comment": "Центральная касса для инкассации",
            "created_at": now_db(),
            "updated_at": now_db(),
        })
        cur.execute("SELECT * FROM cashboxes WHERE cashbox_code = ?", (code,))
        return dict(cur.fetchone())

    raise RuntimeError(
        f"Касса не найдена: {code}. "
        "Допустимые текущие кассы: O, K, BANK; CENTRAL создаётся автоматически при первой передаче."
    )


def recalc_cashbox_balance(cur, cashbox_code):
    cur.execute("SELECT initial_balance FROM cashboxes WHERE cashbox_code = ?", (cashbox_code,))
    row = cur.fetchone()
    if not row:
        raise RuntimeError(f"Касса не найдена: {cashbox_code}")

    initial = float(row[0] or 0)
    cur.execute("""
        SELECT COALESCE(SUM(
            CASE WHEN direction = 'in' THEN amount
                 WHEN direction = 'out' THEN -amount
                 ELSE 0 END
        ), 0)
        FROM cashbox_operations
        WHERE cashbox_code = ?
    """, (cashbox_code,))
    delta = float(cur.fetchone()[0] or 0)
    balance = initial + delta

    update_dynamic_by(
        cur,
        "cashboxes",
        "cashbox_code",
        cashbox_code,
        {"current_balance": balance, "updated_at": now_db()},
    )
    return balance


def recalc_all_cashboxes(cur):
    cur.execute("SELECT cashbox_code FROM cashboxes WHERE COALESCE(is_active, 1) = 1 ORDER BY cashbox_code")
    result = {}
    for row in cur.fetchall():
        code = row[0]
        result[code] = recalc_cashbox_balance(cur, code)
    return result


def find_vehicle(cur, apt=None, plate=None, vehicle_id=None):
    if vehicle_id:
        cur.execute("""
            SELECT
                v.id AS vehicle_id,
                a.apartment_number,
                COALESCE(v.license_plate_normalized, v.license_plate, '') AS plate,
                COALESCE(v.car_model_normalized, v.car_model, '') AS model,
                COALESCE(v.parking_time, '') AS parking_time
            FROM vehicles v
            JOIN apartments a ON a.id = v.apartment_id
            WHERE v.id = ?
        """, (vehicle_id,))
        row = cur.fetchone()
        return dict(row) if row else None

    apt = norm_apartment(apt)
    plate = normalize_plate(plate)

    where, params = [], []
    if apt:
        where.append("a.apartment_number = ?")
        params.append(apt)
    if plate:
        where.append("COALESCE(v.license_plate_normalized, v.license_plate, '') = ?")
        params.append(plate)

    if not where:
        return None

    cur.execute(f"""
        SELECT
            v.id AS vehicle_id,
            a.apartment_number,
            COALESCE(v.license_plate_normalized, v.license_plate, '') AS plate,
            COALESCE(v.car_model_normalized, v.car_model, '') AS model,
            COALESCE(v.parking_time, '') AS parking_time
        FROM vehicles v
        JOIN apartments a ON a.id = v.apartment_id
        WHERE {' AND '.join(where)}
        ORDER BY v.id
    """, tuple(params))

    rows = [dict(row) for row in cur.fetchall()]
    return rows[0] if len(rows) == 1 else None


def duplicate_operation_exists(cur, source_ref):
    if not source_ref:
        return False
    cur.execute("SELECT id FROM cashbox_operations WHERE source_ref = ? LIMIT 1", (source_ref,))
    return cur.fetchone() is not None


def find_charge_for_payment(cur, period_code, apt, vehicle_id, base_service_code, service_item_code):
    if not table_exists(cur, "charges"):
        return None

    cols = table_columns(cur, "charges")
    where_common, params_common = [], []

    if vehicle_id:
        where_common.append("vehicle_id = ?")
        params_common.append(vehicle_id)
    elif apt:
        where_common.append("apartment_number = ?")
        params_common.append(norm_apartment(apt))
    else:
        return None

    # New exact model: service_item_code.
    if "service_item_code" in cols and service_item_code:
        sql = f"""
            SELECT id, amount
            FROM charges
            WHERE service_item_code = ?
              AND {' AND '.join(where_common)}
            ORDER BY id
            LIMIT 1
        """
        cur.execute(sql, tuple([service_item_code] + params_common))
        row = cur.fetchone()
        if row:
            return dict(row)

    # Legacy model: monthly period + service_code.
    if "period_code" in cols and "service_code" in cols and period_code and base_service_code:
        sql = f"""
            SELECT id, amount
            FROM charges
            WHERE period_code = ?
              AND service_code = ?
              AND {' AND '.join(where_common)}
            ORDER BY id
            LIMIT 1
        """
        cur.execute(sql, tuple([period_code, base_service_code] + params_common))
        row = cur.fetchone()
        if row:
            return dict(row)

    return None


def allocated_amount_for_charge(cur, charge_id):
    if not table_exists(cur, "payment_allocations"):
        return 0.0

    cols = table_columns(cur, "payment_allocations")
    amount_col = "amount" if "amount" in cols else ("allocated_amount" if "allocated_amount" in cols else None)
    if not amount_col:
        return 0.0

    cur.execute(
        f"SELECT COALESCE(SUM({amount_col}), 0) FROM payment_allocations WHERE charge_id = ?",
        (charge_id,),
    )
    return float(cur.fetchone()[0] or 0)


def create_allocation(cur, payment_id, charge_id, amount):
    if not table_exists(cur, "payment_allocations"):
        return None

    return insert_dynamic(cur, "payment_allocations", {
        "payment_id": payment_id,
        "charge_id": charge_id,
        "amount": float(amount),
        "allocated_amount": float(amount),
        "created_at": now_db(),
    })


def create_cashbox_operation(
    cur,
    operation_date,
    cashbox_code,
    operation_type,
    direction,
    amount,
    base_service_code="",
    service_item_code="",
    service_type="",
    period_code="",
    apartment_number="",
    vehicle_id=None,
    payment_id=None,
    charge_id=None,
    batch_id="",
    source_type="cashier_journal",
    source_ref="",
    operator_id="",
    actor_type="operator",
    comment="",
):
    if source_ref and duplicate_operation_exists(cur, source_ref):
        raise RuntimeError(f"Операция с source_ref уже существует: {source_ref}")

    return insert_dynamic(cur, "cashbox_operations", {
        "operation_date": operation_date,
        "cashbox_code": cashbox_code,
        "operation_type": operation_type,
        "direction": direction,
        "amount": float(amount),
        "currency": "UAH",
        "period_code": period_code or None,
        "apartment_number": norm_apartment(apartment_number) or None,
        "vehicle_id": vehicle_id,
        "service_code": base_service_code or None,        # совместимость со старой структурой
        "base_service_code": base_service_code or None,
        "service_item_code": service_item_code or None,
        "service_type": service_type or None,
        "payment_id": payment_id,
        "charge_id": charge_id,
        "batch_id": batch_id or None,
        "source_type": source_type,
        "source_ref": source_ref or None,
        "operator_id": operator_id or None,
        "actor_type": actor_type,
        "comment": comment,
        "created_at": now_db(),
    })


def print_operation_preview(title, data):
    print("=" * 78)
    print(title)
    print("=" * 78)
    for key, value in data:
        print(f"{key:22}: {value}")
    print("")


def command_balance(args):
    conn = get_conn()
    cur = conn.cursor()

    if args.recalc:
        recalc_all_cashboxes(cur)
        conn.commit()

    if args.cashbox:
        cur.execute("""
            SELECT cashbox_code, cashbox_name, currency, current_balance, is_active
            FROM cashboxes
            WHERE cashbox_code = ?
            ORDER BY cashbox_code
        """, (args.cashbox.strip().upper(),))
    else:
        cur.execute("""
            SELECT cashbox_code, cashbox_name, currency, current_balance, is_active
            FROM cashboxes
            WHERE COALESCE(is_active, 1) = 1
            ORDER BY cashbox_code
        """)
    boxes = [dict(row) for row in cur.fetchall()]

    print("=" * 78)
    print("ОСТАТКИ КАСС")
    print("=" * 78)
    print("DB:", get_db_file())
    print("MODE:", "TEST/WORK" if USE_TEST_DB else "PROD")
    print("")

    for box in boxes:
        print(f"{box['cashbox_code']:10} {box['cashbox_name']:28} {money(box['current_balance']):>10} {box['currency']}")

    if args.by_service:
        print("")
        print("РАЗРЕЗ ПО СТАТЬЯМ")
        print("-" * 78)

        where, params = [], []
        if args.cashbox:
            where.append("cashbox_code = ?")
            params.append(args.cashbox.strip().upper())

        where_sql = "WHERE " + " AND ".join(where) if where else ""
        cur.execute(f"""
            SELECT
                cashbox_code,
                COALESCE(NULLIF(service_item_code, ''), NULLIF(service_code, ''), 'UNCLASSIFIED') AS item_code,
                COALESCE(NULLIF(base_service_code, ''), NULLIF(service_code, ''), 'UNCLASSIFIED') AS base_code,
                COALESCE(NULLIF(service_type, ''), 'GENERAL') AS service_type,
                COALESCE(SUM(
                    CASE WHEN direction = 'in' THEN amount
                         WHEN direction = 'out' THEN -amount
                         ELSE 0 END
                ), 0) AS balance
            FROM cashbox_operations
            {where_sql}
            GROUP BY cashbox_code, item_code, base_code, service_type
            HAVING ABS(COALESCE(SUM(
                CASE WHEN direction = 'in' THEN amount
                     WHEN direction = 'out' THEN -amount
                     ELSE 0 END
            ), 0)) > 0.00001
            ORDER BY cashbox_code, service_type, base_code, item_code
        """, tuple(params))

        rows = [dict(row) for row in cur.fetchall()]
        if not rows:
            print("Операций по статьям пока нет.")
        else:
            current_box = None
            for row in rows:
                if row["cashbox_code"] != current_box:
                    current_box = row["cashbox_code"]
                    print(f"\n{current_box}:")
                print(
                    f"  {row['item_code']:28} | {row['base_code']:22} | "
                    f"{money(row['balance']):>10} UAH"
                )

    conn.close()
    print("=" * 78)


def command_income(args):
    operation_date = args.date or today()
    cashbox_code = args.cashbox.strip().upper().replace("О", "O")
    amount = float(args.amount)
    operator_id = args.operator or "cashier_cli"
    batch_id = args.batch_id or ""

    conn = get_conn()
    cur = conn.cursor()

    ensure_cashbox(cur, cashbox_code)

    vehicle = find_vehicle(cur, args.apt, args.plate, args.vehicle_id)
    vehicle_id = vehicle["vehicle_id"] if vehicle else None
    apt = norm_apartment(args.apt or (vehicle["apartment_number"] if vehicle else ""))

    base_service_code = resolve_base_service(args.service, vehicle)
    service_type = get_service_type(cur, base_service_code)
    period = normalize_period(args.period, operation_date, required=(service_type == "MONTHLY"))
    item_code, service_type, item_period = resolve_service_item(
        cur,
        base_service_code=base_service_code,
        period_code=period,
        explicit_item=args.item,
        amount=amount,
        comment=args.comment or "",
    )

    charge = find_charge_for_payment(
        cur,
        period_code=item_period,
        apt=apt,
        vehicle_id=vehicle_id,
        base_service_code=base_service_code,
        service_item_code=item_code,
    )
    charge_id = charge["id"] if charge else None
    allocation_amount = 0.0

    if charge:
        available = max(0.0, float(charge["amount"] or 0) - allocated_amount_for_charge(cur, charge_id))
        allocation_amount = min(amount, available)

    print_operation_preview("ПРИЁМ ПЛАТЕЖА", [
        ("Касса", cashbox_code),
        ("Дата", operation_date),
        ("Период", item_period or "-"),
        ("Квартира", apt or "-"),
        ("Авто", f"{vehicle_id} | {vehicle.get('plate')} | {vehicle.get('model')}" if vehicle else "-"),
        ("Базовый сервис", base_service_code),
        ("Статья", item_code),
        ("Тип услуги", service_type),
        ("Сумма", f"{money(amount)} UAH"),
        ("Начисление", charge_id or "-"),
        ("Будет распределено", f"{money(allocation_amount)} UAH"),
        ("Применить", args.apply),
    ])

    if not args.apply:
        conn.close()
        print("DRY RUN ONLY. Добавьте --apply для сохранения.")
        return

    if args.source_ref and duplicate_operation_exists(cur, args.source_ref):
        conn.close()
        raise RuntimeError(f"source_ref уже использован: {args.source_ref}")

    payment_id = insert_dynamic(cur, "payments", {
        "payment_date": operation_date,
        "period_code": item_period or None,
        "apartment_number": apt or None,
        "vehicle_id": vehicle_id,
        "service_code": base_service_code,
        "base_service_code": base_service_code,
        "service_item_code": item_code,
        "service_type": service_type,
        "amount": amount,
        "currency": "UAH",
        "payment_method": "cash",
        "source": "cashier_journal",
        "source_ref": args.source_ref or f"cashier:{uuid4()}",
        "comment": args.comment or "",
        "cashbox_code": cashbox_code,
        "cashier_batch_id": batch_id or None,
        "operator_id": operator_id,
        "created_at": now_db(),
    })

    operation_id = create_cashbox_operation(
        cur=cur,
        operation_date=operation_date,
        cashbox_code=cashbox_code,
        operation_type="income",
        direction="in",
        amount=amount,
        base_service_code=base_service_code,
        service_item_code=item_code,
        service_type=service_type,
        period_code=item_period,
        apartment_number=apt,
        vehicle_id=vehicle_id,
        payment_id=payment_id,
        charge_id=charge_id,
        batch_id=batch_id,
        source_type="cashier_journal",
        source_ref=args.source_ref or f"cashier_operation:{payment_id}",
        operator_id=operator_id,
        actor_type="operator",
        comment=args.comment or "",
    )

    update_dynamic(cur, "payments", payment_id, {
        "cashbox_operation_id": operation_id,
    })

    allocation_id = None
    if charge_id and allocation_amount > 0:
        allocation_id = create_allocation(cur, payment_id, charge_id, allocation_amount)

    new_balance = recalc_cashbox_balance(cur, cashbox_code)

    audit_log(
        conn=conn,
        actor_type="operator",
        operator_id=operator_id,
        user_id=operator_id,
        action_type="cashier_income_created",
        table_name="payments",
        row_id=payment_id,
        field_name="amount",
        old_value="",
        new_value=amount,
        source_context="cashier_journal.py income",
        comment=f"Принят платёж в кассу {cashbox_code}; статья {item_code}; операция {operation_id}",
        extra={
            "cashbox_code": cashbox_code,
            "cashbox_operation_id": operation_id,
            "payment_id": payment_id,
            "allocation_id": allocation_id,
            "charge_id": charge_id,
            "base_service_code": base_service_code,
            "service_item_code": item_code,
            "period_code": item_period,
        },
        commit=False,
    )

    conn.commit()
    conn.close()

    print("Создан платёж:", payment_id)
    print("Создана операция кассы:", operation_id)
    print("Распределение:", allocation_id or "-")
    print("Остаток кассы:", money(new_balance), "UAH")


def command_expense(args):
    operation_date = args.date or today()
    cashbox_code = args.cashbox.strip().upper().replace("О", "O")
    amount = float(args.amount)
    operator_id = args.operator or "cashier_cli"

    conn = get_conn()
    cur = conn.cursor()

    ensure_cashbox(cur, cashbox_code)
    base_service_code = resolve_base_service(args.service)
    service_type = get_service_type(cur, base_service_code)
    period = normalize_period(args.period, operation_date, required=(service_type == "MONTHLY"))
    item_code, service_type, item_period = resolve_service_item(
        cur,
        base_service_code=base_service_code,
        period_code=period,
        explicit_item=args.item,
        amount=None,
        comment=args.comment or "",
    )

    print_operation_preview("РАСХОД ИЗ КАССЫ", [
        ("Касса", cashbox_code),
        ("Дата", operation_date),
        ("Период", item_period or "-"),
        ("Базовый сервис", base_service_code),
        ("Статья", item_code),
        ("Сумма", f"{money(amount)} UAH"),
        ("Применить", args.apply),
    ])

    if not args.apply:
        conn.close()
        print("DRY RUN ONLY. Добавьте --apply для сохранения.")
        return

    source_ref = args.source_ref or f"cashier_expense:{uuid4()}"

    operation_id = create_cashbox_operation(
        cur=cur,
        operation_date=operation_date,
        cashbox_code=cashbox_code,
        operation_type="expense",
        direction="out",
        amount=amount,
        base_service_code=base_service_code,
        service_item_code=item_code,
        service_type=service_type,
        period_code=item_period,
        apartment_number=args.apt,
        vehicle_id=args.vehicle_id or None,
        source_type="cashier_journal",
        source_ref=source_ref,
        operator_id=operator_id,
        actor_type="operator",
        comment=args.comment or "",
    )

    new_balance = recalc_cashbox_balance(cur, cashbox_code)

    audit_log(
        conn=conn,
        actor_type="operator",
        operator_id=operator_id,
        user_id=operator_id,
        action_type="cashier_expense_created",
        table_name="cashbox_operations",
        row_id=operation_id,
        field_name="amount",
        old_value="",
        new_value=amount,
        source_context="cashier_journal.py expense",
        comment=f"Расход из кассы {cashbox_code}; статья {item_code}",
        extra={
            "cashbox_code": cashbox_code,
            "cashbox_operation_id": operation_id,
            "base_service_code": base_service_code,
            "service_item_code": item_code,
            "period_code": item_period,
        },
        commit=False,
    )

    conn.commit()
    conn.close()

    print("Создан расход:", operation_id)
    print("Остаток кассы:", money(new_balance), "UAH")


def command_transfer(args):
    operation_date = args.date or today()
    from_code = args.from_cashbox.strip().upper().replace("О", "O")
    to_code = args.to_cashbox.strip().upper().replace("О", "O")
    amount = float(args.amount)
    operator_id = args.operator or "cashier_cli"

    if from_code == to_code:
        raise RuntimeError("Источник и получатель передачи совпадают.")

    conn = get_conn()
    cur = conn.cursor()

    ensure_cashbox(cur, from_code)
    ensure_cashbox(cur, to_code)

    base_service_code = resolve_base_service(args.service)
    service_type = get_service_type(cur, base_service_code)
    period = normalize_period(args.period, operation_date, required=(service_type == "MONTHLY"))
    item_code, service_type, item_period = resolve_service_item(
        cur,
        base_service_code=base_service_code,
        period_code=period,
        explicit_item=args.item,
        amount=None,
        comment=args.comment or "",
    )

    transfer_ref = args.source_ref or f"transfer:{uuid4()}"

    print_operation_preview("ПЕРЕДАЧА МЕЖДУ КАССАМИ", [
        ("Дата", operation_date),
        ("Из кассы", from_code),
        ("В кассу", to_code),
        ("Статья", item_code),
        ("Сумма", f"{money(amount)} UAH"),
        ("Применить", args.apply),
    ])

    if not args.apply:
        conn.close()
        print("DRY RUN ONLY. Добавьте --apply для сохранения.")
        return

    out_id = create_cashbox_operation(
        cur=cur,
        operation_date=operation_date,
        cashbox_code=from_code,
        operation_type="transfer_out",
        direction="out",
        amount=amount,
        base_service_code=base_service_code,
        service_item_code=item_code,
        service_type=service_type,
        period_code=item_period,
        source_type="cashier_transfer",
        source_ref=f"{transfer_ref}:out",
        operator_id=operator_id,
        actor_type="operator",
        comment=args.comment or f"Передача в кассу {to_code}",
    )

    in_id = create_cashbox_operation(
        cur=cur,
        operation_date=operation_date,
        cashbox_code=to_code,
        operation_type="transfer_in",
        direction="in",
        amount=amount,
        base_service_code=base_service_code,
        service_item_code=item_code,
        service_type=service_type,
        period_code=item_period,
        source_type="cashier_transfer",
        source_ref=f"{transfer_ref}:in",
        operator_id=operator_id,
        actor_type="operator",
        comment=args.comment or f"Получено из кассы {from_code}",
    )

    from_balance = recalc_cashbox_balance(cur, from_code)
    to_balance = recalc_cashbox_balance(cur, to_code)

    audit_log(
        conn=conn,
        actor_type="operator",
        operator_id=operator_id,
        user_id=operator_id,
        action_type="cashbox_transfer",
        table_name="cashbox_operations",
        row_id=f"{out_id},{in_id}",
        field_name="amount",
        old_value="",
        new_value=amount,
        source_context="cashier_journal.py transfer",
        comment=f"Передача {from_code} → {to_code}; статья {item_code}",
        extra={
            "from_cashbox": from_code,
            "to_cashbox": to_code,
            "out_operation_id": out_id,
            "in_operation_id": in_id,
            "base_service_code": base_service_code,
            "service_item_code": item_code,
        },
        commit=False,
    )

    conn.commit()
    conn.close()

    print("Операция списания:", out_id)
    print("Операция поступления:", in_id)
    print(f"{from_code}:", money(from_balance), "UAH")
    print(f"{to_code}:", money(to_balance), "UAH")


def command_service_item_add(args):
    base_service_code = resolve_base_service(args.service)
    meta = get_service_meta(base_service_code)
    service_type = get_service_type(get_conn().cursor(), base_service_code) if False else meta["type"]

    period = ""
    if meta["type"] == "MONTHLY":
        period = normalize_period(args.period, required=True)

    conn = get_conn()
    cur = conn.cursor()

    ensure_service_catalog(cur, base_service_code)
    service_type = get_service_type(cur, base_service_code)

    item_code = args.code.strip()
    item_name = args.name.strip()
    if not item_code or not item_name:
        raise RuntimeError("Для новой статьи нужны --code и --name.")

    print_operation_preview("НОВАЯ СТАТЬЯ СЕРВИСА", [
        ("Код статьи", item_code),
        ("Базовый сервис", base_service_code),
        ("Название", item_name),
        ("Тип", service_type),
        ("Период", period or "-"),
        ("Очередь", args.sequence if args.sequence is not None else "-"),
        ("Сумма по умолчанию", money(args.amount) if args.amount is not None else "-"),
        ("Применить", args.apply),
    ])

    if not args.apply:
        conn.close()
        print("DRY RUN ONLY. Добавьте --apply для сохранения.")
        return

    created = ensure_service_item(
        cur=cur,
        item_code=item_code,
        base_service_code=base_service_code,
        item_name=item_name,
        service_type=service_type,
        period_code=period,
        sequence_no=args.sequence,
        amount_default=args.amount,
        date_from=args.date_from,
        date_to=args.date_to,
        description=args.description,
        comment=args.comment,
    )

    audit_log(
        conn=conn,
        actor_type="operator",
        operator_id=args.operator or "cashier_cli",
        user_id=args.operator or "cashier_cli",
        action_type="service_item_created" if created else "service_item_updated",
        table_name="service_items",
        row_id=item_code,
        field_name="service_item_code",
        old_value="",
        new_value=item_code,
        source_context="cashier_journal.py service-item add",
        comment=args.comment or item_name,
        extra={
            "base_service_code": base_service_code,
            "service_type": service_type,
            "period_code": period,
            "sequence_no": args.sequence,
            "amount_default": args.amount,
        },
        commit=False,
    )

    conn.commit()
    conn.close()
    print("Статья сохранена:", item_code)


def command_service_item_list(args):
    conn = get_conn()
    cur = conn.cursor()

    where, params = [], []
    if args.service:
        where.append("service_code = ?")
        params.append(resolve_base_service(args.service))
    if args.active_only:
        where.append("COALESCE(is_active, 1) = 1")

    where_sql = "WHERE " + " AND ".join(where) if where else ""
    cur.execute(f"""
        SELECT
            service_item_code,
            service_code,
            service_type,
            period_code,
            sequence_no,
            amount_default,
            date_from,
            date_to,
            service_item_name,
            comment
        FROM service_items
        {where_sql}
        ORDER BY service_code, period_code, sequence_no, service_item_code
    """, tuple(params))

    rows = [dict(row) for row in cur.fetchall()]
    conn.close()

    print("=" * 120)
    print("СТАТЬИ СЕРВИСОВ")
    print("=" * 120)
    if not rows:
        print("Нет строк.")
        return

    print("code | base service | type | period | queue | amount | name | comment")
    print("-" * 120)
    for row in rows:
        print(
            f"{row['service_item_code']} | {row['service_code']} | {row['service_type']} | "
            f"{row['period_code'] or '-'} | {row['sequence_no'] if row['sequence_no'] is not None else '-'} | "
            f"{money(row['amount_default']) if row['amount_default'] is not None else '-'} | "
            f"{row['service_item_name']} | {row['comment'] or ''}"
        )


def load_operations(cur, cashbox="", limit=200):
    where, params = [], []
    if cashbox:
        where.append("cashbox_code = ?")
        params.append(cashbox.strip().upper())

    where_sql = "WHERE " + " AND ".join(where) if where else ""
    params.append(limit)

    cur.execute(f"""
        SELECT
            id, operation_date, cashbox_code, operation_type, direction, amount,
            period_code, apartment_number, vehicle_id,
            service_code, base_service_code, service_item_code, service_type,
            payment_id, charge_id, operator_id, comment, created_at
        FROM cashbox_operations
        {where_sql}
        ORDER BY id DESC
        LIMIT ?
    """, tuple(params))
    return [dict(row) for row in cur.fetchall()]


def command_report(args):
    conn = get_conn()
    cur = conn.cursor()

    recalc_all_cashboxes(cur)
    conn.commit()

    operations = load_operations(cur, cashbox=args.cashbox, limit=args.limit)
    cur.execute("""
        SELECT cashbox_code, cashbox_name, currency, current_balance
        FROM cashboxes
        WHERE COALESCE(is_active, 1) = 1
        ORDER BY cashbox_code
    """)
    boxes = [dict(row) for row in cur.fetchall()]

    cur.execute("""
        SELECT
            cashbox_code,
            COALESCE(NULLIF(service_item_code, ''), NULLIF(service_code, ''), 'UNCLASSIFIED') AS item_code,
            COALESCE(NULLIF(base_service_code, ''), NULLIF(service_code, ''), 'UNCLASSIFIED') AS base_code,
            COALESCE(SUM(CASE WHEN direction='in' THEN amount WHEN direction='out' THEN -amount ELSE 0 END), 0) AS balance
        FROM cashbox_operations
        GROUP BY cashbox_code, item_code, base_code
        HAVING ABS(COALESCE(SUM(CASE WHEN direction='in' THEN amount WHEN direction='out' THEN -amount ELSE 0 END), 0)) > 0.00001
        ORDER BY cashbox_code, base_code, item_code
    """)
    service_balances = [dict(row) for row in cur.fetchall()]
    conn.close()

    report_dir = paths.OSBB_EXPORTS_DIR / "cashier"
    report_dir.mkdir(parents=True, exist_ok=True)
    report_file = report_dir / f"cashier_report_{datetime.now():%Y-%m-%d_%H-%M-%S}.txt"

    lines = [
        "=" * 120,
        "ЖУРНАЛ КАССИРА",
        "=" * 120,
        f"Generated: {now_db()}",
        f"DB       : {get_db_file()}",
        f"MODE     : {'TEST/WORK' if USE_TEST_DB else 'PROD'}",
        "",
        "ОСТАТКИ КАСС:",
    ]

    for box in boxes:
        lines.append(f"  {box['cashbox_code']}: {money(box['current_balance'])} {box['currency']} — {box['cashbox_name']}")

    lines.extend(["", "ОСТАТКИ ПО СТАТЬЯМ:", "cashbox | item | base service | balance", "-" * 120])
    for row in service_balances:
        lines.append(f"{row['cashbox_code']} | {row['item_code']} | {row['base_code']} | {money(row['balance'])}")

    lines.extend([
        "",
        "ПОСЛЕДНИЕ ОПЕРАЦИИ:",
        "id | date | cashbox | type | direction | amount | item | apt | vehicle | payment | charge | operator | comment",
        "-" * 120,
    ])

    for op in operations:
        lines.append(
            f"{op['id']} | {op['operation_date']} | {op['cashbox_code']} | {op['operation_type']} | "
            f"{op['direction']} | {money(op['amount'])} | {op.get('service_item_code') or op.get('service_code') or '-'} | "
            f"{op['apartment_number'] or '-'} | {op['vehicle_id'] or '-'} | {op['payment_id'] or '-'} | "
            f"{op['charge_id'] or '-'} | {op['operator_id'] or '-'} | {op['comment'] or ''}"
        )

    report_file.write_text("\n".join(lines), encoding="utf-8")

    xlsx_file = None
    if args.xlsx:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise RuntimeError("Для Excel-отчёта нужен пакет openpyxl.")

        xlsx_file = report_file.with_suffix(".xlsx")
        wb = Workbook()

        ws = wb.active
        ws.title = "Операции"
        ws.append([
            "id", "Дата", "Касса", "Тип", "Направление", "Сумма",
            "Период", "Квартира", "vehicle_id",
            "Базовый сервис", "Статья", "Тип сервиса",
            "payment_id", "charge_id", "Оператор", "Комментарий", "Создано",
        ])
        for op in operations:
            ws.append([
                op["id"], op["operation_date"], op["cashbox_code"],
                op["operation_type"], op["direction"], op["amount"],
                op["period_code"], op["apartment_number"], op["vehicle_id"],
                op.get("base_service_code") or op.get("service_code"),
                op.get("service_item_code"), op.get("service_type"),
                op["payment_id"], op["charge_id"], op["operator_id"],
                op["comment"], op["created_at"],
            ])

        ws2 = wb.create_sheet("Остатки касс")
        ws2.append(["Касса", "Название", "Валюта", "Остаток"])
        for box in boxes:
            ws2.append([
                box["cashbox_code"], box["cashbox_name"],
                box["currency"], box["current_balance"],
            ])

        ws3 = wb.create_sheet("Остатки по статьям")
        ws3.append(["Касса", "Код статьи", "Базовый сервис", "Остаток"])
        for row in service_balances:
            ws3.append([
                row["cashbox_code"], row["item_code"],
                row["base_code"], row["balance"],
            ])

        for sheet in wb.worksheets:
            for cell in sheet[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="D9EAF7")
                cell.alignment = Alignment(vertical="top", wrap_text=True)

            for row in sheet.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

            for col_idx in range(1, sheet.max_column + 1):
                letter = get_column_letter(col_idx)
                max_len = 10
                for cell in sheet[letter]:
                    if cell.value is not None:
                        max_len = max(max_len, min(len(str(cell.value)), 45))
                sheet.column_dimensions[letter].width = max_len + 2
            sheet.freeze_panes = "A2"

        wb.save(xlsx_file)

    print("=" * 78)
    print("ОТЧЁТ КАССИРА")
    print("=" * 78)
    print("TXT :", report_file)
    if xlsx_file:
        print("XLSX:", xlsx_file)
    print("=" * 78)


def main():
    parser = argparse.ArgumentParser(
        description="Журнал кассира ОСББ: платежи, расходы, передачи, остатки и сервисные статьи."
    )
    sub = parser.add_subparsers(dest="command", required=True)

    p_balance = sub.add_parser("balance", help="Остатки касс")
    p_balance.add_argument("--cashbox", default="")
    p_balance.add_argument("--recalc", action="store_true")
    p_balance.add_argument("--by-service", action="store_true")
    p_balance.set_defaults(func=command_balance)

    p_income = sub.add_parser("income", help="Принять платёж")
    p_income.add_argument("--cashbox", default="O")
    p_income.add_argument("--amount", required=True, type=float)
    p_income.add_argument("--service", required=True)
    p_income.add_argument("--item", default="", help="Код конкретной статьи: 01_Imp, 01_ParkEquip...")
    p_income.add_argument("--period", default="", help="Один месяц: 2026-06 или 06_26")
    p_income.add_argument("--date", default="")
    p_income.add_argument("--apt", default="")
    p_income.add_argument("--plate", default="")
    p_income.add_argument("--vehicle-id", type=int, default=0)
    p_income.add_argument("--operator", default="cashier_cli")
    p_income.add_argument("--comment", default="")
    p_income.add_argument("--source-ref", default="")
    p_income.add_argument("--batch-id", default="")
    p_income.add_argument("--apply", action="store_true")
    p_income.set_defaults(func=command_income)

    p_expense = sub.add_parser("expense", help="Записать расход")
    p_expense.add_argument("--cashbox", default="O")
    p_expense.add_argument("--amount", required=True, type=float)
    p_expense.add_argument("--service", required=True)
    p_expense.add_argument("--item", default="", required=False)
    p_expense.add_argument("--period", default="")
    p_expense.add_argument("--date", default="")
    p_expense.add_argument("--apt", default="")
    p_expense.add_argument("--vehicle-id", type=int, default=0)
    p_expense.add_argument("--operator", default="cashier_cli")
    p_expense.add_argument("--comment", default="")
    p_expense.add_argument("--source-ref", default="")
    p_expense.add_argument("--apply", action="store_true")
    p_expense.set_defaults(func=command_expense)

    p_transfer = sub.add_parser("transfer", help="Передать деньги между кассами по одной статье")
    p_transfer.add_argument("--from-cashbox", required=True)
    p_transfer.add_argument("--to-cashbox", required=True)
    p_transfer.add_argument("--amount", required=True, type=float)
    p_transfer.add_argument("--service", required=True)
    p_transfer.add_argument("--item", default="", required=False)
    p_transfer.add_argument("--period", default="")
    p_transfer.add_argument("--date", default="")
    p_transfer.add_argument("--operator", default="cashier_cli")
    p_transfer.add_argument("--comment", default="")
    p_transfer.add_argument("--source-ref", default="")
    p_transfer.add_argument("--apply", action="store_true")
    p_transfer.set_defaults(func=command_transfer)

    p_item = sub.add_parser("service-item", help="Управление конкретными статьями сервисов")
    item_sub = p_item.add_subparsers(dest="item_command", required=True)

    p_item_add = item_sub.add_parser("add", help="Создать/обновить статью")
    p_item_add.add_argument("--code", required=True)
    p_item_add.add_argument("--service", required=True)
    p_item_add.add_argument("--name", required=True)
    p_item_add.add_argument("--period", default="")
    p_item_add.add_argument("--sequence", type=int, default=None)
    p_item_add.add_argument("--amount", type=float, default=None)
    p_item_add.add_argument("--date-from", default="")
    p_item_add.add_argument("--date-to", default="")
    p_item_add.add_argument("--description", default="")
    p_item_add.add_argument("--comment", default="")
    p_item_add.add_argument("--operator", default="cashier_cli")
    p_item_add.add_argument("--apply", action="store_true")
    p_item_add.set_defaults(func=command_service_item_add)

    p_item_list = item_sub.add_parser("list", help="Показать статьи")
    p_item_list.add_argument("--service", default="")
    p_item_list.add_argument("--active-only", action="store_true")
    p_item_list.set_defaults(func=command_service_item_list)

    p_report = sub.add_parser("report", help="Отчёт кассира")
    p_report.add_argument("--cashbox", default="")
    p_report.add_argument("--limit", type=int, default=300)
    p_report.add_argument("--xlsx", action="store_true")
    p_report.set_defaults(func=command_report)

    args = parser.parse_args()
    args.func(args)


if __name__ == "__main__":
    main()
