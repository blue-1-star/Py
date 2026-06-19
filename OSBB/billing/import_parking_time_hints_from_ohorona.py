from pathlib import Path
import sys
import sqlite3
import re
from datetime import datetime
from collections import defaultdict

from openpyxl import load_workbook

MODULE_DIR = Path(__file__).resolve().parent
OSBB_ROOT = MODULE_DIR.parent
PY_ROOT = OSBB_ROOT.parent

for p in (OSBB_ROOT, PY_ROOT):
    if str(p) not in sys.path:
        sys.path.insert(0, str(p))

from config import paths


SOURCE_FILE = (
    getattr(paths, "OSBB_RAW_TYPED_DIR", paths.OSBB_RAW_DIR / "typed")
    / "Охорона.xlsx"
)

DAY_WORDS = ("день",)
NIGHT_WORDS = ("сутки", "сутки", "ніч", "ночь", "night")

TARIFF_RE = re.compile(
    r"(\d+)\s*[-–—]?\s*(день|сутки|ніч|ночь|night)",
    re.IGNORECASE,
)


def now_ts():
    return datetime.now().strftime("%Y-%m-%d_%H-%M-%S")


def norm_apartment(value):
    if value is None:
        return None

    text = str(value).strip()

    if not text:
        return None

    if text.endswith(".0"):
        text = text[:-2]

    if text.isdigit():
        return text

    return None


def parse_tariff_text(value):
    if value is None:
        return []

    text = str(value).strip().lower()

    result = []

    for count_text, kind in TARIFF_RE.findall(text):
        count = int(count_text)

        if kind in DAY_WORDS:
            parking_time = "Day"
        else:
            parking_time = "Night"

        result.append({
            "count": count,
            "parking_time": parking_time,
            "raw": str(value).strip(),
        })

    return result


def read_ohorona_list1():
    wb = load_workbook(SOURCE_FILE, data_only=True)

    if "Лист1" in wb.sheetnames:
        ws = wb["Лист1"]
    else:
        ws = wb[wb.sheetnames[0]]

    hints = defaultdict(lambda: {
        "Day": 0,
        "Night": 0,
        "rows": [],
    })

    current_apartment = None

    # По вашему описанию:
    # B = квартира
    # C = "1-день", "2-сутки" и т.п.
    APARTMENT_COL = 2
    TARIFF_COL = 3

    for row_idx in range(1, ws.max_row + 1):
        apt_raw = ws.cell(row=row_idx, column=APARTMENT_COL).value
        tariff_raw = ws.cell(row=row_idx, column=TARIFF_COL).value

        apt = norm_apartment(apt_raw)

        if apt:
            current_apartment = apt

        entries = parse_tariff_text(tariff_raw)

        if not entries:
            continue

        if not current_apartment:
            continue

        for item in entries:
            pt = item["parking_time"]
            cnt = item["count"]

            hints[current_apartment][pt] += cnt
            hints[current_apartment]["rows"].append({
                "excel_row": row_idx,
                "raw": item["raw"],
                "count": cnt,
                "parking_time": pt,
            })

    return hints


def load_db_vehicles():
    conn = sqlite3.connect(paths.OSBB_DB_FILE)
    cur = conn.cursor()

    cur.execute("""
        SELECT
            a.apartment_number,
            v.id,
            v.license_plate,
            v.license_plate_normalized,
            v.car_model,
            v.car_model_normalized,
            v.parking_time
        FROM vehicles v
        JOIN apartments a ON a.id = v.apartment_id
        ORDER BY
            CASE
                WHEN a.apartment_number GLOB '[0-9]*'
                THEN CAST(a.apartment_number AS INTEGER)
                ELSE 999999
            END,
            a.apartment_number,
            v.id
    """)

    rows = cur.fetchall()
    conn.close()

    vehicles = defaultdict(list)

    for (
        apartment_number,
        vehicle_id,
        plate,
        plate_norm,
        model,
        model_norm,
        parking_time,
    ) in rows:
        vehicles[str(apartment_number)].append({
            "vehicle_id": vehicle_id,
            "plate": plate,
            "plate_norm": plate_norm,
            "model": model,
            "model_norm": model_norm,
            "parking_time": parking_time,
        })

    return vehicles


def apt_sort_key(apt):
    s = str(apt)
    return (0, int(s)) if s.isdigit() else (1, s)


def vehicle_line(v):
    return (
        f"id={v['vehicle_id']} | "
        f"{v['plate_norm'] or v['plate'] or '-'} | "
        f"{v['model_norm'] or v['model'] or '-'} | "
        f"parking_time={v['parking_time'] or '-'}"
    )


def main():
    if not SOURCE_FILE.exists():
        raise FileNotFoundError(f"Файл не найден:\n{SOURCE_FILE}")

    hints = read_ohorona_list1()
    vehicles = load_db_vehicles()

    report_dir = paths.OSBB_EXPORTS_DIR / "audits"
    report_dir.mkdir(parents=True, exist_ok=True)

    report_file = report_dir / f"parking_time_hints_from_ohorona_{now_ts()}.txt"

    all_apts = sorted(
        set(hints.keys()) | set(vehicles.keys()),
        key=apt_sort_key,
    )

    auto_safe = []
    operator_needed = []
    hints_without_db = []
    db_missing_without_hint = []

    for apt in all_apts:
        h = hints.get(apt)
        vs = vehicles.get(apt, [])

        if not h:
            missing = [
                v for v in vs
                if not v["parking_time"]
            ]

            if missing:
                db_missing_without_hint.append((apt, missing))

            continue

        day_count = h["Day"]
        night_count = h["Night"]
        expected_total = day_count + night_count

        if not vs:
            hints_without_db.append((apt, h))
            continue

        missing_vehicles = [
            v for v in vs
            if not v["parking_time"]
        ]

        existing_filled = [
            v for v in vs
            if v["parking_time"]
        ]

        if (
            expected_total == len(vs)
            and len(missing_vehicles) == len(vs)
            and (
                (day_count == len(vs) and night_count == 0)
                or (night_count == len(vs) and day_count == 0)
            )
        ):
            auto_time = "Day" if day_count else "Night"
            auto_safe.append((apt, h, vs, auto_time))
        else:
            operator_needed.append((
                apt,
                h,
                vs,
                missing_vehicles,
                existing_filled,
            ))

    lines = []
    lines.append("=" * 80)
    lines.append("PARKING_TIME HINTS FROM ОХОРОНА.XLSX")
    lines.append("=" * 80)
    lines.append(f"Generated : {datetime.now():%Y-%m-%d %H:%M:%S}")
    lines.append(f"Source    : {SOURCE_FILE}")
    lines.append("")

    lines.append("=" * 80)
    lines.append("SUMMARY")
    lines.append("=" * 80)
    lines.append(f"Apartments with hints              : {len(hints)}")
    lines.append(f"Apartments with vehicles in DB      : {len(vehicles)}")
    lines.append(f"Auto-fill safe apartments           : {len(auto_safe)}")
    lines.append(f"Operator-needed apartments          : {len(operator_needed)}")
    lines.append(f"Hints without vehicles in DB        : {len(hints_without_db)}")
    lines.append(f"DB missing parking_time without hint: {len(db_missing_without_hint)}")
    lines.append("")

    lines.append("=" * 80)
    lines.append("AUTO-FILL SAFE")
    lines.append("=" * 80)
    lines.append("Эти квартиры можно заполнить автоматически: все авто одной категории Day/Night.")
    lines.append("")

    for apt, h, vs, auto_time in auto_safe:
        lines.append("-" * 80)
        lines.append(f"Apartment : {apt}")
        lines.append(f"Hint      : Day={h['Day']} | Night={h['Night']}")
        lines.append(f"Action    : set parking_time = {auto_time} for all vehicles")
        lines.append("Vehicles:")
        for v in vs:
            lines.append(f"  {vehicle_line(v)}")

    lines.append("")
    lines.append("=" * 80)
    lines.append("OPERATOR NEEDED")
    lines.append("=" * 80)
    lines.append("Здесь файл даёт количество Day/Night, но нужно назначить тариф конкретным авто.")
    lines.append("")

    for apt, h, vs, missing_vehicles, existing_filled in operator_needed:
        lines.append("-" * 80)
        lines.append(f"Apartment : {apt}")
        lines.append(f"Hint      : Day={h['Day']} | Night={h['Night']}")
        lines.append(f"DB cars   : {len(vs)}")
        lines.append(f"Missing   : {len(missing_vehicles)}")
        lines.append("Hint rows:")
        for r in h["rows"]:
            lines.append(
                f"  Excel row {r['excel_row']}: "
                f"{r['raw']} -> {r['count']} {r['parking_time']}"
            )
        lines.append("Vehicles:")
        for v in vs:
            lines.append(f"  {vehicle_line(v)}")

    lines.append("")
    lines.append("=" * 80)
    lines.append("HINTS WITHOUT VEHICLES IN DB")
    lines.append("=" * 80)

    for apt, h in hints_without_db:
        lines.append("-" * 80)
        lines.append(f"Apartment : {apt}")
        lines.append(f"Hint      : Day={h['Day']} | Night={h['Night']}")
        for r in h["rows"]:
            lines.append(
                f"  Excel row {r['excel_row']}: "
                f"{r['raw']} -> {r['count']} {r['parking_time']}"
            )

    lines.append("")
    lines.append("=" * 80)
    lines.append("DB MISSING PARKING_TIME WITHOUT HINT")
    lines.append("=" * 80)

    for apt, missing in db_missing_without_hint:
        lines.append("-" * 80)
        lines.append(f"Apartment : {apt}")
        for v in missing:
            lines.append(f"  {vehicle_line(v)}")

    report_file.write_text("\n".join(lines), encoding="utf-8")

    print("Parking time hints report created.")
    print("Report:")
    print(report_file)


if __name__ == "__main__":
    main()