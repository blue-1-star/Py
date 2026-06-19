from pathlib import Path
import sys
import sqlite3
import re
from datetime import datetime
from collections import defaultdict

from openpyxl import load_workbook

MODULE_DIR = Path(__file__).resolve().parent
OSBB_ROOT = MODULE_DIR.parent
PY_ROOT = OSBB_ROOT.parent

for p in (OSBB_ROOT, PY_ROOT):
    if str(p) not in sys.path:
        sys.path.insert(0, str(p))

from config import paths


SOURCE_FILE = (
    getattr(paths, "OSBB_RAW_TYPED_DIR", paths.OSBB_RAW_DIR / "typed")
    / "Охорона.xlsx"
)

TARIFF_RE = re.compile(
    r"(\d+)\s*[-–—]?\s*(день|сутки|ніч|ночь|night)",
    re.IGNORECASE,
)

SERVICE_CHAT_KEYWORDS = ["дима_охр", "дима охр"]


def now_ts():
    return datetime.now().strftime("%Y-%m-%d_%H-%M-%S")


def apt_sort_key(value):
    s = str(value)
    return (0, int(s)) if s.isdigit() else (1, s)


def norm_apartment(value):
    if value is None:
        return None
    text = str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]
    return text if text.isdigit() else None


def parse_tariff_text(value):
    if value is None:
        return []

    result = []
    text = str(value).strip()

    for count_text, kind in TARIFF_RE.findall(text):
        count = int(count_text)
        kind = kind.lower()

        parking_time = "Day" if "день" in kind else "Night"

        result.append({
            "count": count,
            "parking_time": parking_time,
            "raw": text,
        })

    return result


def load_ohorona_hints():
    hints = defaultdict(lambda: {
        "Day": 0,
        "Night": 0,
        "rows": [],
        "raw_unparsed": [],
    })

    if not SOURCE_FILE.exists():
        return hints

    wb = load_workbook(SOURCE_FILE, data_only=True)
    ws = wb["Лист1"] if "Лист1" in wb.sheetnames else wb[wb.sheetnames[0]]

    current_apartment = None

    APARTMENT_COL = 2
    TARIFF_COL = 3

    for row_idx in range(1, ws.max_row + 1):
        apt_raw = ws.cell(row=row_idx, column=APARTMENT_COL).value
        tariff_raw = ws.cell(row=row_idx, column=TARIFF_COL).value

        apt = norm_apartment(apt_raw)

        if apt:
            current_apartment = apt

        if tariff_raw is None:
            continue

        entries = parse_tariff_text(tariff_raw)

        if entries and current_apartment:
            for item in entries:
                pt = item["parking_time"]
                hints[current_apartment][pt] += item["count"]
                hints[current_apartment]["rows"].append({
                    "excel_row": row_idx,
                    "raw": item["raw"],
                    "count": item["count"],
                    "parking_time": pt,
                })
        else:
            text = str(tariff_raw).strip().lower()
            if current_apartment and (
                "д" in text or "сут" in text or "ніч" in text or "ноч" in text
            ):
                hints[current_apartment]["raw_unparsed"].append({
                    "excel_row": row_idx,
                    "raw": str(tariff_raw).strip(),
                })

    return hints


def load_vehicles(cur):
    cur.execute("""
        SELECT
            a.apartment_number,
            v.id,
            v.license_plate,
            v.license_plate_normalized,
            v.car_model,
            v.car_model_normalized,
            v.parking_time
        FROM vehicles v
        JOIN apartments a ON a.id = v.apartment_id
        ORDER BY
            CASE
                WHEN a.apartment_number GLOB '[0-9]*'
                THEN CAST(a.apartment_number AS INTEGER)
                ELSE 999999
            END,
            a.apartment_number,
            v.id
    """)

    result = defaultdict(list)

    for apt, vid, plate, plate_norm, model, model_norm, pt in cur.fetchall():
        result[str(apt)].append({
            "id": vid,
            "plate": plate_norm or plate,
            "model": model_norm or model,
            "parking_time": pt,
        })

    return result


def load_open_tasks(cur):
    cur.execute("""
        SELECT
            id,
            apartment_number,
            vehicle_id,
            license_plate,
            car_model,
            current_parking_time,
            task_type,
            priority
        FROM parking_time_review_tasks
        WHERE status = 'new'
    """)

    result = defaultdict(list)

    for row in cur.fetchall():
        apt = str(row[1])
        result[apt].append({
            "task_id": row[0],
            "vehicle_id": row[2],
            "license_plate": row[3],
            "car_model": row[4],
            "current_parking_time": row[5],
            "task_type": row[6],
            "priority": row[7],
        })

    return result


def load_telegram_hints():
    result = defaultdict(list)

    if not paths.OSBB_TELEGRAM_DB_FILE.exists():
        return result

    conn = sqlite3.connect(paths.OSBB_TELEGRAM_DB_FILE)
    cur = conn.cursor()

    cur.execute("""
        SELECT
            apartment_number,
            license_plate_normalized,
            person_name,
            phone_normalized,
            car_model,
            comment
        FROM telegram_facts
        WHERE fact_type = 'vehicle_info'
          AND apartment_number IS NOT NULL
    """)

    for apt, plate, person, phone, model, comment in cur.fetchall():
        c = (comment or "").lower()
        if any(k in c for k in SERVICE_CHAT_KEYWORDS):
            continue

        if plate and plate.upper().startswith(("KB", "KV")):
            continue

        result[str(apt)].append({
            "plate": plate,
            "person": person,
            "phone": phone,
            "model": model,
            "comment": comment,
        })

    conn.close()
    return result


def recommend_for_apartment(vehicles, hint):
    if not hint:
        return {
            "confidence": "NO_HINT",
            "reason": "No Охорона.xlsx hint",
            "actions": [],
        }

    day_need = hint["Day"]
    night_need = hint["Night"]
    total_need = day_need + night_need

    filled_day = [v for v in vehicles if v["parking_time"] == "Day"]
    filled_night = [v for v in vehicles if v["parking_time"] == "Night"]
    missing = [v for v in vehicles if not v["parking_time"]]

    day_missing = max(day_need - len(filled_day), 0)
    night_missing = max(night_need - len(filled_night), 0)

    actions = []

    if hint["raw_unparsed"]:
        return {
            "confidence": "LOW",
            "reason": "Охорона.xlsx has unparsed tariff text",
            "actions": [],
        }

    if total_need == len(vehicles) and len(missing) == len(vehicles):
        if day_need == len(vehicles) and night_need == 0:
            return {
                "confidence": "VERY_HIGH",
                "reason": "All vehicles missing; hint says all Day",
                "actions": [(v, "Day") for v in vehicles],
            }

        if night_need == len(vehicles) and day_need == 0:
            return {
                "confidence": "VERY_HIGH",
                "reason": "All vehicles missing; hint says all Night",
                "actions": [(v, "Night") for v in vehicles],
            }

    if len(missing) > 0:
        if day_missing == len(missing) and night_missing == 0:
            return {
                "confidence": "HIGH",
                "reason": "Existing filled values satisfy Night; all missing should be Day",
                "actions": [(v, "Day") for v in missing],
            }

        if night_missing == len(missing) and day_missing == 0:
            return {
                "confidence": "HIGH",
                "reason": "Existing filled values satisfy Day; all missing should be Night",
                "actions": [(v, "Night") for v in missing],
            }

    if total_need != len(vehicles):
        return {
            "confidence": "MEDIUM",
            "reason": (
                f"Hint total={total_need}, DB vehicles={len(vehicles)}; "
                "operator must verify vehicle list"
            ),
            "actions": [],
        }

    if len(missing) > 0:
        return {
            "confidence": "MEDIUM",
            "reason": (
                f"Hint Day={day_need}, Night={night_need}; "
                f"missing vehicles={len(missing)}; distribution requires operator"
            ),
            "actions": [],
        }

    return {
        "confidence": "CHECK",
        "reason": "No missing parking_time, but hint exists for comparison",
        "actions": [],
    }


def confidence_rank(conf):
    return {
        "VERY_HIGH": 1,
        "HIGH": 2,
        "MEDIUM": 3,
        "LOW": 4,
        "NO_HINT": 5,
        "CHECK": 6,
    }.get(conf, 9)


def main():
    conn = sqlite3.connect(paths.OSBB_DB_FILE)
    cur = conn.cursor()

    hints = load_ohorona_hints()
    vehicles_by_apt = load_vehicles(cur)
    tasks_by_apt = load_open_tasks(cur)
    telegram_by_apt = load_telegram_hints()

    target_apts = sorted(
        set(tasks_by_apt.keys()) | set(hints.keys()),
        key=apt_sort_key,
    )

    blocks = []

    for apt in target_apts:
        vehicles = vehicles_by_apt.get(apt, [])
        hint = hints.get(apt)
        tasks = tasks_by_apt.get(apt, [])
        telegram = telegram_by_apt.get(apt, [])

        rec = recommend_for_apartment(vehicles, hint)

        if not tasks and rec["confidence"] == "CHECK":
            continue

        blocks.append({
            "apt": apt,
            "vehicles": vehicles,
            "hint": hint,
            "tasks": tasks,
            "telegram": telegram,
            "recommendation": rec,
        })

    blocks.sort(
        key=lambda b: (
            confidence_rank(b["recommendation"]["confidence"]),
            apt_sort_key(b["apt"]),
        )
    )

    report_dir = paths.OSBB_EXPORTS_DIR / "audits"
    report_dir.mkdir(parents=True, exist_ok=True)

    report_file = report_dir / f"parking_time_with_hints_{now_ts()}.txt"

    lines = []
    lines.append("=" * 80)
    lines.append("PARKING_TIME WITH HINTS REPORT")
    lines.append("=" * 80)
    lines.append(f"Generated: {datetime.now():%Y-%m-%d %H:%M:%S}")
    lines.append(f"Source   : {SOURCE_FILE}")
    lines.append("")

    counts = defaultdict(int)
    for b in blocks:
        counts[b["recommendation"]["confidence"]] += 1

    lines.append("=" * 80)
    lines.append("SUMMARY")
    lines.append("=" * 80)
    for key in ["VERY_HIGH", "HIGH", "MEDIUM", "LOW", "NO_HINT", "CHECK"]:
        lines.append(f"{key:10}: {counts.get(key, 0)}")
    lines.append("")

    for b in blocks:
        apt = b["apt"]
        hint = b["hint"]
        rec = b["recommendation"]

        lines.append("=" * 80)
        lines.append(f"APARTMENT {apt} | {rec['confidence']}")
        lines.append("=" * 80)

        lines.append("Vehicles:")
        for v in b["vehicles"]:
            lines.append(
                f"  id={v['id']} | {v['plate'] or '-'} | "
                f"{v['model'] or '-'} | parking_time={v['parking_time'] or '-'}"
            )

        lines.append("")
        lines.append("Open tasks:")
        if b["tasks"]:
            for t in b["tasks"]:
                lines.append(
                    f"  Task {t['task_id']} | vehicle_id={t['vehicle_id']} | "
                    f"{t['license_plate'] or '-'} | {t['car_model'] or '-'}"
                )
        else:
            lines.append("  -")

        lines.append("")
        lines.append("Охорона.xlsx hint:")
        if hint:
            lines.append(f"  Day={hint['Day']} | Night={hint['Night']}")
            for r in hint["rows"]:
                lines.append(
                    f"  Excel row {r['excel_row']}: "
                    f"{r['raw']} -> {r['count']} {r['parking_time']}"
                )
            for r in hint["raw_unparsed"]:
                lines.append(
                    f"  UNPARSED row {r['excel_row']}: {r['raw']}"
                )
        else:
            lines.append("  -")

        lines.append("")
        lines.append("Telegram hints:")
        if b["telegram"]:
            seen = set()
            for tg in b["telegram"]:
                key = (tg["plate"], tg["person"], tg["phone"], tg["model"])
                if key in seen:
                    continue
                seen.add(key)

                lines.append(
                    f"  plate={tg['plate'] or '-'} | "
                    f"person={tg['person'] or '-'} | "
                    f"phone={tg['phone'] or '-'} | "
                    f"model={tg['model'] or '-'}"
                )
        else:
            lines.append("  -")

        lines.append("")
        lines.append("Recommendation:")
        lines.append(f"  Confidence: {rec['confidence']}")
        lines.append(f"  Reason    : {rec['reason']}")

        if rec["actions"]:
            lines.append("  Proposed actions:")
            for v, new_value in rec["actions"]:
                lines.append(
                    f"    vehicle_id={v['id']} | "
                    f"{v['plate'] or '-'} -> {new_value}"
                )
        else:
            lines.append("  Proposed actions: -")

        lines.append("")
        lines.append("Operator options:")
        lines.append("  [Apply proposed] [Set Day] [Set Night] [Ask owner] [Skip]")
        lines.append("")

    report_file.write_text("\n".join(lines), encoding="utf-8")

    conn.close()

    print("Parking time with hints report created.")
    print("Report:")
    print(report_file)


if __name__ == "__main__":
    main()