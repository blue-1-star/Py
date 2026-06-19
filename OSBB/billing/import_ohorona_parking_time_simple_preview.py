from pathlib import Path
import sys
import sqlite3
import re
import argparse
from datetime import datetime
from collections import defaultdict

from openpyxl import load_workbook


MODULE_DIR = Path(__file__).resolve().parent
OSBB_ROOT = MODULE_DIR.parent
PY_ROOT = OSBB_ROOT.parent

for p in (OSBB_ROOT, PY_ROOT):
    if str(p) not in sys.path:
        sys.path.insert(0, str(p))

from config import paths


SOURCE_FILE = (
    getattr(paths, "OSBB_RAW_TYPED_DIR", paths.OSBB_RAW_DIR / "typed")
    / "Охорона.xlsx"
)

SHEET_NAME = "Лист1"
APARTMENT_COL = 2  # B
TARIFF_COL = 3     # C

TARIFF_RE = re.compile(
    r"(\d+)\s*[-–—]?\s*(день|сутки|ніч|ночь|night)",
    re.IGNORECASE,
)

DAY_WORDS = {"день"}
NIGHT_WORDS = {"сутки", "ніч", "ночь", "night"}


def now_ts() -> str:
    return datetime.now().strftime("%Y-%m-%d_%H-%M-%S")


def norm_apartment(value):
    if value is None:
        return None

    text = str(value).strip()
    if not text:
        return None

    if text.endswith(".0"):
        text = text[:-2]

    return text if text.isdigit() else None


def parse_tariff_cell(value):
    """
    Расшифровка:
    1-день  -> 1 авто Day
    2-день  -> 2 авто Day
    1-сутки -> 1 авто Night
    2-сутки -> 2 авто Night

    В одной ячейке допускается несколько фрагментов, например:
    1-день, 1-сутки
    """
    if value is None:
        return []

    raw = str(value).strip()
    if not raw:
        return []

    result = []

    for count_text, kind in TARIFF_RE.findall(raw):
        count = int(count_text)
        kind_norm = kind.lower()

        if kind_norm in DAY_WORDS:
            parking_time = "Day"
        elif kind_norm in NIGHT_WORDS:
            parking_time = "Night"
        else:
            parking_time = None

        if parking_time:
            result.append({
                "count": count,
                "parking_time": parking_time,
                "raw": raw,
            })

    return result


def read_ohorona_hints(source_file: Path):
    """
    Читает B/C листа Лист1.
    Если B пустая, используется последняя найденная квартира выше.
    Это как раз покрывает строки 5-6, 13-14, 17-18 и подобные случаи.
    """
    wb = load_workbook(source_file, data_only=True)

    if SHEET_NAME in wb.sheetnames:
        ws = wb[SHEET_NAME]
    else:
        ws = wb[wb.sheetnames[0]]

    hints = defaultdict(lambda: {
        "Day": 0,
        "Night": 0,
        "rows": [],
    })

    current_apartment = None

    for row_idx in range(1, ws.max_row + 1):
        apt_raw = ws.cell(row=row_idx, column=APARTMENT_COL).value
        tariff_raw = ws.cell(row=row_idx, column=TARIFF_COL).value

        apt = norm_apartment(apt_raw)
        if apt:
            current_apartment = apt

        entries = parse_tariff_cell(tariff_raw)
        if not entries:
            continue

        if not current_apartment:
            continue

        for item in entries:
            pt = item["parking_time"]
            cnt = item["count"]

            hints[current_apartment][pt] += cnt
            hints[current_apartment]["rows"].append({
                "excel_row": row_idx,
                "excel_apartment_cell": apt_raw,
                "effective_apartment": current_apartment,
                "raw": item["raw"],
                "count": cnt,
                "parking_time": pt,
            })

    return hints


def load_db_vehicles():
    conn = sqlite3.connect(paths.OSBB_DB_FILE)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    cur.execute("""
        SELECT
            a.apartment_number,
            v.id AS vehicle_id,
            v.license_plate,
            v.license_plate_normalized,
            v.car_model,
            v.car_model_normalized,
            v.parking_time
        FROM vehicles v
        JOIN apartments a ON a.id = v.apartment_id
        ORDER BY
            CASE
                WHEN a.apartment_number GLOB '[0-9]*'
                THEN CAST(a.apartment_number AS INTEGER)
                ELSE 999999
            END,
            a.apartment_number,
            v.id
    """)

    vehicles = defaultdict(list)

    for row in cur.fetchall():
        apt = str(row["apartment_number"])
        vehicles[apt].append(dict(row))

    conn.close()
    return vehicles


def apt_sort_key(apt):
    s = str(apt)
    return (0, int(s)) if s.isdigit() else (1, s)


def vehicle_label(v):
    plate = v.get("license_plate_normalized") or v.get("license_plate") or "-"
    model = v.get("car_model_normalized") or v.get("car_model") or "-"
    pt = v.get("parking_time") or "-"
    return f"id={v['vehicle_id']} | {plate} | {model} | parking_time={pt}"


def classify_apartments(hints, vehicles):
    auto_safe = []
    operator_needed = []
    hints_without_db = []
    db_missing_without_hint = []

    all_apts = sorted(set(hints.keys()) | set(vehicles.keys()), key=apt_sort_key)

    for apt in all_apts:
        h = hints.get(apt)
        vs = vehicles.get(apt, [])

        if not h:
            missing = [v for v in vs if not v.get("parking_time")]
            if missing:
                db_missing_without_hint.append((apt, missing))
            continue

        day_count = h["Day"]
        night_count = h["Night"]
        expected_total = day_count + night_count

        if not vs:
            hints_without_db.append((apt, h))
            continue

        missing_vehicles = [v for v in vs if not v.get("parking_time")]
        existing_filled = [v for v in vs if v.get("parking_time")]

        # Безопасное автозаполнение:
        # 1) количество тарифов совпало с количеством авто в базе;
        # 2) у всех этих авто parking_time пустой;
        # 3) все авто одной категории: только Day или только Night.
        if (
            expected_total == len(vs)
            and len(missing_vehicles) == len(vs)
            and (
                (day_count == len(vs) and night_count == 0)
                or (night_count == len(vs) and day_count == 0)
            )
        ):
            auto_time = "Day" if day_count else "Night"
            auto_safe.append((apt, h, vs, auto_time))
        else:
            operator_needed.append((apt, h, vs, missing_vehicles, existing_filled))

    return auto_safe, operator_needed, hints_without_db, db_missing_without_hint


def format_tariff_hint_short(h):
    parts = []

    if h["Day"]:
        parts.append(f"{h['Day']} Day")

    if h["Night"]:
        parts.append(f"{h['Night']} Night")

    return " + ".join(parts) if parts else "-"


def write_simple_preview_section(lines, hints, vehicles):
    lines.append("=" * 90)
    lines.append("SIMPLE CHECK TABLE")
    lines.append("=" * 90)
    lines.append("Квартира | Номер авто | Тариф из Охорона")
    lines.append("-" * 90)

    for apt in sorted(hints.keys(), key=apt_sort_key):
        h = hints[apt]
        tariff_text = format_tariff_hint_short(h)
        vs = vehicles.get(apt, [])

        if not vs:
            lines.append(f"{apt} | - | {tariff_text}")
            continue

        for v in vs:
            plate = v.get("license_plate_normalized") or v.get("license_plate") or "-"
            lines.append(f"{apt} | {plate} | {tariff_text}")

    lines.append("")


def write_report(hints, vehicles, classified, report_file: Path):
    auto_safe, operator_needed, hints_without_db, db_missing_without_hint = classified

    lines = []
    write_simple_preview_section(lines, hints, vehicles)

    lines.append("=" * 90)
    lines.append("PARKING_TIME PARSE REPORT FROM ОХОРОНА.XLSX")
    lines.append("=" * 90)
    lines.append(f"Generated : {datetime.now():%Y-%m-%d %H:%M:%S}")
    lines.append(f"Source    : {SOURCE_FILE}")
    lines.append(f"Sheet     : {SHEET_NAME}")
    lines.append("Columns   : B = apartment, C = tariff text")
    lines.append("Rule      : empty B cell means previous apartment")
    lines.append("")

    lines.append("=" * 90)
    lines.append("SUMMARY")
    lines.append("=" * 90)
    lines.append(f"Apartments with tariff hints        : {len(hints)}")
    lines.append(f"Apartments with vehicles in DB      : {len(vehicles)}")
    lines.append(f"Auto-fill safe apartments           : {len(auto_safe)}")
    lines.append(f"Operator-needed apartments          : {len(operator_needed)}")
    lines.append(f"Hints without vehicles in DB        : {len(hints_without_db)}")
    lines.append(f"DB missing parking_time without hint: {len(db_missing_without_hint)}")
    lines.append("")

    lines.append("=" * 90)
    lines.append("PARSED HINTS BY APARTMENT")
    lines.append("=" * 90)
    for apt in sorted(hints.keys(), key=apt_sort_key):
        h = hints[apt]
        lines.append("-" * 90)
        lines.append(f"Apartment : {apt}")
        lines.append(f"Decoded   : Day={h['Day']} | Night={h['Night']} | Total={h['Day'] + h['Night']}")
        for r in h["rows"]:
            apt_cell = r["excel_apartment_cell"]
            apt_cell_text = "" if apt_cell is None else str(apt_cell)
            lines.append(
                f"  row {r['excel_row']:>4}: "
                f"B='{apt_cell_text}' -> apt {r['effective_apartment']} | "
                f"C='{r['raw']}' -> {r['count']} {r['parking_time']}"
            )
    lines.append("")

    lines.append("=" * 90)
    lines.append("AUTO-FILL SAFE")
    lines.append("=" * 90)
    lines.append("Эти квартиры можно заполнить автоматически: все авто одной категории Day/Night.")
    lines.append("")
    for apt, h, vs, auto_time in auto_safe:
        lines.append("-" * 90)
        lines.append(f"Apartment : {apt}")
        lines.append(f"Hint      : Day={h['Day']} | Night={h['Night']}")
        lines.append(f"Action    : set parking_time = {auto_time} for all vehicles of this apartment")
        lines.append("Vehicles:")
        for v in vs:
            lines.append(f"  {vehicle_label(v)}")
    lines.append("")

    lines.append("=" * 90)
    lines.append("OPERATOR NEEDED")
    lines.append("=" * 90)
    lines.append("Здесь нужно назначить Day/Night конкретным авто или проверить конфликт.")
    lines.append("")
    for apt, h, vs, missing_vehicles, existing_filled in operator_needed:
        lines.append("-" * 90)
        lines.append(f"Apartment : {apt}")
        lines.append(f"Hint      : Day={h['Day']} | Night={h['Night']}")
        lines.append(f"DB cars   : {len(vs)}")
        lines.append(f"Missing   : {len(missing_vehicles)}")
        lines.append(f"Filled    : {len(existing_filled)}")
        lines.append("Hint rows:")
        for r in h["rows"]:
            lines.append(
                f"  row {r['excel_row']:>4}: "
                f"B='{r['excel_apartment_cell'] or ''}' -> apt {r['effective_apartment']} | "
                f"C='{r['raw']}' -> {r['count']} {r['parking_time']}"
            )
        lines.append("Vehicles:")
        for v in vs:
            lines.append(f"  {vehicle_label(v)}")
    lines.append("")

    lines.append("=" * 90)
    lines.append("HINTS WITHOUT VEHICLES IN DB")
    lines.append("=" * 90)
    for apt, h in hints_without_db:
        lines.append("-" * 90)
        lines.append(f"Apartment : {apt}")
        lines.append(f"Hint      : Day={h['Day']} | Night={h['Night']}")
        for r in h["rows"]:
            lines.append(
                f"  row {r['excel_row']:>4}: "
                f"C='{r['raw']}' -> {r['count']} {r['parking_time']}"
            )
    lines.append("")

    lines.append("=" * 90)
    lines.append("DB MISSING PARKING_TIME WITHOUT HINT")
    lines.append("=" * 90)
    for apt, missing in db_missing_without_hint:
        lines.append("-" * 90)
        lines.append(f"Apartment : {apt}")
        for v in missing:
            lines.append(f"  {vehicle_label(v)}")

    report_file.write_text("\n".join(lines), encoding="utf-8")


def apply_auto_safe(auto_safe):
    """
    В базу записываются только AUTO-FILL SAFE случаи.
    Неоднозначные квартиры НЕ трогаются.
    """
    conn = sqlite3.connect(paths.OSBB_DB_FILE)
    cur = conn.cursor()

    updated = 0

    for apt, h, vs, auto_time in auto_safe:
        for v in vs:
            cur.execute(
                """
                UPDATE vehicles
                SET parking_time = ?
                WHERE id = ?
                  AND (parking_time IS NULL OR TRIM(parking_time) = '')
                """,
                (auto_time, v["vehicle_id"]),
            )
            updated += cur.rowcount

    conn.commit()
    conn.close()
    return updated


def main():
    parser = argparse.ArgumentParser(
        description="Parse Охорона.xlsx parking tariffs and optionally apply safe parking_time updates."
    )
    parser.add_argument(
        "--apply",
        action="store_true",
        help="Apply only AUTO-FILL SAFE parking_time updates to SQLite. Without this flag only report is created.",
    )
    parser.add_argument(
        "--source",
        type=str,
        default=str(SOURCE_FILE),
        help="Path to Охорона.xlsx. Default: configured raw/typed/Охорона.xlsx",
    )
    args = parser.parse_args()

    source_file = Path(args.source)

    if not source_file.exists():
        raise FileNotFoundError(f"Файл не найден:\n{source_file}")

    hints = read_ohorona_hints(source_file)
    vehicles = load_db_vehicles()
    classified = classify_apartments(hints, vehicles)

    report_dir = paths.OSBB_EXPORTS_DIR / "audits"
    report_dir.mkdir(parents=True, exist_ok=True)
    report_file = report_dir / f"parking_time_parse_from_ohorona_{now_ts()}.txt"

    write_report(hints, vehicles, classified, report_file)

    print("Report created:")
    print(report_file)

    if args.apply:
        auto_safe, *_ = classified
        updated = apply_auto_safe(auto_safe)
        print()
        print("APPLY MODE ENABLED")
        print(f"Updated vehicles: {updated}")
    else:
        print()
        print("DRY RUN ONLY. Database was not changed.")
        print("To apply only AUTO-FILL SAFE updates, run with --apply")


if __name__ == "__main__":
    main()
