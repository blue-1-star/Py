from pathlib import Path
import sys
import sqlite3
import argparse
from datetime import datetime, timedelta
from collections import Counter

OSBB_ROOT = Path(__file__).resolve().parent
PY_ROOT = OSBB_ROOT.parent

if str(PY_ROOT) not in sys.path:
    sys.path.insert(0, str(PY_ROOT))

from config import paths, USE_TEST_DB

try:
    from plate_consensus_report import DEFAULT_PERIOD_CODE, build_consensus
except Exception:
    DEFAULT_PERIOD_CODE = "2026-05_2026-06"
    build_consensus = None


def get_db_file():
    return paths.OSBB_TEST_DB_FILE if USE_TEST_DB else paths.OSBB_DB_FILE


def get_quarantine_db_file():
    return getattr(paths, "OSBB_QUARANTINE_DB_FILE", paths.OSBB_DB_DIR / "osbb_quarantine.db")


def now_ts():
    return datetime.now().strftime("%Y-%m-%d_%H-%M-%S")


def money(value):
    value = float(value or 0)
    return str(int(value)) if value.is_integer() else f"{value:.2f}"


def table_exists(cur, table_name):
    cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table_name,))
    return cur.fetchone() is not None


def table_columns(cur, table_name):
    cur.execute(f"PRAGMA table_info({table_name})")
    return [row[1] for row in cur.fetchall()]


def pick_table(cur, *names):
    for name in names:
        if table_exists(cur, name):
            return name
    return None


def count_table(cur, table_name):
    if not table_exists(cur, table_name):
        return None
    cur.execute(f"SELECT COUNT(*) FROM {table_name}")
    return cur.fetchone()[0]


def ensure_audit_log(cur):
    """
    Safe idempotent migration for the WORK/TEST database.
    Does not destroy old data.
    """
    cur.execute("""
        CREATE TABLE IF NOT EXISTS operator_audit_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            created_at TEXT,
            operator_id TEXT,
            user_id TEXT,
            actor_type TEXT,
            action_type TEXT,
            table_name TEXT,
            row_id TEXT,
            field_name TEXT,
            old_value TEXT,
            new_value TEXT,
            action_status TEXT DEFAULT 'applied',
            review_status TEXT DEFAULT 'pending',
            source_context TEXT,
            comment TEXT
        )
    """)

    cols = table_columns(cur, "operator_audit_log")

    migrations = {
        "created_at": "TEXT",
        "operator_id": "TEXT",
        "user_id": "TEXT",
        "actor_type": "TEXT",
        "action_type": "TEXT",
        "table_name": "TEXT",
        "row_id": "TEXT",
        "field_name": "TEXT",
        "old_value": "TEXT",
        "new_value": "TEXT",
        "action_status": "TEXT DEFAULT 'applied'",
        "review_status": "TEXT DEFAULT 'pending'",
        "source_context": "TEXT",
        "comment": "TEXT",
        "db_mode": "TEXT",
        "db_file": "TEXT",
        "reviewed_by": "TEXT",
        "reviewed_at": "TEXT",
        "review_comment": "TEXT",
        "extra_json": "TEXT",
    }

    for col, col_type in migrations.items():
        if col not in cols:
            cur.execute(f"ALTER TABLE operator_audit_log ADD COLUMN {col} {col_type}")

    cur.execute("CREATE INDEX IF NOT EXISTS idx_operator_audit_created ON operator_audit_log(created_at)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_operator_audit_operator ON operator_audit_log(operator_id)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_operator_audit_actor ON operator_audit_log(actor_type)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_operator_audit_action ON operator_audit_log(action_type)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_operator_audit_review ON operator_audit_log(review_status)")


def load_recent_audit(cur, limit=50, operator_id="", actor_type="", days=0):
    if not table_exists(cur, "operator_audit_log"):
        return []

    cols = table_columns(cur, "operator_audit_log")
    id_col = "id" if "id" in cols else "rowid"
    created_col = "created_at" if "created_at" in cols else None
    operator_col = "operator_id" if "operator_id" in cols else ("user_id" if "user_id" in cols else None)
    actor_col = "actor_type" if "actor_type" in cols else None

    select_parts = [
        f"{id_col} AS id",
        f"{created_col if created_col else 'NULL'} AS created_at",
        f"{operator_col if operator_col else 'NULL'} AS operator_id",
        f"{actor_col if actor_col else 'NULL'} AS actor_type",
        "action_type" if "action_type" in cols else "NULL AS action_type",
        "table_name" if "table_name" in cols else "NULL AS table_name",
        "row_id" if "row_id" in cols else "NULL AS row_id",
        "field_name" if "field_name" in cols else "NULL AS field_name",
        "old_value" if "old_value" in cols else "NULL AS old_value",
        "new_value" if "new_value" in cols else "NULL AS new_value",
        "review_status" if "review_status" in cols else "NULL AS review_status",
        "action_status" if "action_status" in cols else "NULL AS action_status",
        "source_context" if "source_context" in cols else "NULL AS source_context",
        "comment" if "comment" in cols else "NULL AS comment",
    ]

    where = []
    params = []

    if operator_id and operator_col:
        where.append(f"{operator_col} = ?")
        params.append(str(operator_id))

    if actor_type and actor_col:
        where.append(f"{actor_col} = ?")
        params.append(str(actor_type))

    if days and created_col:
        since = (datetime.now() - timedelta(days=days)).strftime("%Y-%m-%d %H:%M:%S")
        where.append(f"{created_col} >= ?")
        params.append(since)

    where_sql = "WHERE " + " AND ".join(where) if where else ""
    order_col = created_col if created_col else id_col

    params.append(limit)

    cur.execute(f"""
        SELECT {", ".join(select_parts)}
        FROM operator_audit_log
        {where_sql}
        ORDER BY {order_col} DESC, {id_col} DESC
        LIMIT ?
    """, tuple(params))

    return [dict(row) for row in cur.fetchall()]


def audit_summary(cur):
    rows = load_recent_audit(cur, limit=100000)
    by_actor = Counter(row.get("actor_type") or "-" for row in rows)
    by_operator = Counter(row.get("operator_id") or "-" for row in rows)
    by_action = Counter(row.get("action_type") or "-" for row in rows)
    by_review = Counter(row.get("review_status") or "-" for row in rows)

    return {
        "total": len(rows),
        "by_actor": by_actor,
        "by_operator": by_operator,
        "by_action": by_action,
        "by_review": by_review,
    }


def vehicle_quality_summary(cur):
    total = count_table(cur, "vehicles") or 0
    no_tariff = 0
    suspicious_plate = 0

    if not table_exists(cur, "vehicles"):
        return {"total": 0, "no_tariff": 0, "suspicious_plate": 0}

    cols = table_columns(cur, "vehicles")

    if "parking_time" in cols:
        cur.execute("""
            SELECT COUNT(*)
            FROM vehicles
            WHERE parking_time IS NULL OR TRIM(CAST(parking_time AS TEXT)) = ''
        """)
        no_tariff = cur.fetchone()[0]

    plate_col = "license_plate_normalized" if "license_plate_normalized" in cols else (
        "license_plate" if "license_plate" in cols else None
    )

    if plate_col:
        cur.execute(f"SELECT {plate_col} FROM vehicles")
        for (plate,) in cur.fetchall():
            text = "" if plate is None else str(plate).strip().upper()
            digits = "".join(ch for ch in text if ch.isdigit())
            if len(digits) != 4 or len(text) < 6:
                suspicious_plate += 1

    return {"total": total, "no_tariff": no_tariff, "suspicious_plate": suspicious_plate}


def billing_summary(cur, period_code):
    charges_table = pick_table(cur, "charges", "service_charges")
    payments_table = pick_table(cur, "payments", "service_payments")
    allocations_table = pick_table(cur, "payment_allocations", "service_payment_allocations")

    result = {
        "charges_table": charges_table,
        "payments_table": payments_table,
        "allocations_table": allocations_table,
        "charges_count": 0,
        "payments_count": 0,
        "charges_total": 0,
        "payments_total": 0,
        "unallocated_payments": 0,
    }

    if charges_table:
        cur.execute(f"SELECT COUNT(*), COALESCE(SUM(amount), 0) FROM {charges_table} WHERE period_code = ?", (period_code,))
        result["charges_count"], result["charges_total"] = cur.fetchone()

    if payments_table:
        cur.execute(f"SELECT COUNT(*), COALESCE(SUM(amount), 0) FROM {payments_table} WHERE period_code = ?", (period_code,))
        result["payments_count"], result["payments_total"] = cur.fetchone()

    if payments_table and allocations_table:
        cur.execute(f"""
            SELECT
                COALESCE(SUM(p.amount), 0)
                -
                COALESCE((
                    SELECT SUM(pa.amount)
                    FROM {allocations_table} pa
                    JOIN {payments_table} p2 ON p2.id = pa.payment_id
                    WHERE p2.period_code = ?
                ), 0)
            FROM {payments_table} p
            WHERE p.period_code = ?
        """, (period_code, period_code))
        result["unallocated_payments"] = cur.fetchone()[0] or 0

    return result


def consensus_summary(period_code):
    if build_consensus is None:
        return {"available": False, "groups": 0, "majority": 0, "tie": 0, "single": 0}

    try:
        items = build_consensus(period_code=period_code, include_single=False)
    except Exception as exc:
        return {"available": False, "error": str(exc), "groups": 0, "majority": 0, "tie": 0, "single": 0}

    counts = Counter(item.get("status") for item in items)

    return {
        "available": True,
        "groups": len(items),
        "majority": counts.get("majority", 0),
        "tie": counts.get("tie", 0),
        "single": counts.get("single_variant", 0),
    }


def collect_dashboard_data(args):
    db_file = get_db_file()
    conn = sqlite3.connect(db_file)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    if not args.no_ensure_audit:
        ensure_audit_log(cur)
        conn.commit()

    table_counts = {}
    for table in [
        "apartments", "contacts", "vehicles", "operator_audit_log",
        "service_catalog", "service_tariffs", "charges", "payments", "payment_allocations",
    ]:
        table_counts[table] = count_table(cur, table)

    vehicle_q = vehicle_quality_summary(cur)
    billing = billing_summary(cur, args.period)
    audit = audit_summary(cur)
    recent_rows = load_recent_audit(
        cur,
        limit=args.limit,
        operator_id=args.operator,
        actor_type=args.actor_type,
        days=args.days,
    )
    consensus = consensus_summary(args.period)

    conn.close()

    return {
        "db_file": db_file,
        "qdb_file": get_quarantine_db_file(),
        "mode": "TEST/WORK" if USE_TEST_DB else "PROD",
        "period": args.period,
        "table_counts": table_counts,
        "vehicle_q": vehicle_q,
        "billing": billing,
        "audit": audit,
        "recent_rows": recent_rows,
        "consensus": consensus,
        "filters": {
            "operator": args.operator,
            "actor_type": args.actor_type,
            "days": args.days,
            "limit": args.limit,
        },
    }


def write_txt_report(report_file, data):
    lines = []
    lines.append("=" * 140)
    lines.append("OSBB SUPERVISOR DASHBOARD")
    lines.append("=" * 140)
    lines.append(f"Generated : {datetime.now():%Y-%m-%d %H:%M:%S}")
    lines.append(f"DB        : {data['db_file']}")
    lines.append(f"QDB       : {data['qdb_file']}")
    lines.append(f"MODE      : {data['mode']}")
    lines.append(f"Period    : {data['period']}")
    lines.append(f"Filter operator : {data['filters']['operator'] or 'ALL'}")
    lines.append(f"Filter actor    : {data['filters']['actor_type'] or 'ALL'}")
    lines.append(f"Filter days     : {data['filters']['days'] or 'ALL'}")
    lines.append(f"Limit actions   : {data['filters']['limit']}")
    lines.append("")

    lines.append("=" * 140)
    lines.append("1. DATABASE STATE")
    lines.append("=" * 140)
    for table, count in data["table_counts"].items():
        lines.append(f"{table:24}: {'-' if count is None else count}")
    lines.append("")

    lines.append("=" * 140)
    lines.append("2. VEHICLE QUALITY")
    lines.append("=" * 140)
    lines.append(f"Vehicles total          : {data['vehicle_q']['total']}")
    lines.append(f"Vehicles without tariff : {data['vehicle_q']['no_tariff']}")
    lines.append(f"Suspicious plate rows   : {data['vehicle_q']['suspicious_plate']}")
    lines.append("")

    lines.append("=" * 140)
    lines.append("3. BILLING STATE")
    lines.append("=" * 140)
    b = data["billing"]
    lines.append(f"Charges table           : {b['charges_table']}")
    lines.append(f"Payments table          : {b['payments_table']}")
    lines.append(f"Allocations table       : {b['allocations_table']}")
    lines.append(f"Charges count           : {b['charges_count']}")
    lines.append(f"Payments count          : {b['payments_count']}")
    lines.append(f"Charges total           : {money(b['charges_total'])}")
    lines.append(f"Payments total          : {money(b['payments_total'])}")
    lines.append(f"Unallocated payments    : {money(b['unallocated_payments'])}")
    lines.append("")

    lines.append("=" * 140)
    lines.append("4. PLATE CONSENSUS")
    lines.append("=" * 140)
    c = data["consensus"]
    if c.get("available"):
        lines.append(f"Consensus groups        : {c['groups']}")
        lines.append(f"Majority                : {c['majority']}")
        lines.append(f"Tie                     : {c['tie']}")
        lines.append(f"Single                  : {c['single']}")
    else:
        lines.append(f"Consensus unavailable   : {c.get('error', '')}")
    lines.append("")

    lines.append("=" * 140)
    lines.append("5. AUDIT SUMMARY")
    lines.append("=" * 140)
    a = data["audit"]
    lines.append(f"Audit rows total        : {a['total']}")
    if a['total'] == 0:
        lines.append("WARNING                 : audit log is empty. Earlier actions cannot be reconstructed if they were not logged.")
    lines.append("")
    lines.append("By actor type:")
    for key, value in a["by_actor"].most_common(20):
        lines.append(f"  {key}: {value}")
    lines.append("")
    lines.append("By operator/admin/system:")
    for key, value in a["by_operator"].most_common(20):
        lines.append(f"  {key}: {value}")
    lines.append("")
    lines.append("By action:")
    for key, value in a["by_action"].most_common(20):
        lines.append(f"  {key}: {value}")
    lines.append("")
    lines.append("By review status:")
    for key, value in a["by_review"].most_common(20):
        lines.append(f"  {key}: {value}")
    lines.append("")

    lines.append("=" * 140)
    lines.append("6. RECENT ACTIONS")
    lines.append("=" * 140)
    if not data["recent_rows"]:
        lines.append("нет записей")
    else:
        lines.append("id | created_at | actor | operator | action | table | row | field | old -> new | review")
        lines.append("-" * 140)
        for row in data["recent_rows"]:
            lines.append(
                f"{row.get('id')} | {row.get('created_at') or '-'} | "
                f"{row.get('actor_type') or '-'} | {row.get('operator_id') or '-'} | "
                f"{row.get('action_type') or '-'} | {row.get('table_name') or '-'} | "
                f"{row.get('row_id') or '-'} | {row.get('field_name') or '-'} | "
                f"{row.get('old_value') or ''} -> {row.get('new_value') or ''} | "
                f"{row.get('review_status') or '-'}"
            )

    report_file.write_text("\n".join(lines), encoding="utf-8")


def write_xlsx_report(xlsx_file, data):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError(
            "Для Excel-выгрузки нужен пакет openpyxl. Установите: pip install openpyxl"
        ) from exc

    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"

    fill_title = PatternFill("solid", fgColor="1F4E78")
    fill_header = PatternFill("solid", fgColor="D9EAF7")
    white = Font(color="FFFFFF", bold=True, size=14)
    bold = Font(bold=True)
    border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    ws["A1"] = "OSBB Supervisor Dashboard"
    ws["A1"].fill = fill_title
    ws["A1"].font = white
    ws.merge_cells("A1:D1")

    meta = [
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("DB", str(data["db_file"])),
        ("QDB", str(data["qdb_file"])),
        ("MODE", data["mode"]),
        ("Period", data["period"]),
        ("Filter operator", data["filters"]["operator"] or "ALL"),
        ("Filter actor", data["filters"]["actor_type"] or "ALL"),
        ("Filter days", data["filters"]["days"] or "ALL"),
        ("Limit actions", data["filters"]["limit"]),
    ]

    r = 3
    for k, v in meta:
        ws.cell(r, 1).value = k
        ws.cell(r, 1).font = bold
        ws.cell(r, 2).value = v
        r += 1

    def section(title):
        nonlocal r
        r += 1
        ws.cell(r, 1).value = title
        ws.cell(r, 1).fill = fill_header
        ws.cell(r, 1).font = bold
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        r += 1

    def kv(key, value):
        nonlocal r
        ws.cell(r, 1).value = key
        ws.cell(r, 2).value = value
        r += 1

    section("Database State")
    for table, count in data["table_counts"].items():
        kv(table, "" if count is None else count)

    section("Vehicle Quality")
    kv("Vehicles total", data["vehicle_q"]["total"])
    kv("Vehicles without tariff", data["vehicle_q"]["no_tariff"])
    kv("Suspicious plate rows", data["vehicle_q"]["suspicious_plate"])

    section("Billing State")
    b = data["billing"]
    kv("Charges count", b["charges_count"])
    kv("Payments count", b["payments_count"])
    kv("Charges total", b["charges_total"])
    kv("Payments total", b["payments_total"])
    kv("Unallocated payments", b["unallocated_payments"])

    section("Plate Consensus")
    c = data["consensus"]
    kv("Consensus groups", c.get("groups", 0))
    kv("Majority", c.get("majority", 0))
    kv("Tie", c.get("tie", 0))
    kv("Single", c.get("single", 0))

    section("Audit State")
    kv("Audit rows total", data["audit"]["total"])
    if data["audit"]["total"] == 0:
        kv("Warning", "Audit log is empty. Earlier actions cannot be reconstructed if they were not logged.")

    # Audit summary sheet
    ws2 = wb.create_sheet("Audit Summary")
    ws2.append(["Category", "Key", "Count"])

    for category, counter in [
        ("Actor type", data["audit"]["by_actor"]),
        ("Operator/Admin/System", data["audit"]["by_operator"]),
        ("Action", data["audit"]["by_action"]),
        ("Review status", data["audit"]["by_review"]),
    ]:
        for key, value in counter.most_common():
            ws2.append([category, key, value])

    # Filters sheet
    ws_filters = wb.create_sheet("Run Filters")
    ws_filters.append(["Parameter", "Value"])
    ws_filters.append(["operator", data["filters"]["operator"] or "ALL"])
    ws_filters.append(["actor_type", data["filters"]["actor_type"] or "ALL"])
    ws_filters.append(["days", data["filters"]["days"] or "ALL"])
    ws_filters.append(["limit", data["filters"]["limit"]])
    ws_filters.append(["audit_rows_total", data["audit"]["total"]])
    if data["audit"]["total"] == 0:
        ws_filters.append(["warning", "No audit rows. Filters produce identical outputs because there is no audit data yet."])

    # Recent actions sheet
    ws3 = wb.create_sheet("Recent Actions")
    headers = [
        "id", "created_at", "actor_type", "operator_id", "action_type",
        "table_name", "row_id", "field_name", "old_value", "new_value",
        "review_status", "action_status", "source_context", "comment",
    ]
    ws3.append(headers)

    for row in data["recent_rows"]:
        ws3.append([row.get(h, "") for h in headers])

    for sheet in [ws, ws2, ws_filters, ws3]:
        for row_cells in sheet.iter_rows():
            for cell in row_cells:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        for cell in sheet[1]:
            cell.font = bold

        for col in range(1, sheet.max_column + 1):
            letter = get_column_letter(col)
            max_len = 10
            for cell in sheet[letter]:
                if cell.value is not None:
                    max_len = max(max_len, min(len(str(cell.value)), 60))
            sheet.column_dimensions[letter].width = max_len + 2

        sheet.freeze_panes = "A2"

    wb.save(xlsx_file)


def main():
    parser = argparse.ArgumentParser(description="Supervisor dashboard: TXT + optional Excel.")
    parser.add_argument("--period", default=DEFAULT_PERIOD_CODE)
    parser.add_argument("--limit", type=int, default=100)
    parser.add_argument("--operator", default="")
    parser.add_argument("--actor-type", default="", choices=["", "operator", "admin", "system", "user"])
    parser.add_argument("--days", type=int, default=0)
    parser.add_argument("--no-ensure-audit", action="store_true")
    parser.add_argument("--xlsx", action="store_true")

    args = parser.parse_args()

    data = collect_dashboard_data(args)

    report_dir = paths.OSBB_EXPORTS_DIR / "supervisor"
    report_dir.mkdir(parents=True, exist_ok=True)

    base = report_dir / f"supervisor_dashboard_{args.period}_{now_ts()}"

    txt_file = base.with_suffix(".txt")
    write_txt_report(txt_file, data)

    xlsx_file = None
    if args.xlsx:
        xlsx_file = base.with_suffix(".xlsx")
        write_xlsx_report(xlsx_file, data)

    print("=" * 70)
    print("OSBB SUPERVISOR DASHBOARD")
    print("=" * 70)
    print("DB:", get_db_file())
    print("MODE:", "TEST/WORK" if USE_TEST_DB else "PROD")
    print("Period:", args.period)
    if data["audit"]["total"] == 0:
        print("WARNING: operator_audit_log is empty; filters will look identical until new actions are logged.")
    print("TXT:")
    print(txt_file)
    if xlsx_file:
        print("XLSX:")
        print(xlsx_file)
    print("=" * 70)


if __name__ == "__main__":
    main()
