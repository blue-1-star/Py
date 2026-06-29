# cashier_parking_payments_audit_v4.py
# OSBB parking payments audit/export - READ ONLY
#
# Назначение:
#   - ничего не пишет в БД;
#   - читает charges, payment_allocations, payments, vehicles, apartments;
#   - НЕ угадывает период: умеет --list-periods;
#   - экспортирует Excel/TXT/CSV с нормальным разделением:
#       AllCharges, Paid, Unpaid, Partial, Overpaid, Problems, Periods, Summary;
#   - для переплаты по одной машине ищет вторую/другую машину той же квартиры
#     в том же периоде и сервисе, где есть недоплата.
#
# Адаптированный запуск из VS Code PowerShell, когда терминал открыт в G:\Programming\Py:
#   python .\OSBB\tools\cashier_parking_payments_audit_v4.py
#
# Список реальных периодов:
#   python .\OSBB\tools\cashier_parking_payments_audit_v4.py --list-periods
#
# Явный период:
#   python .\OSBB\tools\cashier_parking_payments_audit_v4.py --period 2026-05_2026-06

from __future__ import annotations

import argparse
import csv
import sqlite3
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Sequence, Tuple

THIS_FILE = Path(__file__).resolve()
OSBB_ROOT = THIS_FILE.parents[1]
PY_ROOT = OSBB_ROOT.parent
if str(PY_ROOT) not in sys.path:
    sys.path.insert(0, str(PY_ROOT))

from config import USE_TEST_DB, paths  # noqa: E402

DEFAULT_SERVICES = ("PARKING_DAY", "PARKING_NIGHT", "BARRIER_PHONE")
EPS = 0.00001


def now_stamp() -> str:
    return datetime.now().strftime("%Y-%m-%d_%H-%M-%S")


def now_text() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def money(value: Any) -> str:
    try:
        f = float(value or 0)
    except Exception:
        return str(value)
    if f.is_integer():
        return str(int(f))
    return f"{f:.2f}".rstrip("0").rstrip(".")


def norm_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def get_db_file(which: str) -> Path:
    if which == "prod":
        return paths.OSBB_DB_FILE
    if which == "test":
        return paths.OSBB_TEST_DB_FILE
    return paths.OSBB_TEST_DB_FILE if USE_TEST_DB else paths.OSBB_DB_FILE


def connect(db_file: Path) -> sqlite3.Connection:
    conn = sqlite3.connect(str(db_file))
    conn.row_factory = sqlite3.Row
    return conn


def table_exists(cur: sqlite3.Cursor, table_name: str) -> bool:
    cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table_name,))
    return cur.fetchone() is not None


def table_columns(cur: sqlite3.Cursor, table_name: str) -> List[str]:
    cur.execute(f"PRAGMA table_info({table_name})")
    return [row[1] for row in cur.fetchall()]


def require_tables(cur: sqlite3.Cursor, tables: Sequence[str]) -> None:
    missing = [t for t in tables if not table_exists(cur, t)]
    if missing:
        raise RuntimeError("Не найдены нужные таблицы: " + ", ".join(missing))


def qmarks(values: Sequence[Any]) -> str:
    return ",".join("?" for _ in values)


def get_paid_by_charge_sql(cur: sqlite3.Cursor) -> str:
    if not table_exists(cur, "payment_allocations"):
        return "SELECT NULL AS charge_id, 0 AS paid_amount WHERE 0"
    cols = table_columns(cur, "payment_allocations")
    amount_col = "amount" if "amount" in cols else ("allocated_amount" if "allocated_amount" in cols else None)
    if not amount_col or "charge_id" not in cols:
        return "SELECT NULL AS charge_id, 0 AS paid_amount WHERE 0"
    return f"""
        SELECT charge_id, COALESCE(SUM({amount_col}), 0) AS paid_amount
        FROM payment_allocations
        GROUP BY charge_id
    """


def load_periods(conn: sqlite3.Connection, services: Sequence[str]) -> List[Dict[str, Any]]:
    cur = conn.cursor()
    require_tables(cur, ["charges"])
    paid_sql = get_paid_by_charge_sql(cur)

    params: List[Any] = list(services)
    service_filter = ""
    if services:
        service_filter = f"AND COALESCE(c.base_service_code, c.service_code, '') IN ({qmarks(services)})"

    sql = f"""
        WITH paid AS ({paid_sql})
        SELECT
            COALESCE(c.period_code, '') AS period_code,
            COALESCE(c.base_service_code, c.service_code, '') AS service_code,
            COALESCE(c.service_item_code, '') AS service_item_code,
            COUNT(*) AS rows_count,
            COALESCE(SUM(c.amount), 0) AS amount_sum,
            COALESCE(SUM(COALESCE(p.paid_amount, 0)), 0) AS paid_sum,
            COALESCE(SUM(c.amount - COALESCE(p.paid_amount, 0)), 0) AS balance_sum
        FROM charges c
        LEFT JOIN paid p ON p.charge_id = c.id
        WHERE COALESCE(c.status, '') <> 'cancelled'
          {service_filter}
        GROUP BY COALESCE(c.period_code, ''), COALESCE(c.base_service_code, c.service_code, ''), COALESCE(c.service_item_code, '')
        ORDER BY period_code, service_code, service_item_code
    """
    cur.execute(sql, tuple(params))
    return [dict(r) for r in cur.fetchall()]


def choose_period(period_rows: List[Dict[str, Any]]) -> str:
    totals: Dict[str, float] = defaultdict(float)
    for row in period_rows:
        period = norm_text(row.get("period_code"))
        if not period:
            continue
        totals[period] += abs(float(row.get("balance_sum") or 0))
    if not totals:
        return ""
    return sorted(totals.items(), key=lambda kv: (kv[1], kv[0]), reverse=True)[0][0]


def load_charges(conn: sqlite3.Connection, period: str, services: Sequence[str]) -> List[Dict[str, Any]]:
    cur = conn.cursor()
    require_tables(cur, ["charges", "apartments"])
    paid_sql = get_paid_by_charge_sql(cur)

    has_vehicles = table_exists(cur, "vehicles")
    service_filter = ""
    params: List[Any] = [period]
    if services:
        service_filter = f"AND COALESCE(c.base_service_code, c.service_code, '') IN ({qmarks(services)})"
        params.extend(services)

    vehicle_join = ""
    vehicle_cols = "NULL AS plate, NULL AS model, NULL AS parking_time, NULL AS vehicle_status"
    if has_vehicles:
        vehicle_join = "LEFT JOIN vehicles v ON v.id = c.vehicle_id"
        vehicle_cols = """
            COALESCE(v.license_plate_normalized, v.license_plate, '') AS plate,
            COALESCE(v.car_model_normalized, v.car_model, '') AS model,
            COALESCE(v.parking_time, '') AS parking_time,
            COALESCE(v.status, '') AS vehicle_status
        """

    sql = f"""
        WITH paid AS ({paid_sql})
        SELECT
            COALESCE(a.entrance_number, CAST(a.entrance AS INTEGER), 0) AS entrance_number,
            COALESCE(a.entrance, '') AS entrance,
            a.id AS apartment_id,
            c.apartment_number,
            c.id AS charge_id,
            c.vehicle_id,
            COALESCE(c.base_service_code, c.service_code, '') AS service_code,
            COALESCE(c.service_item_code, '') AS service_item_code,
            COALESCE(c.service_type, '') AS service_type,
            c.period_code,
            c.amount AS charge_amount,
            COALESCE(p.paid_amount, 0) AS paid_amount,
            c.amount - COALESCE(p.paid_amount, 0) AS balance_amount,
            c.status AS charge_status,
            {vehicle_cols}
        FROM charges c
        LEFT JOIN paid p ON p.charge_id = c.id
        LEFT JOIN apartments a ON a.apartment_number = c.apartment_number
        {vehicle_join}
        WHERE c.period_code = ?
          AND COALESCE(c.status, '') <> 'cancelled'
          {service_filter}
        ORDER BY
            COALESCE(a.entrance_number, CAST(a.entrance AS INTEGER), 9999),
            CAST(c.apartment_number AS INTEGER),
            c.apartment_number,
            c.id
    """
    cur.execute(sql, tuple(params))
    rows = [dict(r) for r in cur.fetchall()]

    # Apartment-level context for multi-car situations.
    by_apt_service: Dict[Tuple[str, str], List[Dict[str, Any]]] = defaultdict(list)
    by_apt_all: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for row in rows:
        apt = norm_text(row.get("apartment_number"))
        svc = norm_text(row.get("service_code"))
        by_apt_service[(apt, svc)].append(row)
        by_apt_all[apt].append(row)

    for row in rows:
        amount = float(row.get("charge_amount") or 0)
        paid = float(row.get("paid_amount") or 0)
        balance = amount - paid
        row["balance_amount"] = balance

        if paid <= EPS:
            status = "UNPAID"
        elif abs(balance) <= EPS:
            status = "PAID"
        elif balance > EPS:
            status = "PARTIAL"
        else:
            status = "OVERPAID"
        row["payment_status"] = status

        flags: List[str] = []
        hints: List[str] = []
        service = norm_text(row.get("service_code"))
        parking_time = norm_text(row.get("parking_time"))
        vehicle_id = row.get("vehicle_id")
        apt = norm_text(row.get("apartment_number"))

        if service in {"PARKING_DAY", "PARKING_NIGHT"} and not vehicle_id:
            flags.append("NO_VEHICLE")
        if service in {"PARKING_DAY", "PARKING_NIGHT"} and not parking_time:
            flags.append("NO_PARKING_TIME")
        if not row.get("apartment_id"):
            flags.append("NO_APARTMENT_MATCH")

        siblings_same_service = [r for r in by_apt_service[(apt, service)] if r.get("charge_id") != row.get("charge_id")]
        siblings_all = [r for r in by_apt_all[apt] if r.get("charge_id") != row.get("charge_id")]
        sibling_underpaid = [r for r in siblings_same_service if float(r.get("charge_amount") or 0) - float(r.get("paid_amount") or 0) > EPS]
        row["apt_charge_count"] = len(by_apt_all[apt])
        row["same_service_charge_count"] = len(by_apt_service[(apt, service)])
        row["sibling_vehicles"] = "; ".join(
            f"charge={r.get('charge_id')} vehicle={r.get('vehicle_id') or '-'} {r.get('plate') or '-'} rest={money(float(r.get('charge_amount') or 0) - float(r.get('paid_amount') or 0))}"
            for r in siblings_all
        )

        if status == "OVERPAID":
            over = paid - amount
            row["overpaid_amount"] = over
            if sibling_underpaid:
                total_sibling_rest = sum(float(r.get("charge_amount") or 0) - float(r.get("paid_amount") or 0) for r in sibling_underpaid)
                flags.append("OVERPAID_HAS_SIBLING_UNDERPAID")
                hints.append(f"Вероятно оплата за второе авто этой квартиры; у соседних начислений остаток {money(total_sibling_rest)}")
            else:
                flags.append("OVERPAID_NO_SIBLING_UNDERPAID")
                hints.append("Переплата есть, но второй недоплаченной машины той же квартиры/услуги не найдено")
        else:
            row["overpaid_amount"] = 0.0

        if status == "PARTIAL":
            if len(by_apt_all[apt]) > 1:
                flags.append("PARTIAL_MULTI_VEHICLE_APT")
                hints.append("Недоплата при нескольких начислениях квартиры: проверить распределение между авто")
            else:
                flags.append("PARTIAL_SINGLE_CHARGE")
                hints.append("Оплата меньше начисления; простой пары внутри квартиры не видно")

        row["flags"] = ",".join(flags)
        row["hint"] = " | ".join(hints)
        row["problem"] = 1 if flags else 0
    return rows


def split_rows(rows: List[Dict[str, Any]]) -> Dict[str, List[Dict[str, Any]]]:
    return {
        "ALL": rows,
        "PAID": [r for r in rows if r["payment_status"] == "PAID"],
        "UNPAID": [r for r in rows if r["payment_status"] == "UNPAID"],
        "PARTIAL": [r for r in rows if r["payment_status"] == "PARTIAL"],
        "OVERPAID": [r for r in rows if r["payment_status"] == "OVERPAID"],
        "PROBLEMS": [r for r in rows if r.get("flags")],
    }


def summarize_by_entrance(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    grouped: Dict[str, Dict[str, Any]] = {}
    for row in rows:
        key = norm_text(row.get("entrance_number") or row.get("entrance") or "-")
        item = grouped.setdefault(key, {"entrance": key, "rows": 0, "charged": 0.0, "paid": 0.0, "balance": 0.0, "problems": 0})
        item["rows"] += 1
        item["charged"] += float(row.get("charge_amount") or 0)
        item["paid"] += float(row.get("paid_amount") or 0)
        item["balance"] += float(row.get("balance_amount") or 0)
        item["problems"] += int(row.get("problem") or 0)

    def sort_key(x: Dict[str, Any]) -> Tuple[int, str]:
        s = norm_text(x["entrance"])
        return (int(s) if s.isdigit() else 9999, s)

    return sorted(grouped.values(), key=sort_key)


HEADERS = [
    "entrance", "apartment", "charge_id", "vehicle_id", "service", "service_item", "period",
    "plate", "model", "parking_time", "charge_amount", "paid_amount", "balance_amount",
    "overpaid_amount", "payment_status", "apt_charge_count", "same_service_charge_count",
    "sibling_vehicles", "flags", "hint",
]


def row_to_list(r: Dict[str, Any]) -> List[Any]:
    return [
        r.get("entrance_number") or r.get("entrance") or "-",
        r.get("apartment_number") or "",
        r.get("charge_id"),
        r.get("vehicle_id"),
        r.get("service_code") or "",
        r.get("service_item_code") or "",
        r.get("period_code") or "",
        r.get("plate") or "",
        r.get("model") or "",
        r.get("parking_time") or "",
        float(r.get("charge_amount") or 0),
        float(r.get("paid_amount") or 0),
        float(r.get("balance_amount") or 0),
        float(r.get("overpaid_amount") or 0),
        r.get("payment_status") or "",
        r.get("apt_charge_count") or 0,
        r.get("same_service_charge_count") or 0,
        r.get("sibling_vehicles") or "",
        r.get("flags") or "",
        r.get("hint") or "",
    ]


def autosize_openpyxl(ws: Any, max_width: int = 50) -> None:
    from openpyxl.utils import get_column_letter
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_len = 8
        for cell in column_cells:
            if cell.value is not None:
                max_len = max(max_len, min(len(str(cell.value)), max_width))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, max_width)


def write_xlsx(xlsx_file: Path, *, db_file: Path, period: str, services: Sequence[str], period_rows: List[Dict[str, Any]], rows: List[Dict[str, Any]], summary: List[Dict[str, Any]]) -> bool:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except ImportError:
        return False

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    split = split_rows(rows)

    ws.append(["OSBB parking payments audit", "READ ONLY"])
    ws.append(["Generated", now_text()])
    ws.append(["USE_TEST_DB", str(USE_TEST_DB)])
    ws.append(["DB", str(db_file)])
    ws.append(["Period", period])
    ws.append(["Services", ", ".join(services)])
    ws.append([])
    ws.append(["Status", "Rows", "Charged", "Paid", "Balance"])
    for name in ["ALL", "PAID", "UNPAID", "PARTIAL", "OVERPAID", "PROBLEMS"]:
        data = split[name]
        ws.append([
            name,
            len(data),
            sum(float(r.get("charge_amount") or 0) for r in data),
            sum(float(r.get("paid_amount") or 0) for r in data),
            sum(float(r.get("balance_amount") or 0) for r in data),
        ])
    ws.append([])
    ws.append(["Entrance", "Rows", "Charged", "Paid", "Balance", "Problems"])
    for r in summary:
        ws.append([r["entrance"], r["rows"], r["charged"], r["paid"], r["balance"], r["problems"]])

    sheet_map = [
        ("AllCharges", split["ALL"]),
        ("Paid", split["PAID"]),
        ("Unpaid", split["UNPAID"]),
        ("Partial", split["PARTIAL"]),
        ("Overpaid", split["OVERPAID"]),
        ("Problems", split["PROBLEMS"]),
    ]
    for sheet_name, data in sheet_map:
        sh = wb.create_sheet(sheet_name)
        sh.append(HEADERS)
        for r in data:
            sh.append(row_to_list(r))

    wp = wb.create_sheet("Periods")
    wp.append(["period_code", "service", "service_item", "rows", "amount", "paid", "balance"])
    for r in period_rows:
        wp.append([
            r.get("period_code") or "",
            r.get("service_code") or "",
            r.get("service_item_code") or "",
            r.get("rows_count") or 0,
            float(r.get("amount_sum") or 0),
            float(r.get("paid_sum") or 0),
            float(r.get("balance_sum") or 0),
        ])

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        if sheet.title == "Summary":
            sheet["A1"].font = Font(bold=True, size=14)
            for header_row in (8, 16):
                for cell in sheet[header_row]:
                    cell.font = Font(bold=True)
                    cell.fill = header_fill
            sheet.freeze_panes = "A8"
        else:
            for cell in sheet[1]:
                cell.font = Font(bold=True)
                cell.fill = header_fill
            sheet.freeze_panes = "A2"
        autosize_openpyxl(sheet)

    for sh in wb.worksheets:
        if sh.title == "Summary":
            for row in sh.iter_rows(min_row=9, min_col=3, max_col=5):
                for cell in row:
                    cell.number_format = '#,##0.00'
            for row in sh.iter_rows(min_row=17, min_col=3, max_col=5):
                for cell in row:
                    cell.number_format = '#,##0.00'
        elif sh.title == "Periods":
            for row in sh.iter_rows(min_row=2, min_col=5, max_col=7):
                for cell in row:
                    cell.number_format = '#,##0.00'
        else:
            for row in sh.iter_rows(min_row=2, min_col=11, max_col=14):
                for cell in row:
                    cell.number_format = '#,##0.00'

    for sheet in wb.worksheets:
        if sheet.max_row >= 2 and sheet.max_column >= 2 and sheet.title != "Summary":
            try:
                tab = Table(displayName=sheet.title.replace(" ", "") + "Table", ref=sheet.dimensions)
                tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
                sheet.add_table(tab)
            except Exception:
                pass

    wb.save(xlsx_file)
    return True


def write_txt(report_file: Path, *, db_file: Path, period: str, services: Sequence[str], period_rows: List[Dict[str, Any]], rows: List[Dict[str, Any]], summary: List[Dict[str, Any]]) -> None:
    split = split_rows(rows)
    lines: List[str] = []
    lines.append("OSBB parking payments audit - READ ONLY")
    lines.append(f"Generated: {now_text()}")
    lines.append(f"USE_TEST_DB: {USE_TEST_DB}")
    lines.append(f"DB: {db_file}")
    lines.append(f"Period: {period}")
    lines.append("Services: " + ", ".join(services))
    lines.append("")
    lines.append("STATUS SUMMARY")
    lines.append("-" * 90)
    lines.append("status | rows | charged | paid | balance")
    for name in ["ALL", "PAID", "UNPAID", "PARTIAL", "OVERPAID", "PROBLEMS"]:
        data = split[name]
        lines.append(
            f"{name} | {len(data)} | "
            f"{money(sum(float(r.get('charge_amount') or 0) for r in data))} | "
            f"{money(sum(float(r.get('paid_amount') or 0) for r in data))} | "
            f"{money(sum(float(r.get('balance_amount') or 0) for r in data))}"
        )
    lines.append("")
    lines.append("SUMMARY BY ENTRANCE")
    lines.append("-" * 90)
    lines.append("entr | rows | charged | paid | balance | problems")
    for r in summary:
        lines.append(f"{r['entrance']} | {r['rows']} | {money(r['charged'])} | {money(r['paid'])} | {money(r['balance'])} | {r['problems']}")
    lines.append("")
    lines.append("PROBLEMS")
    lines.append("-" * 180)
    lines.append(" | ".join(HEADERS))
    for r in split["PROBLEMS"]:
        lines.append(" | ".join(str(x) for x in row_to_list(r)))
    lines.append("")
    lines.append("PERIODS IN CHARGES")
    lines.append("-" * 110)
    lines.append("period_code | service | item | rows | amount | paid | balance")
    for r in period_rows:
        lines.append(
            f"{r.get('period_code') or '-'} | {r.get('service_code') or '-'} | {r.get('service_item_code') or '-'} | "
            f"{r.get('rows_count')} | {money(r.get('amount_sum'))} | {money(r.get('paid_sum'))} | {money(r.get('balance_sum'))}"
        )
    lines.append("READ ONLY COMPLETED")
    report_file.write_text("\n".join(lines), encoding="utf-8")


def write_csv(csv_file: Path, rows: List[Dict[str, Any]]) -> None:
    with csv_file.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f, delimiter=";")
        writer.writerow(HEADERS)
        for r in rows:
            writer.writerow(row_to_list(r))


def print_periods(db_file: Path, period_rows: List[Dict[str, Any]]) -> None:
    print("OSBB cashier periods - READ ONLY")
    print("Generated:", now_text())
    print("USE_TEST_DB:", USE_TEST_DB)
    print("DB:", db_file)
    print("")
    print("period_code | service | item | rows | amount | paid | balance")
    print("-" * 100)
    for r in period_rows:
        print(
            f"{r.get('period_code') or '-'} | {r.get('service_code') or '-'} | {r.get('service_item_code') or '-'} | "
            f"{r.get('rows_count')} | {money(r.get('amount_sum'))} | {money(r.get('paid_sum'))} | {money(r.get('balance_sum'))}"
        )
    print("READ ONLY COMPLETED")


def main() -> None:
    parser = argparse.ArgumentParser(description="OSBB parking payments audit/export - READ ONLY")
    parser.add_argument("--which", choices=["config", "test", "prod"], default="config", help="DB source. Default: config.py USE_TEST_DB")
    parser.add_argument("--period", default="", help="Exact period_code. If omitted, auto-selects period with largest absolute balance.")
    parser.add_argument("--services", default=",".join(DEFAULT_SERVICES), help="Comma-separated base service codes")
    parser.add_argument("--list-periods", action="store_true", help="Only show period_code summary from charges")
    parser.add_argument("--no-xlsx", action="store_true", help="Do not create Excel file")
    parser.add_argument("--no-csv", action="store_true", help="Do not create CSV file")
    args = parser.parse_args()

    services = tuple(s.strip() for s in args.services.split(",") if s.strip())
    db_file = get_db_file(args.which)
    if not db_file.exists():
        raise RuntimeError(f"DB file not found: {db_file}")

    conn = connect(db_file)
    try:
        period_rows = load_periods(conn, services)
        if args.list_periods:
            print_periods(db_file, period_rows)
            return
        period = args.period.strip() or choose_period(period_rows)
        if not period:
            raise RuntimeError("Не найден период с начислениями. Запустите с --list-periods.")
        rows = load_charges(conn, period, services)
        summary = summarize_by_entrance(rows)
    finally:
        conn.close()

    export_dir = paths.OSBB_EXPORTS_DIR / "cashier"
    export_dir.mkdir(parents=True, exist_ok=True)
    stamp = now_stamp()
    safe_period = period.replace("/", "-").replace("\\", "-").replace(":", "-")
    base = export_dir / f"cashier_payments_audit_{safe_period}_{stamp}"
    txt_file = base.with_suffix(".txt")
    xlsx_file = base.with_suffix(".xlsx")
    csv_file = base.with_suffix(".csv")

    write_txt(txt_file, db_file=db_file, period=period, services=services, period_rows=period_rows, rows=rows, summary=summary)
    wrote_xlsx = False
    if not args.no_xlsx:
        wrote_xlsx = write_xlsx(xlsx_file, db_file=db_file, period=period, services=services, period_rows=period_rows, rows=rows, summary=summary)
    if not args.no_csv:
        write_csv(csv_file, rows)

    split = split_rows(rows)
    print("OSBB parking payments audit/export - READ ONLY")
    print("Generated:", now_text())
    print("USE_TEST_DB:", USE_TEST_DB)
    print("DB:", db_file)
    print("Period:", period)
    print("Services:", ", ".join(services))
    print("")
    print("Rows all:", len(rows))
    print("Paid:", len(split["PAID"]))
    print("Unpaid:", len(split["UNPAID"]))
    print("Partial:", len(split["PARTIAL"]))
    print("Overpaid:", len(split["OVERPAID"]))
    print("Problems:", len(split["PROBLEMS"]))
    print("")
    print("TXT :", txt_file)
    if wrote_xlsx:
        print("XLSX:", xlsx_file)
    elif not args.no_xlsx:
        print("XLSX: not created - openpyxl is not installed")
    if not args.no_csv:
        print("CSV :", csv_file)
    print("READ ONLY COMPLETED")


if __name__ == "__main__":
    main()
