from pathlib import Path
import sys
import sqlite3
import argparse
import re
from datetime import datetime
from decimal import Decimal, InvalidOperation
from openpyxl import load_workbook

OSBB_ROOT = Path(__file__).resolve().parent
PY_ROOT = OSBB_ROOT.parent
if str(PY_ROOT) not in sys.path:
    sys.path.insert(0, str(PY_ROOT))

from config import paths, USE_TEST_DB

DEFAULT_SOURCE_FILE = (getattr(paths, "OSBB_RAW_TYPED_DIR", paths.OSBB_RAW_DIR / "typed") / "Охорона.xlsx")
SHEET_NAME = "Sheet1"
DEFAULT_PERIOD_CODE = "2026-05_2026-06"

COL_DATE = 1
COL_ENTRANCE = 2
COL_APARTMENT = 3
COL_NAME = 4
COL_PLATE = 5
COL_PURPOSE = 6
COL_INCOME = 7
COL_EXPENSE = 8
COL_CASHBOX = 9

PURPOSE_MAP = {
    "П": "PARKING",
    "ПАРКОВКА": "PARKING",
    "P": "PARKING",
    "Ш": "BARRIER_CALL",
    "ШЛАГБАУМ": "BARRIER_CALL",
    "Б": "IMPROVEMENT",
    "БЛАГОУСТРОЙСТВО": "IMPROVEMENT",
    "БЛАГОУСТРІЙ": "IMPROVEMENT",
}


def get_db_file():
    return paths.OSBB_TEST_DB_FILE if USE_TEST_DB else paths.OSBB_DB_FILE


def now_ts():
    return datetime.now().strftime("%Y-%m-%d_%H-%M-%S")


def norm_text(value):
    return "" if value is None else str(value).strip()


def norm_apartment(value):
    text = norm_text(value)
    return text[:-2] if text.endswith(".0") else text


def normalize_plate(value):
    text = norm_text(value).upper()
    if not text:
        return ""

    text = text.replace("О", "O")
    if re.fullmatch(r"O\d{3}", text):
        text = "0" + text[1:]

    cyr_to_lat = str.maketrans({
        "А": "A", "В": "B", "Е": "E", "І": "I", "К": "K",
        "М": "M", "Н": "H", "О": "O", "Р": "P", "С": "C",
        "Т": "T", "Х": "X", "У": "Y",
    })
    text = text.translate(cyr_to_lat)
    text = re.sub(r"[^A-Z0-9]", "", text)

    if re.fullmatch(r"O\d{3}", text):
        text = "0" + text[1:]

    return text


def normalize_purpose(value):
    text = norm_text(value).upper().replace(".", "").strip()
    return PURPOSE_MAP.get(text, text)


def normalize_cashbox(value):
    return norm_text(value).upper().replace("О", "O")


def parse_money(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = norm_text(value).replace(" ", "").replace(",", ".")
    if not text:
        return None
    match = re.search(r"[-+]?\d+(?:\.\d+)?", text)
    if not match:
        return None
    try:
        return float(Decimal(match.group(0)))
    except (InvalidOperation, ValueError):
        return None


def normalize_date(value):
    if value is None:
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")
    return norm_text(value)


def table_exists(cur, table_name):
    cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table_name,))
    return cur.fetchone() is not None


def table_columns(cur, table_name):
    cur.execute(f"PRAGMA table_info({table_name})")
    return [row[1] for row in cur.fetchall()]


def pick_table(cur, *names):
    for name in names:
        if table_exists(cur, name):
            return name
    return None


def detect_header_row(ws):
    expected = ["дата", "кв", "plate", "назнач", "приход", "расход", "касса"]
    best_row = 1
    best_score = -1

    for row_idx in range(1, min(ws.max_row, 30) + 1):
        joined = " | ".join(
            norm_text(ws.cell(row=row_idx, column=col).value).lower()
            for col in range(1, min(ws.max_column, 12) + 1)
        )
        score = sum(1 for word in expected if word in joined)
        if score > best_score:
            best_score = score
            best_row = row_idx

    return best_row


def read_sheet1_rows(source_file):
    wb = load_workbook(source_file, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Лист {SHEET_NAME!r} не найден. Доступные листы: {wb.sheetnames}")

    ws = wb[SHEET_NAME]
    header_row = detect_header_row(ws)
    rows = []

    for row_idx in range(header_row + 1, ws.max_row + 1):
        values = [ws.cell(row=row_idx, column=col).value for col in range(1, 10)]
        if not any(norm_text(v) for v in values):
            continue

        purpose_raw = norm_text(ws.cell(row=row_idx, column=COL_PURPOSE).value)
        purpose = normalize_purpose(purpose_raw)
        income = parse_money(ws.cell(row=row_idx, column=COL_INCOME).value)
        expense = parse_money(ws.cell(row=row_idx, column=COL_EXPENSE).value)

        row = {
            "excel_row": row_idx,
            "payment_date": normalize_date(ws.cell(row=row_idx, column=COL_DATE).value),
            "entrance": norm_text(ws.cell(row=row_idx, column=COL_ENTRANCE).value),
            "apartment_number": norm_apartment(ws.cell(row=row_idx, column=COL_APARTMENT).value),
            "payer_name": norm_text(ws.cell(row=row_idx, column=COL_NAME).value),
            "plate_raw": norm_text(ws.cell(row=row_idx, column=COL_PLATE).value),
            "plate": normalize_plate(ws.cell(row=row_idx, column=COL_PLATE).value),
            "purpose_raw": purpose_raw,
            "purpose": purpose,
            "income": income,
            "expense": expense,
            "cashbox": normalize_cashbox(ws.cell(row=row_idx, column=COL_CASHBOX).value),
            "parse_status": "ok",
            "parse_error": "",
            "row_kind": "payment",
        }

        is_total_row = (
            not row["payment_date"]
            and not row["apartment_number"]
            and not row["payer_name"]
            and not row["plate_raw"]
            and not row["purpose_raw"]
            and (income is not None or expense is not None or row["cashbox"])
        )

        if is_total_row:
            row["row_kind"] = "cashbox_total"
            row["parse_status"] = "skip"
            row["parse_error"] = "cashbox_total_row"
            rows.append(row)
            continue

        errors = []
        if income is None:
            errors.append("missing_income")
        if purpose not in ["PARKING", "BARRIER_CALL", "IMPROVEMENT"]:
            errors.append(f"unknown_purpose:{purpose_raw}")
        if errors:
            row["parse_status"] = "warning"
            row["parse_error"] = "; ".join(errors)

        rows.append(row)

    return rows, header_row


def load_vehicle_index(cur):
    cur.execute("""
        SELECT v.id, a.apartment_number,
               COALESCE(v.license_plate_normalized, v.license_plate) AS plate,
               v.parking_time
        FROM vehicles v
        JOIN apartments a ON a.id = v.apartment_id
        ORDER BY v.id
    """)
    vehicles = []
    for vehicle_id, apartment_number, plate, parking_time in cur.fetchall():
        vehicles.append({
            "vehicle_id": vehicle_id,
            "apartment_number": str(apartment_number),
            "plate": normalize_plate(plate),
            "plate_raw": plate,
            "parking_time": parking_time,
        })
    return vehicles


def find_vehicle(cur, plate, apartment_number):
    vehicles = load_vehicle_index(cur)
    apt = str(apartment_number) if apartment_number else ""
    plate_norm = normalize_plate(plate)

    if plate_norm:
        exact = [v for v in vehicles if v["plate"] == plate_norm]
        if len(exact) == 1:
            return exact[0], "plate_exact"
        if len(exact) > 1:
            if apt:
                exact_apt = [v for v in exact if v["apartment_number"] == apt]
                if len(exact_apt) == 1:
                    return exact_apt[0], "plate_and_apartment"
            return None, "plate_multiple"

        if re.fullmatch(r"\d{4}", plate_norm):
            partial = [
                v for v in vehicles
                if plate_norm in re.sub(r"[^0-9]", "", v["plate"])
            ]
            if apt:
                partial_apt = [v for v in partial if v["apartment_number"] == apt]
                if len(partial_apt) == 1:
                    return partial_apt[0], "plate_4digits_and_apartment"
            if len(partial) == 1:
                return partial[0], "plate_4digits_unique"
            if len(partial) > 1:
                return None, "plate_4digits_multiple"

    if apt:
        apt_vehicles = [v for v in vehicles if v["apartment_number"] == apt]
        if len(apt_vehicles) == 1:
            return apt_vehicles[0], "apartment_single_vehicle"
        if len(apt_vehicles) > 1:
            return None, "apartment_multiple_vehicles"

    return None, "vehicle_not_found"


def service_code_for(item, vehicle):
    if item["purpose"] == "BARRIER_CALL":
        return "BARRIER_CALL"
    if item["purpose"] == "IMPROVEMENT":
        return "IMPROVEMENT"
    if item["purpose"] == "PARKING":
        if vehicle:
            if vehicle["parking_time"] == "Day":
                return "PARKING_DAY"
            if vehicle["parking_time"] == "Night":
                return "PARKING_NIGHT"
        return None
    return None


def find_charge(cur, charges_table, period_code, vehicle_id, service_code, amount):
    if not charges_table or not vehicle_id or not service_code:
        return None, "no_charge_lookup"

    columns = table_columns(cur, charges_table)
    status_col = "charge_status" if "charge_status" in columns else ("status" if "status" in columns else None)

    sql = f"""
        SELECT id, amount
        FROM {charges_table}
        WHERE period_code = ?
          AND vehicle_id = ?
          AND service_code = ?
    """
    params = [period_code, vehicle_id, service_code]

    if status_col:
        sql += f" AND COALESCE({status_col}, '') <> 'cancelled'"

    sql += " ORDER BY id LIMIT 1"
    cur.execute(sql, tuple(params))
    result = cur.fetchone()

    if not result:
        return None, "charge_not_found"

    charge_id, charge_amount = result
    diff = float(amount or 0) - float(charge_amount or 0)

    if abs(diff) < 0.01:
        return charge_id, "charge_exact"
    if diff > 0:
        return charge_id, f"charge_overpay:{diff:g}"
    return charge_id, f"charge_underpay:{abs(diff):g}"


def enrich_rows(conn, rows, period_code):
    cur = conn.cursor()
    charges_table = pick_table(cur, "charges", "service_charges")
    enriched = []

    for row in rows:
        item = dict(row)
        vehicle, vehicle_status = find_vehicle(cur, row["plate"], row["apartment_number"])

        item["vehicle_match_status"] = vehicle_status
        item["vehicle_id"] = vehicle["vehicle_id"] if vehicle else None
        item["vehicle_apartment"] = vehicle["apartment_number"] if vehicle else ""
        item["vehicle_plate"] = vehicle["plate"] if vehicle else ""
        item["vehicle_parking_time"] = vehicle["parking_time"] if vehicle else ""

        item["effective_apartment"] = item["vehicle_apartment"] or item["apartment_number"]
        item["service_code"] = service_code_for(item, vehicle)

        charge_id, charge_status = find_charge(
            cur,
            charges_table,
            period_code,
            item["vehicle_id"],
            item["service_code"],
            item["income"],
        )
        item["charge_id"] = charge_id
        item["charge_match_status"] = charge_status

        enriched.append(item)

    return enriched


def duplicate_payment_exists(cur, payments_table, source_ref):
    columns = table_columns(cur, payments_table)
    if "source_ref" not in columns:
        return False
    cur.execute(f"SELECT id FROM {payments_table} WHERE source_ref = ? LIMIT 1", (source_ref,))
    return cur.fetchone() is not None


def insert_dynamic(cur, table_name, values):
    columns = table_columns(cur, table_name)
    insert_cols = [col for col in values if col in columns]
    placeholders = ",".join("?" for _ in insert_cols)

    cur.execute(f"""
        INSERT INTO {table_name} ({", ".join(insert_cols)})
        VALUES ({placeholders})
    """, tuple(values[col] for col in insert_cols))

    return cur.lastrowid


def apply_rows(conn, enriched, period_code):
    cur = conn.cursor()
    payments_table = pick_table(cur, "payments", "service_payments")
    allocations_table = pick_table(cur, "payment_allocations", "service_payment_allocations")

    if not payments_table:
        raise RuntimeError("Не найдена таблица payments/service_payments")

    inserted_payments = 0
    inserted_allocations = 0
    skipped_duplicates = 0
    skipped_errors = 0

    for item in enriched:
        if item.get("row_kind") != "payment":
            skipped_errors += 1
            continue

        if item["income"] is None or item["income"] <= 0:
            skipped_errors += 1
            continue

        if item["parse_status"] != "ok":
            skipped_errors += 1
            continue

        source_ref = f"ohorona_sheet1:{period_code}:row:{item['excel_row']}"

        if duplicate_payment_exists(cur, payments_table, source_ref):
            skipped_duplicates += 1
            continue

        comment = (
            f"Охорона.xlsx Sheet1 row {item['excel_row']}; "
            f"purpose={item['purpose_raw']}; cashbox={item['cashbox']}; "
            f"payer={item['payer_name']}; plate_raw={item['plate_raw']}; "
            f"vehicle_match={item['vehicle_match_status']}; "
            f"charge_match={item['charge_match_status']}"
        )

        payment_values = {
            "payment_date": item["payment_date"] or None,
            "period_code": period_code,
            "apartment_number": item["effective_apartment"],
            "vehicle_id": item["vehicle_id"],
            "service_code": item["service_code"],
            "amount": item["income"],
            "currency": "UAH",
            "payment_method": "cash",
            "source": "import_ohorona_sheet1",
            "source_ref": source_ref,
            "created_by": "import_ohorona_sheet1_payments.py",
            "comment": comment,
        }

        payment_id = insert_dynamic(cur, payments_table, payment_values)
        inserted_payments += 1

        if item.get("charge_id") and allocations_table:
            allocation_values = {
                "payment_id": payment_id,
                "charge_id": item["charge_id"],
                "amount": item["income"],
            }
            insert_dynamic(cur, allocations_table, allocation_values)
            inserted_allocations += 1

    conn.commit()

    return {
        "payments_table": payments_table,
        "allocations_table": allocations_table,
        "inserted_payments": inserted_payments,
        "inserted_allocations": inserted_allocations,
        "skipped_duplicates": skipped_duplicates,
        "skipped_errors": skipped_errors,
    }


def cashbox_summary(enriched):
    result = {}
    for item in enriched:
        if item.get("row_kind") != "payment":
            continue

        cashbox = item["cashbox"] or "-"
        income = item["income"] or 0
        expense = item["expense"] or 0
        if cashbox not in result:
            result[cashbox] = {"income": 0, "expense": 0, "net": 0, "rows": 0}
        result[cashbox]["income"] += income
        result[cashbox]["expense"] += expense
        result[cashbox]["net"] += income - expense
        result[cashbox]["rows"] += 1
    return result


def cashbox_total_rows(enriched):
    return [item for item in enriched if item.get("row_kind") == "cashbox_total"]


def unknown_plates(enriched):
    result = []
    for item in enriched:
        if item.get("row_kind") != "payment":
            continue
        if item["purpose"] != "PARKING":
            continue
        if item.get("vehicle_match_status") == "vehicle_not_found":
            result.append(item)
    return result


def skipped_system_rows(enriched):
    """
    Строки, которые есть в Excel, но не являются платежами к импорту:
    кассовые итоги, строки без назначения, строки без прихода.
    """
    result = []

    for item in enriched:
        if item.get("row_kind") == "cashbox_total":
            result.append(item)
            continue

        if item.get("row_kind") != "payment":
            result.append(item)
            continue

        if item.get("parse_status") != "ok" and (
            "unknown_purpose" in item.get("parse_error", "")
            or "missing_income" in item.get("parse_error", "")
        ):
            result.append(item)

    return result


def manual_review_required(enriched):
    """
    Реальные платежные строки, которые можно хранить как оплаты,
    но они требуют внимания человека:
    - парковка без найденного авто;
    - неоднозначное авто;
    - недоплата/переплата относительно начисления;
    - шлагбаум/благоустройство без начисления — это нормально для денег,
      но полезно видеть отдельно.
    """
    result = []

    for item in enriched:
        if item.get("row_kind") != "payment":
            continue

        if item.get("parse_status") != "ok":
            continue

        if item.get("purpose") == "PARKING":
            # Комбинат/коммерческая структура: П + сумма есть, но нет квартиры/номера.
            # Это не системная строка, а ручная строка к учёту.
            if not item.get("apartment_number") and not item.get("plate"):
                result.append(item)
                continue

            if item.get("vehicle_match_status") not in {
                "plate_exact",
                "plate_and_apartment",
                "plate_4digits_unique",
                "plate_4digits_and_apartment",
                "apartment_single_vehicle",
            }:
                result.append(item)
                continue

            if item.get("charge_match_status") != "charge_exact":
                result.append(item)
                continue

        if item.get("purpose") in {"BARRIER_CALL", "IMPROVEMENT"}:
            if item.get("charge_match_status") in {"charge_not_found", "no_charge_lookup"}:
                result.append(item)

    return result


def write_report(report_file, enriched, apply_result, db_file, period_code, apply_mode, header_row, source_file):
    lines = []
    lines.append("=" * 120)
    lines.append("OHORONA SHEET1 PAYMENTS DRY RUN" if not apply_mode else "OHORONA SHEET1 PAYMENTS APPLY")
    lines.append("=" * 120)
    lines.append(f"Generated : {datetime.now():%Y-%m-%d %H:%M:%S}")
    lines.append(f"DB        : {db_file}")
    lines.append(f"MODE      : {'TEST/WORK' if USE_TEST_DB else 'PROD'}")
    lines.append(f"Source    : {source_file}")
    lines.append(f"Sheet     : {SHEET_NAME}")
    lines.append(f"Period    : {period_code}")
    lines.append(f"Header row: {header_row}")
    lines.append("Columns   : Date | Entrance | Apt | Name | Plate | Purpose | Income | Expense | Cashbox")
    lines.append("")

    lines.append("=" * 120)
    lines.append("SIMPLE CHECK TABLE")
    lines.append("=" * 120)
    lines.append("row | kind | date | apt | effective_apt | plate_raw | plate_norm | purpose | income | cashbox | service | vehicle_match | charge_match")
    lines.append("-" * 120)

    for item in enriched:
        lines.append(
            f"{item['excel_row']} | {item.get('row_kind', 'payment')} | {item['payment_date']} | "
            f"{item['apartment_number']} | {item['effective_apartment']} | "
            f"{item['plate_raw']} | {item['plate']} | "
            f"{item['purpose_raw']} | {item['income']} | "
            f"{item['cashbox']} | {item.get('service_code') or '-'} | "
            f"{item.get('vehicle_match_status')} | {item.get('charge_match_status')}"
        )

    lines.append("")
    lines.append("=" * 120)
    lines.append("SKIPPED_SYSTEM_ROWS — НЕ ИМПОРТИРУЮТСЯ")
    lines.append("=" * 120)

    system_rows = skipped_system_rows(enriched)

    if not system_rows:
        lines.append("нет системных/служебных строк")
    else:
        lines.append("row | kind | date | apt | plate | purpose | income | expense | cashbox | reason")
        lines.append("-" * 120)
        for item in system_rows:
            lines.append(
                f"{item['excel_row']} | {item.get('row_kind')} | "
                f"{item['payment_date']} | {item['apartment_number']} | "
                f"{item['plate_raw']} | {item['purpose_raw']} | "
                f"{item['income']} | {item['expense']} | {item['cashbox']} | "
                f"{item.get('parse_error', '')}"
            )

    lines.append("")
    lines.append("=" * 120)
    lines.append("MANUAL_REVIEW_REQUIRED")
    lines.append("=" * 120)

    good_vehicle_statuses = {
        "plate_exact",
        "plate_and_apartment",
        "plate_4digits_unique",
        "plate_4digits_and_apartment",
        "apartment_single_vehicle",
    }

    warnings = manual_review_required(enriched)

    if not warnings:
        lines.append("нет строк, требующих ручного просмотра")
    else:
        for item in warnings:
            lines.append(
                f"row {item['excel_row']}: apt={item['apartment_number']} "
                f"effective_apt={item['effective_apartment']} "
                f"plate={item['plate_raw']}->{item['plate']} "
                f"purpose={item['purpose_raw']} income={item['income']} cashbox={item['cashbox']} | "
                f"parse={item['parse_status']} {item['parse_error']} | "
                f"vehicle={item.get('vehicle_match_status')} | "
                f"charge={item.get('charge_match_status')}"
            )

    lines.append("")
    lines.append("=" * 120)
    lines.append("UNKNOWN PLATES FROM OHORONA")
    lines.append("=" * 120)

    unknowns = unknown_plates(enriched)
    if not unknowns:
        lines.append("нет неопознанных номеров")
    else:
        lines.append("row | date | apt | plate_raw | plate_norm | income | cashbox")
        lines.append("-" * 120)
        for item in unknowns:
            lines.append(
                f"{item['excel_row']} | {item['payment_date']} | "
                f"{item['apartment_number']} | {item['plate_raw']} | "
                f"{item['plate']} | {item['income']} | {item['cashbox']}"
            )

    lines.append("")
    lines.append("=" * 120)
    lines.append("CASHBOX TOTAL ROWS — НЕ ИМПОРТИРУЮТСЯ КАК ПЛАТЕЖИ")
    lines.append("=" * 120)

    total_rows = cashbox_total_rows(enriched)
    if not total_rows:
        lines.append("нет итоговых строк")
    else:
        lines.append("row | income | expense | cashbox/net")
        lines.append("-" * 120)
        for item in total_rows:
            lines.append(
                f"{item['excel_row']} | {item['income']} | "
                f"{item['expense']} | {item['cashbox']}"
            )

    lines.append("")
    lines.append("=" * 120)
    lines.append("CASHBOX SUMMARY")
    lines.append("=" * 120)
    lines.append("cashbox | rows | income | expense | net")
    lines.append("-" * 120)

    for cashbox, data in sorted(cashbox_summary(enriched).items()):
        lines.append(f"{cashbox} | {data['rows']} | {data['income']:g} | {data['expense']:g} | {data['net']:g}")

    payment_rows = [item for item in enriched if item.get("row_kind") == "payment"]
    total_income = sum(item["income"] or 0 for item in payment_rows)
    total_expense = sum(item["expense"] or 0 for item in payment_rows)
    parking_income = sum(item["income"] or 0 for item in payment_rows if item["purpose"] == "PARKING")
    barrier_income = sum(item["income"] or 0 for item in payment_rows if item["purpose"] == "BARRIER_CALL")
    improvement_income = sum(item["income"] or 0 for item in payment_rows if item["purpose"] == "IMPROVEMENT")

    lines.append("")
    lines.append("=" * 120)
    lines.append("SUMMARY")
    lines.append("=" * 120)
    lines.append(f"Parsed rows        : {len(enriched)}")
    lines.append(f"Payment rows       : {len(payment_rows)}")
    lines.append(f"Cashbox total rows : {len(cashbox_total_rows(enriched))}")
    lines.append(f"Skipped system rows: {len(skipped_system_rows(enriched))}")
    lines.append(f"Manual review rows : {len(warnings)}")
    lines.append(f"Unknown plates     : {len(unknown_plates(enriched))}")
    lines.append(f"Total income       : {total_income:g}")
    lines.append(f"Total expense      : {total_expense:g}")
    lines.append(f"Net cash movement  : {total_income - total_expense:g}")
    lines.append(f"Parking income     : {parking_income:g}")
    lines.append(f"Barrier income     : {barrier_income:g}")
    lines.append(f"Improvement income : {improvement_income:g}")

    if apply_result:
        lines.append("")
        lines.append("APPLY RESULT")
        for key, value in apply_result.items():
            lines.append(f"{key}: {value}")
    else:
        lines.append("")
        lines.append("DRY RUN ONLY. Database was not changed.")
        lines.append("To apply, run with --apply")

    report_file.write_text("\n".join(lines), encoding="utf-8")


def main():
    parser = argparse.ArgumentParser(description="Import payments from Sheet1 of Охорона.xlsx.")
    parser.add_argument("--source", default=str(DEFAULT_SOURCE_FILE))
    parser.add_argument("--period", default=DEFAULT_PERIOD_CODE)
    parser.add_argument("--apply", action="store_true")
    args = parser.parse_args()

    source_file = Path(args.source)
    if not source_file.exists():
        raise FileNotFoundError(f"Файл не найден: {source_file}")

    rows, header_row = read_sheet1_rows(source_file)
    db_file = get_db_file()
    conn = sqlite3.connect(db_file)

    enriched = enrich_rows(conn, rows, args.period)

    apply_result = None
    if args.apply:
        apply_result = apply_rows(conn, enriched, args.period)

    report_dir = paths.OSBB_EXPORTS_DIR / "billing"
    report_dir.mkdir(parents=True, exist_ok=True)

    mode = "apply" if args.apply else "dry_run"
    report_file = report_dir / f"ohorona_sheet1_payments_{args.period}_{mode}_{now_ts()}.txt"

    write_report(
        report_file=report_file,
        enriched=enriched,
        apply_result=apply_result,
        db_file=db_file,
        period_code=args.period,
        apply_mode=args.apply,
        header_row=header_row,
        source_file=source_file,
    )

    conn.close()

    print("=" * 70)
    print("OHORONA SHEET1 PAYMENTS")
    print("=" * 70)
    print("DB:", db_file)
    print("MODE:", "TEST/WORK" if USE_TEST_DB else "PROD")
    print("Source:", source_file)
    print("Period:", args.period)
    print("Apply:", args.apply)
    print("Rows parsed:", len(rows))

    if apply_result:
        print("Inserted payments:", apply_result["inserted_payments"])
        print("Inserted allocations:", apply_result["inserted_allocations"])
        print("Skipped duplicates:", apply_result["skipped_duplicates"])
        print("Skipped errors:", apply_result["skipped_errors"])
    else:
        print("DRY RUN ONLY. Database was not changed.")

    print("")
    print("Report:")
    print(report_file)
    print("=" * 70)


if __name__ == "__main__":
    main()
