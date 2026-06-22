from pathlib import Path
import sys
import sqlite3
import argparse
import re
from datetime import datetime
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook

OSBB_ROOT = Path(__file__).resolve().parent
PY_ROOT = OSBB_ROOT.parent

if str(PY_ROOT) not in sys.path:
    sys.path.insert(0, str(PY_ROOT))

from config import paths, USE_TEST_DB
from audit_logger import audit_log

# Используем единую финансовую логику нового журнала кассира.
from cashier_journal import (
    get_conn,
    now_db,
    norm_text,
    norm_apartment,
    normalize_plate,
    normalize_period,
    monthly_item_code,
    table_exists,
    table_columns,
    insert_dynamic,
    update_dynamic,
    ensure_cashbox,
    recalc_cashbox_balance,
    find_vehicle,
    get_service_type,
    ensure_service_catalog,
    ensure_service_item,
    create_cashbox_operation,
    duplicate_operation_exists,
    money,
)


DEFAULT_SOURCE_FILE = (
    getattr(paths, "OSBB_RAW_TYPED_DIR", paths.OSBB_RAW_DIR / "typed") / "Охорона.xlsx"
)
SHEET_NAME = "Sheet1"

# Это старый технический ключ, использованный предыдущим импортом 30 платежей.
# Он нужен только для связи исторической строки с уже существующей записью payments.
DEFAULT_LEGACY_PERIOD = "2026-05_2026-06"

COL_DATE = 1
COL_ENTRANCE = 2
COL_APARTMENT = 3
COL_NAME = 4
COL_PLATE = 5
COL_PURPOSE = 6
COL_INCOME = 7
COL_EXPENSE = 8
COL_CASHBOX = 9

PURPOSE_MAP = {
    "П": "PARKING",
    "ПАРКОВКА": "PARKING",
    "P": "PARKING",

    # ВАЖНО: Ш = сбор/ремонт шлагбаума, НЕ телефонное открытие.
    "Ш": "BARRIER_REPAIR",
    "ШЛАГБАУМ": "BARRIER_REPAIR",
    "РЕМОНТ ШЛАГБАУМА": "BARRIER_REPAIR",

    "Б": "IMPROVEMENT",
    "БЛАГОУСТРОЙСТВО": "IMPROVEMENT",
    "БЛАГОУСТРІЙ": "IMPROVEMENT",
}


def get_db_file():
    return paths.OSBB_TEST_DB_FILE if USE_TEST_DB else paths.OSBB_DB_FILE


def now_ts():
    return datetime.now().strftime("%Y-%m-%d_%H-%M-%S")


def parse_money(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)

    text = norm_text(value).replace(" ", "").replace(",", ".")
    if not text:
        return None

    match = re.search(r"[-+]?\d+(?:\.\d+)?", text)
    if not match:
        return None

    try:
        return float(Decimal(match.group(0)))
    except (InvalidOperation, ValueError):
        return None


def normalize_cashbox(value):
    return norm_text(value).upper().replace("О", "O")


def normalize_date(value):
    if value is None:
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")

    text = norm_text(value)
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", text):
        return text
    return text


def normalize_purpose(value):
    text = norm_text(value).upper().replace(".", "").strip()
    return PURPOSE_MAP.get(text, text)


def detect_header_row(ws):
    expected = ["дата", "кв", "plate", "назнач", "приход", "расход", "касса"]
    best_row, best_score = 1, -1

    for row_idx in range(1, min(ws.max_row, 30) + 1):
        joined = " | ".join(
            norm_text(ws.cell(row=row_idx, column=col).value).lower()
            for col in range(1, min(ws.max_column, 12) + 1)
        )
        score = sum(1 for word in expected if word in joined)
        if score > best_score:
            best_row, best_score = row_idx, score

    return best_row


def read_source_rows(source_file):
    wb = load_workbook(source_file, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise RuntimeError(f"Лист {SHEET_NAME!r} не найден. Есть листы: {wb.sheetnames}")

    ws = wb[SHEET_NAME]
    header_row = detect_header_row(ws)
    result = []

    for row_idx in range(header_row + 1, ws.max_row + 1):
        values = [ws.cell(row=row_idx, column=col).value for col in range(1, 10)]
        if not any(norm_text(value) for value in values):
            continue

        income = parse_money(ws.cell(row=row_idx, column=COL_INCOME).value)
        expense = parse_money(ws.cell(row=row_idx, column=COL_EXPENSE).value)
        purpose_raw = norm_text(ws.cell(row=row_idx, column=COL_PURPOSE).value)

        item = {
            "excel_row": row_idx,
            "payment_date": normalize_date(ws.cell(row=row_idx, column=COL_DATE).value),
            "entrance": norm_text(ws.cell(row=row_idx, column=COL_ENTRANCE).value),
            "apartment_number": norm_apartment(ws.cell(row=row_idx, column=COL_APARTMENT).value),
            "payer_name": norm_text(ws.cell(row=row_idx, column=COL_NAME).value),
            "plate_raw": norm_text(ws.cell(row=row_idx, column=COL_PLATE).value),
            "plate": normalize_plate(ws.cell(row=row_idx, column=COL_PLATE).value),
            "purpose_raw": purpose_raw,
            "purpose": normalize_purpose(purpose_raw),
            "income": income,
            "expense": expense,
            "cashbox_raw": normalize_cashbox(ws.cell(row=row_idx, column=COL_CASHBOX).value),
        }

        is_total = (
            not item["payment_date"]
            and not item["apartment_number"]
            and not item["payer_name"]
            and not item["plate_raw"]
            and not item["purpose_raw"]
            and (income is not None or expense is not None or item["cashbox_raw"])
        )
        item["row_kind"] = "cashbox_total" if is_total else "operation"
        result.append(item)

    return result, header_row


def rows_to_dicts(rows):
    return [dict(row) for row in rows]


def payment_candidates(cur, where_sql, params):
    cur.execute(f"SELECT * FROM payments WHERE {where_sql} ORDER BY id", tuple(params))
    return rows_to_dicts(cur.fetchall())


def unique_unclaimed(candidates, claimed_payment_ids):
    available = [
        item for item in candidates
        if item.get("id") not in (claimed_payment_ids or set())
    ]
    if len(available) == 1:
        return available[0], "unique"
    if not available:
        return None, "none"
    return None, "ambiguous"


def find_legacy_payment(cur, source_row, legacy_period, vehicle_id=None, claimed_payment_ids=None):
    """
    Платежи были импортированы ранним модулем import_ohorona_sheet1_payments.py.
    В части старых БД source_ref не сохранился, хотя строка Excel всегда
    записывалась в comment. Поэтому порядок поиска такой:

      1. source_ref (если есть);
      2. source + comment с номером строки Excel;
      3. точное совпадение vehicle/date/amount;
      4. точное совпадение apartment/date/amount;
      5. date/amount — только если кандидат строго один.

    Неоднозначные совпадения НЕ связываются автоматически.
    """
    if not table_exists(cur, "payments"):
        return None, "no_payments_table"

    cols = table_columns(cur, "payments")
    claimed_payment_ids = claimed_payment_ids or set()
    row_no = source_row["excel_row"]
    amount = source_row.get("income")
    pay_date = source_row.get("payment_date")
    apartment = norm_apartment(source_row.get("apartment_number"))

    if amount in (None, 0) or not is_valid_iso_date(pay_date):
        return None, "not_applicable"

    # 1. Прямой ключ старого импортёра.
    if "source_ref" in cols:
        source_ref = f"ohorona_sheet1:{legacy_period}:row:{row_no}"
        direct = payment_candidates(cur, "source_ref = ?", [source_ref])
        payment, state = unique_unclaimed(direct, claimed_payment_ids)
        if payment:
            return payment, "source_ref"
        if state == "ambiguous":
            return None, "ambiguous_source_ref"

    # 2. Самый надёжный fallback: ранний импорт записывал номер строки в comment.
    if "comment" in cols:
        marker = f"Охорона.xlsx Sheet1 row {row_no};"
        where = ["comment LIKE ?"]
        params = [f"%{marker}%"]

        if "source" in cols:
            where.insert(0, "source = ?")
            params.insert(0, "import_ohorona_sheet1")

        candidates = payment_candidates(cur, " AND ".join(where), params)
        payment, state = unique_unclaimed(candidates, claimed_payment_ids)
        if payment:
            return payment, "comment_row"
        if state == "ambiguous":
            return None, "ambiguous_comment_row"

        # На всякий случай: comment без source.
        if "source" in cols:
            candidates = payment_candidates(cur, "comment LIKE ?", [f"%{marker}%"])
            payment, state = unique_unclaimed(candidates, claimed_payment_ids)
            if payment:
                return payment, "comment_row_without_source"
            if state == "ambiguous":
                return None, "ambiguous_comment_row"

    # Общая основа для осторожного точного совпадения.
    base_where = ["ABS(COALESCE(amount, 0) - ?) < 0.00001", "payment_date = ?"]
    base_params = [float(amount), pay_date]

    if "source" in cols:
        base_where.insert(0, "source = ?")
        base_params.insert(0, "import_ohorona_sheet1")

    # 3. vehicle/date/amount.
    if vehicle_id and "vehicle_id" in cols:
        candidates = payment_candidates(
            cur,
            " AND ".join(base_where + ["vehicle_id = ?"]),
            base_params + [vehicle_id],
        )
        payment, state = unique_unclaimed(candidates, claimed_payment_ids)
        if payment:
            return payment, "vehicle_date_amount"
        if state == "ambiguous":
            return None, "ambiguous_vehicle_date_amount"

    # 4. apartment/date/amount.
    if apartment and "apartment_number" in cols:
        candidates = payment_candidates(
            cur,
            " AND ".join(base_where + ["apartment_number = ?"]),
            base_params + [apartment],
        )
        payment, state = unique_unclaimed(candidates, claimed_payment_ids)
        if payment:
            return payment, "apartment_date_amount"
        if state == "ambiguous":
            return None, "ambiguous_apartment_date_amount"

    # 5. Только дата и сумма — связываем лишь при одном кандидате.
    candidates = payment_candidates(cur, " AND ".join(base_where), base_params)
    payment, state = unique_unclaimed(candidates, claimed_payment_ids)
    if payment:
        return payment, "date_amount_unique"
    if state == "ambiguous":
        return None, "ambiguous_date_amount"

    return None, "no_match"

def is_valid_iso_date(value):
    return bool(re.fullmatch(r"\d{4}-\d{2}-\d{2}", norm_text(value)))


def get_vehicle_for_row(cur, payment, source_row):
    vehicle_id = payment.get("vehicle_id") if payment else None
    if vehicle_id:
        vehicle = find_vehicle(cur, vehicle_id=vehicle_id)
        if vehicle:
            return vehicle

    return find_vehicle(
        cur,
        apt=source_row.get("apartment_number", ""),
        plate=source_row.get("plate", ""),
    )


def historical_item_for_unclassified(cur, operation_date):
    month = normalize_period("", operation_date, required=True)
    year_short = month[2:4]
    month_num = month[5:7]
    code = f"HIST_{month_num}_{year_short}_Unclassified"

    ensure_service_item(
        cur=cur,
        item_code=code,
        base_service_code="HISTORICAL_UNCLASSIFIED",
        item_name=f"Исторические нераспределённые операции — {month}",
        service_type="HISTORICAL",
        period_code=month,
        description="Строка Охорона.xlsx с неполными реквизитами",
        comment="Требует последующей классификации по найденным бумажным документам",
    )
    return code, month, "HISTORICAL"


def resolve_historical_service(cur, source_row, payment, vehicle):
    """
    Определяет базовый сервис и конкретную статью для кассовой операции.
    Старый period_code платежа не переписываем: он мог быть двухмесячным костылём.
    Для кассовой статьи используем месяц даты фактического поступления.
    """
    purpose = source_row["purpose"]
    operation_date = source_row["payment_date"]

    if purpose == "PARKING":
        base = ""
        if payment:
            candidate = norm_text(payment.get("base_service_code") or payment.get("service_code"))
            if candidate in {"PARKING_DAY", "PARKING_NIGHT"}:
                base = candidate

        if not base and vehicle:
            parking_time = norm_text(vehicle.get("parking_time"))
            if parking_time == "Day":
                base = "PARKING_DAY"
            elif parking_time == "Night":
                base = "PARKING_NIGHT"

        if base:
            period = normalize_period("", operation_date, required=True)
            item_code = monthly_item_code(period, base)
            ensure_service_item(
                cur=cur,
                item_code=item_code,
                base_service_code=base,
                item_name=f"{'Парковка Day' if base == 'PARKING_DAY' else 'Парковка Night'} — {period}",
                service_type="MONTHLY",
                period_code=period,
                date_from=f"{period}-01",
                description="Исторический платёж, статья создана по дате поступления",
                comment="Период выведен из даты кассовой операции; может быть уточнён позднее",
            )
            return base, item_code, "MONTHLY", period, ""

        item_code, period, service_type = historical_item_for_unclassified(cur, operation_date)
        return (
            "HISTORICAL_UNCLASSIFIED",
            item_code,
            service_type,
            period,
            "Не удалось определить тариф Day/Night для исторического парковочного платежа",
        )

    if purpose == "BARRIER_REPAIR":
        ensure_service_item(
            cur=cur,
            item_code="01_BarrierRepair",
            base_service_code="BARRIER_REPAIR",
            item_name="Ремонт шлагбаума — сбор 01",
            service_type="FUNDRAISING",
            sequence_no=1,
            amount_default=1000,
            description="Исторический сбор из Охорона.xlsx, обозначение Ш",
            comment="Ш = ремонт/сбор на шлагбаум",
        )
        return "BARRIER_REPAIR", "01_BarrierRepair", "FUNDRAISING", "", ""

    if purpose == "IMPROVEMENT":
        ensure_service_item(
            cur=cur,
            item_code="01_Imp",
            base_service_code="IMPROVEMENT",
            item_name="Благоустройство — сбор 01",
            service_type="FUNDRAISING",
            sequence_no=1,
            description="Исторический сбор из Охорона.xlsx, обозначение Б",
            comment="Описание сбора уточняется",
        )
        return "IMPROVEMENT", "01_Imp", "FUNDRAISING", "", ""

    item_code, period, service_type = historical_item_for_unclassified(cur, operation_date)
    return (
        "HISTORICAL_UNCLASSIFIED",
        item_code,
        service_type,
        period,
        f"Неизвестное назначение: {source_row['purpose_raw']!r}",
    )


def effective_cashbox(source_row, default_cashbox):
    raw = source_row["cashbox_raw"]
    if raw:
        return raw
    return norm_text(default_cashbox).upper().replace("О", "O")


def operation_source_ref(legacy_period, excel_row):
    return f"cashier_backfill:ohorona_sheet1:{legacy_period}:row:{excel_row}"


def update_payment_metadata(cur, payment_id, cashbox_code, operation_id, base_code, item_code, service_type):
    update_dynamic(cur, "payments", payment_id, {
        "service_code": base_code,
        "base_service_code": base_code,
        "service_item_code": item_code,
        "service_type": service_type,
        "cashbox_code": cashbox_code,
        "cashbox_operation_id": operation_id,
        "operator_id": "ohorona_backfill",
    })


def make_row_plan(cur, source_row, args, claimed_payment_ids=None):
    plan = {
        "excel_row": source_row["excel_row"],
        "status": "ready",
        "reason": "",
        "operation_date": source_row["payment_date"],
        "cashbox_code": effective_cashbox(source_row, args.default_cashbox),
        "direction": "",
        "operation_type": "",
        "amount": 0,
        "payment": None,
        "payment_id": None,
        "payment_match_method": "",
        "vehicle": None,
        "vehicle_id": None,
        "apartment_number": source_row["apartment_number"],
        "base_service_code": "",
        "service_item_code": "",
        "service_type": "",
        "period_code": "",
        "metadata_warning": "",
    }

    if source_row["row_kind"] == "cashbox_total":
        plan["status"] = "skip"
        plan["reason"] = "итоговая строка кассы — не является операцией"
        return plan

    income = source_row["income"]
    expense = source_row["expense"]

    if income is None and expense is None:
        plan["status"] = "skip"
        plan["reason"] = "нет ни прихода, ни расхода"
        return plan

    if income not in (None, 0) and expense not in (None, 0):
        plan["status"] = "skip"
        plan["reason"] = "в одной строке одновременно приход и расход"
        return plan

    if not is_valid_iso_date(plan["operation_date"]):
        if args.unknown_date:
            plan["operation_date"] = args.unknown_date
            plan["metadata_warning"] = f"Дата в источнике нераспознана; использована вручную заданная дата {args.unknown_date}"
        else:
            plan["status"] = "skip"
            plan["reason"] = "неизвестная дата; добавьте --unknown-date YYYY-MM-DD после проверки бумажного источника"
            return plan

    if not plan["cashbox_code"]:
        plan["status"] = "skip"
        plan["reason"] = "не указана касса; добавьте --default-cashbox O только если уверены"
        return plan

    if income not in (None, 0):
        plan["direction"] = "in"
        plan["operation_type"] = "historical_income"
        plan["amount"] = float(income)
    else:
        plan["direction"] = "out"
        plan["operation_type"] = "historical_expense"
        plan["amount"] = float(expense)

    # Сначала пробуем найти авто по самой строке — это помогает сопоставлению
    # старого платежа в случаях, где source_ref не был сохранён.
    source_vehicle = find_vehicle(
        cur,
        apt=source_row.get("apartment_number", ""),
        plate=source_row.get("plate", ""),
    )
    source_vehicle_id = source_vehicle.get("vehicle_id") if source_vehicle else None

    payment = None
    match_method = "not_applicable"
    if plan["direction"] == "in":
        payment, match_method = find_legacy_payment(
            cur=cur,
            source_row=source_row,
            legacy_period=args.legacy_period,
            vehicle_id=source_vehicle_id,
            claimed_payment_ids=claimed_payment_ids,
        )

    plan["payment"] = payment
    plan["payment_id"] = payment.get("id") if payment else None
    plan["payment_match_method"] = match_method

    vehicle = get_vehicle_for_row(cur, payment, source_row)
    plan["vehicle"] = vehicle
    plan["vehicle_id"] = vehicle.get("vehicle_id") if vehicle else None

    if vehicle and not plan["apartment_number"]:
        plan["apartment_number"] = norm_apartment(vehicle.get("apartment_number"))

    base, item, service_type, period, warning = resolve_historical_service(
        cur=cur,
        source_row=source_row,
        payment=payment,
        vehicle=vehicle,
    )

    plan["base_service_code"] = base
    plan["service_item_code"] = item
    plan["service_type"] = service_type
    plan["period_code"] = period
    plan["metadata_warning"] = "; ".join(x for x in [plan["metadata_warning"], warning] if x)

    if plan["direction"] == "in" and payment:
        plan["metadata_warning"] = "; ".join(
            x for x in [
                plan["metadata_warning"],
                f"Связан payment id={plan['payment_id']} ({plan['payment_match_method']})",
            ]
            if x
        )
    elif plan["direction"] == "in":
        plan["metadata_warning"] = "; ".join(
            x for x in [
                plan["metadata_warning"],
                (
                    "Связанная запись payments не найдена; будет только кассовая операция "
                    f"(поиск: {plan['payment_match_method']})"
                ),
            ]
            if x
        )

    plan["source_ref"] = operation_source_ref(args.legacy_period, source_row["excel_row"])
    if duplicate_operation_exists(cur, plan["source_ref"]):
        plan["status"] = "duplicate"
        plan["reason"] = "кассовая операция уже существует по source_ref"

    return plan


def apply_plans(conn, plans, args):
    cur = conn.cursor()

    result = {
        "inserted_operations": 0,
        "linked_existing_payments": 0,
        "ledger_only_operations": 0,
        "skipped": 0,
        "duplicates": 0,
        "snapshot_operation_id": None,
        "snapshot_adjustment": 0,
        "balances": {},
    }

    for plan in plans:
        if plan["status"] == "skip":
            result["skipped"] += 1
            continue
        if plan["status"] == "duplicate":
            result["duplicates"] += 1
            continue

        ensure_cashbox(cur, plan["cashbox_code"])
        ensure_service_catalog(cur, plan["base_service_code"])

        comment_parts = [
            f"Исторический импорт Охорона.xlsx, строка {plan['excel_row']}",
            f"Назначение: {plan['base_service_code']}",
        ]
        if plan["metadata_warning"]:
            comment_parts.append(f"Внимание: {plan['metadata_warning']}")
        if args.comment:
            comment_parts.append(args.comment)

        operation_id = create_cashbox_operation(
            cur=cur,
            operation_date=plan["operation_date"],
            cashbox_code=plan["cashbox_code"],
            operation_type=plan["operation_type"],
            direction=plan["direction"],
            amount=plan["amount"],
            base_service_code=plan["base_service_code"],
            service_item_code=plan["service_item_code"],
            service_type=plan["service_type"],
            period_code=plan["period_code"],
            apartment_number=plan["apartment_number"],
            vehicle_id=plan["vehicle_id"],
            payment_id=plan["payment_id"],
            source_type="ohorona_sheet1_backfill",
            source_ref=plan["source_ref"],
            operator_id="ohorona_backfill",
            actor_type="system",
            comment="; ".join(comment_parts),
        )
        result["inserted_operations"] += 1

        if plan["payment_id"]:
            update_payment_metadata(
                cur=cur,
                payment_id=plan["payment_id"],
                cashbox_code=plan["cashbox_code"],
                operation_id=operation_id,
                base_code=plan["base_service_code"],
                item_code=plan["service_item_code"],
                service_type=plan["service_type"],
            )
            result["linked_existing_payments"] += 1
        else:
            result["ledger_only_operations"] += 1

        audit_log(
            conn=conn,
            actor_type="system",
            operator_id="ohorona_backfill",
            user_id="ohorona_backfill",
            action_type="cashbox_operation_backfilled",
            table_name="cashbox_operations",
            row_id=operation_id,
            field_name="amount",
            old_value="",
            new_value=plan["amount"],
            source_context="import_ohorona_to_cashbox.py",
            comment=f"Охорона.xlsx строка {plan['excel_row']}; {plan['direction']}; статья {plan['service_item_code']}",
            extra={
                "excel_row": plan["excel_row"],
                "cashbox_code": plan["cashbox_code"],
                "payment_id": plan["payment_id"],
                "base_service_code": plan["base_service_code"],
                "service_item_code": plan["service_item_code"],
                "metadata_warning": plan["metadata_warning"],
            },
            commit=False,
        )

    # Сначала пересчитываем кассы по детальным историческим операциям.
    cur.execute("SELECT cashbox_code FROM cashboxes WHERE COALESCE(is_active, 1) = 1")
    for row in cur.fetchall():
        code = row[0]
        result["balances"][code] = recalc_cashbox_balance(cur, code)

    # Не создаём сверку автоматически. Она делается только по явно указанному
    # пользователем остатку кассы из подтверждённой итоговой строки.
    if args.snapshot_balance is not None:
        snapshot_box = args.snapshot_cashbox.strip().upper().replace("О", "O")
        ensure_cashbox(cur, snapshot_box)

        current = recalc_cashbox_balance(cur, snapshot_box)
        target = float(args.snapshot_balance)
        delta = target - current

        if abs(delta) > 0.00001:
            snapshot_date = args.snapshot_date or today()
            recon_item = f"OHORONA_RECON_{snapshot_date.replace('-', '_')}"
            ensure_service_item(
                cur=cur,
                item_code=recon_item,
                base_service_code="CASH_RECONCILIATION",
                item_name=f"Сверка исторического остатка кассы — {snapshot_date}",
                service_type="RECONCILIATION",
                period_code=normalize_period("", snapshot_date, required=True),
                date_from=snapshot_date,
                description="Разница между детальными историческими строками и подтверждённым остатком кассы",
                comment="Нераспределённая историческая разница; уточняется по бумажным документам",
            )

            snapshot_ref = (
                f"cashier_backfill:ohorona_snapshot:{snapshot_box}:{snapshot_date}:{money(target)}"
            )
            if not duplicate_operation_exists(cur, snapshot_ref):
                operation_id = create_cashbox_operation(
                    cur=cur,
                    operation_date=snapshot_date,
                    cashbox_code=snapshot_box,
                    operation_type="historical_reconciliation",
                    direction="in" if delta > 0 else "out",
                    amount=abs(delta),
                    base_service_code="CASH_RECONCILIATION",
                    service_item_code=recon_item,
                    service_type="RECONCILIATION",
                    period_code=normalize_period("", snapshot_date, required=True),
                    source_type="ohorona_sheet1_reconciliation",
                    source_ref=snapshot_ref,
                    operator_id="ohorona_backfill",
                    actor_type="system",
                    comment=(
                        f"Сверка по подтверждённому остатку {money(target)} UAH. "
                        f"Детальные операции до сверки: {money(current)} UAH. "
                        "Разница не распределена по статьям и требует уточнения."
                    ),
                )
                result["snapshot_operation_id"] = operation_id
                result["snapshot_adjustment"] = delta

                audit_log(
                    conn=conn,
                    actor_type="system",
                    operator_id="ohorona_backfill",
                    user_id="ohorona_backfill",
                    action_type="cashbox_historical_reconciliation",
                    table_name="cashbox_operations",
                    row_id=operation_id,
                    field_name="amount",
                    old_value=current,
                    new_value=target,
                    source_context="import_ohorona_to_cashbox.py",
                    comment=f"Сверка кассы {snapshot_box} по старому остатку",
                    extra={
                        "cashbox_code": snapshot_box,
                        "detail_balance": current,
                        "target_balance": target,
                        "adjustment": delta,
                        "service_item_code": recon_item,
                    },
                    commit=False,
                )

            result["balances"][snapshot_box] = recalc_cashbox_balance(cur, snapshot_box)

    audit_log(
        conn=conn,
        actor_type="system",
        operator_id="ohorona_backfill",
        user_id="ohorona_backfill",
        action_type="ohorona_cashbox_backfill_summary",
        table_name="cashbox_operations",
        row_id="",
        field_name="",
        old_value="",
        new_value="",
        source_context="import_ohorona_to_cashbox.py",
        comment=(
            f"Исторический импорт Охорона.xlsx: операций {result['inserted_operations']}, "
            f"связанных платежей {result['linked_existing_payments']}, "
            f"только кассовых строк {result['ledger_only_operations']}"
        ),
        extra=result,
        commit=False,
    )

    conn.commit()
    return result


def make_report(report_file, source_file, header_row, plans, result, args):
    lines = [
        "=" * 130,
        "ИСТОРИЧЕСКИЙ ИМПОРТ ОХОРОНА.XLSX В ЖУРНАЛ КАССИРА",
        "=" * 130,
        f"Generated        : {now_db()}",
        f"DB               : {get_db_file()}",
        f"MODE             : {'TEST/WORK' if USE_TEST_DB else 'PROD'}",
        f"Source           : {source_file}",
        f"Sheet            : {SHEET_NAME}",
        f"Header row       : {header_row}",
        f"Legacy period key: {args.legacy_period}",
        f"Default cashbox  : {args.default_cashbox or '-'}",
        f"Snapshot balance : {args.snapshot_balance if args.snapshot_balance is not None else '-'}",
        "",
        "ПЛАН ОПЕРАЦИЙ:",
        "row | status | date | cashbox | direction | amount | base service | item | payment_id | payment match | warning/reason",
        "-" * 130,
    ]

    for plan in plans:
        lines.append(
            f"{plan['excel_row']} | {plan['status']} | {plan.get('operation_date', '')} | "
            f"{plan.get('cashbox_code', '') or '-'} | {plan.get('direction', '') or '-'} | "
            f"{money(plan.get('amount', 0))} | {plan.get('base_service_code', '') or '-'} | "
            f"{plan.get('service_item_code', '') or '-'} | {plan.get('payment_id') or '-'} | "
            f"{plan.get('payment_match_method') or '-'} | "
            f"{plan.get('metadata_warning') or plan.get('reason') or ''}"
        )

    lines.extend(["", "СВОДКА:"])

    status_counts = {}
    for plan in plans:
        status_counts[plan["status"]] = status_counts.get(plan["status"], 0) + 1

    for status, count in sorted(status_counts.items()):
        lines.append(f"{status}: {count}")

    match_counts = {}
    for plan in plans:
        method = plan.get("payment_match_method") or "-"
        match_counts[method] = match_counts.get(method, 0) + 1

    lines.extend(["", "СОПОСТАВЛЕНИЕ С УЖЕ ИМПОРТИРОВАННЫМИ PAYMENT:"])
    for method, count in sorted(match_counts.items()):
        lines.append(f"{method}: {count}")

    if result:
        lines.extend(["", "РЕЗУЛЬТАТ APPLY:"])
        for key, value in result.items():
            lines.append(f"{key}: {value}")
    else:
        lines.extend([
            "",
            "DRY RUN ONLY. База не изменялась.",
            "Для записи добавьте --apply.",
            "Сверку по итогу старого файла добавляйте только после проверки: "
            "--snapshot-balance <сумма> --snapshot-date YYYY-MM-DD.",
        ])

    report_file.write_text("\n".join(lines), encoding="utf-8")


def main():
    parser = argparse.ArgumentParser(
        description=(
            "Связывает уже импортированные платежи Охорона.xlsx с новым журналом кассира "
            "и создаёт cashbox_operations без дублирования payments. "
            "Поддерживает старые платежи без source_ref: поиск по comment, авто, квартире, дате и сумме."
        )
    )
    parser.add_argument("--source", default=str(DEFAULT_SOURCE_FILE))
    parser.add_argument("--legacy-period", default=DEFAULT_LEGACY_PERIOD)
    parser.add_argument(
        "--default-cashbox",
        default="",
        help="Касса для строк Охорона.xlsx, где касса пуста. Например: O",
    )
    parser.add_argument(
        "--unknown-date",
        default="",
        help="Явная дата для строк с нераспознанной датой. Иначе такие строки будут пропущены.",
    )
    parser.add_argument(
        "--snapshot-balance",
        type=float,
        default=None,
        help="Подтверждённый остаток кассы после исторического периода. Создаёт отдельную операцию сверки.",
    )
    parser.add_argument(
        "--snapshot-cashbox",
        default="O",
        help="Касса, для которой применяется подтверждённый остаток.",
    )
    parser.add_argument(
        "--snapshot-date",
        default="",
        help="Дата подтверждённого остатка, например 2026-06-01.",
    )
    parser.add_argument("--comment", default="")
    parser.add_argument("--apply", action="store_true")
    args = parser.parse_args()

    source_file = Path(args.source)
    if not source_file.exists():
        raise FileNotFoundError(f"Файл не найден: {source_file}")

    if args.unknown_date and not is_valid_iso_date(args.unknown_date):
        raise RuntimeError("--unknown-date должен быть в формате YYYY-MM-DD.")
    if args.snapshot_date and not is_valid_iso_date(args.snapshot_date):
        raise RuntimeError("--snapshot-date должен быть в формате YYYY-MM-DD.")

    rows, header_row = read_source_rows(source_file)
    conn = get_conn()
    cur = conn.cursor()

    claimed_payment_ids = set()
    plans = []

    for row in rows:
        plan = make_row_plan(
            cur,
            row,
            args,
            claimed_payment_ids=claimed_payment_ids,
        )
        if plan.get("payment_id"):
            claimed_payment_ids.add(plan["payment_id"])
        plans.append(plan)

    result = None

    if args.apply:
        result = apply_plans(conn, plans, args)

    conn.close()

    report_dir = paths.OSBB_EXPORTS_DIR / "cashier"
    report_dir.mkdir(parents=True, exist_ok=True)
    mode = "apply" if args.apply else "dry_run"
    report_file = report_dir / f"ohorona_cashbox_backfill_{mode}_{now_ts()}.txt"

    make_report(report_file, source_file, header_row, plans, result, args)

    print("=" * 78)
    print("ОХОРОНА.XLSX → ЖУРНАЛ КАССИРА")
    print("=" * 78)
    print("DB:", get_db_file())
    print("MODE:", "TEST/WORK" if USE_TEST_DB else "PROD")
    print("Rows:", len(rows))
    print("Apply:", args.apply)

    if result:
        print("Inserted cashbox operations:", result["inserted_operations"])
        print("Linked existing payments    :", result["linked_existing_payments"])
        print("Ledger-only operations      :", result["ledger_only_operations"])
        print("Skipped                     :", result["skipped"])
        print("Duplicates                  :", result["duplicates"])
        print("Snapshot operation          :", result["snapshot_operation_id"] or "-")
        print("Snapshot adjustment         :", money(result["snapshot_adjustment"]))
        print("Balances                    :", result["balances"])
    else:
        print("DRY RUN ONLY. База не изменялась.")

    print("")
    print("Report:")
    print(report_file)
    print("=" * 78)


if __name__ == "__main__":
    main()
