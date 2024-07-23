import openpyxl

# Создание новой рабочей книги
wb = openpyxl.Workbook()

# Выбор активного листа
ws = wb.active

# Запись данных в ячейки
ws['A1'] = 'Привет'
ws['B1'] = 'Мир'

# Сохранение рабочей книги в файл
# file_path = 'C:/path/to/your/directory/example.xlsx'
str1 = 'Всё хорошо прекрасная маркиза! Все хорошо, все хорошо'
str2 = 'Daily Readings to make your Life the best it can be'
many_strings = [str1, str2]
len_ms = len(many_strings) 
# sl =  str.split() 
# print(sl)
file_path = r'G:\Programming\Py\Tax\data\test.xlsx'
target_column = "A"
row = 1
while ws[f"{target_column}{row}"].value:
    row += 1
# nr = 3
# fill_row = str.split()
for i in range(len_ms):
    fill_row = many_strings[i].split()
    for j in range(len(fill_row)):
        ws.cell(row=row, column=j+1).value = fill_row[j]
    row += 1


# sheet.cell(row=row, column=j).value for j in range(1, len(fill_row)) = fill_row[j] 
wb.save(file_path)

