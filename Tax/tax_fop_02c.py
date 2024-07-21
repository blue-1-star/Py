import openpyxl
from openpyxl.utils import get_column_letter
from tkinter import Tk, Button, Label, Entry, StringVar, messagebox

# Путь к файлу Excel
file_path = r"G:\Programming\Py\Tax\data\L_02a.xlsx"
tax_rate = 0.05

def load_workbook():
    wb = openpyxl.load_workbook(file_path)
    return wb

def get_next_quarter(current_quarter):
    quarter, year = map(int, current_quarter.split("_"))
    quarter += 1
    if quarter > 4:
        quarter = 1
        year += 1
    return f"{quarter:02d}_{year:04d}"

def add_data():
    wb = load_workbook()
    sheet = wb.active
    target_column = "E"

    # Найти первую пустую строку в столбце E
    row = 1
    while sheet[f"{target_column}{row}"].value:
        row += 1

    prev_quarter = sheet[f"{target_column}{row - 1}"].value if row > 1 else "00_0000"
    current_quarter = get_next_quarter(prev_quarter)
    sheet[f"{target_column}{row}"].value = current_quarter

    # Ввод квартального дохода
    try:
        quarterly_income = float(income_var.get())
    except ValueError:
        messagebox.showerror("Ошибка", "Некорректный доход")
        return

    # Заполняем ячейки
    income_col = get_column_letter(ord(target_column) + 1)
    total_inc_col = get_column_letter(ord(target_column) + 2)
    total_tax_col = get_column_letter(ord(target_column) + 3)
    prev_tax_col = get_column_letter(ord(target_column) + 4)
    tax_pay_col = get_column_letter(ord(target_column) + 5)

    if current_quarter.startswith("01"):  # Первый квартал
        total_income = quarterly_income
        total_tax = total_income * tax_rate
        prev_tax = 0
    else:
        prev_total_income = sheet[f"G{row - 1}"].value if row > 1 else 0
        if prev_total_income is None:
            prev_total_income = 0

        total_income = quarterly_income + prev_total_income
        prev_total_tax = sheet[f"H{row - 1}"].value if row > 1 else 0
        if prev_total_tax is None:
            prev_total_tax = 0

        total_tax = prev_total_tax + quarterly_income * tax_rate
        prev_tax = prev_total_tax

    sheet[f"E{row}"].value = current_quarter
    sheet[f"F{row}"].value = quarterly_income
    sheet[f"G{row}"].value = total_income
    sheet[f"H{row}"].value = total_tax
    sheet[f"I{row}"].value = prev_tax
    tax_pay = total_tax - prev_tax
    sheet[f"J{row}"].value = tax_pay

    wb.save(file_path)
    messagebox.showinfo("Успех", "Данные успешно добавлены")
    income_var.set("")

def show_last_two_rows(sheet):
    target_column = "E"
    row = 1
    while sheet[f"{target_column}{row}"].value:
        row += 1

    last_rows = []
    for i in range(max(1, row - 2), row):
        row_data = [sheet.cell(row=i, column=j).value for j in range(5, 11)]
        last_rows.append(row_data)
    
    return last_rows

def create_gui():
    root = Tk()
    root.title("Добавление квартальных данных")
    
    Label(root, text="Введите квартальный доход:").grid(row=0, column=0, padx=10, pady=10)
    
    global income_var
    income_var = StringVar()
    Entry(root, textvariable=income_var).grid(row=0, column=1, padx=10, pady=10)

    wb = load_workbook()
    sheet = wb.active
    last_rows = show_last_two_rows(sheet)

    Label(root, text="Последние две записи:").grid(row=1, column=0, columnspan=2)

    headers = ["quarterly", "quart_inc", "total_inc", "total_tax", "prev_tax", "Tax_pay"]
    for col_num, header in enumerate(headers, 0):
        Label(root, text=header).grid(row=2, column=col_num, padx=5, pady=5)

    for row_num, row_data in enumerate(last_rows, 3):
        for col_num, cell_value in enumerate(row_data, 0):
            Label(root, text=str(cell_value)).grid(row=row_num, column=col_num, padx=5, pady=5)

    Button(root, text="Добавить данные", command=add_data).grid(row=5, column=0, columnspan=2, padx=10, pady=10)
    
    root.mainloop()

if __name__ == "__main__":
    create_gui()
