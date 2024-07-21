import openpyxl
from openpyxl.utils import get_column_letter
from tkinter import Tk, Button, Label, Entry, StringVar, messagebox

# Путь к файлу Excel
file_path = r"G:\Programming\Py\Tax\data\L_02a.xlsx"

def load_workbook():
    wb = openpyxl.load_workbook(file_path)
    return wb

def get_next_quarter(current_quarter):
    quarter, year = map(int, current_quarter.split("_"))
    quarter += 1
    if quarter > 4:
        quarter = 1
        year += 1
    return f"{quarter:02d}_{year:04d}"

def add_data():
    wb = load_workbook()
    sheet = wb.active
    target_column = "E"

    # Найти первую пустую строку в столбце E
    row = 1
    while sheet[f"{target_column}{row}"].value:
        row += 1

    prev_quarter = sheet[f"{target_column}{row - 1}"].value if row > 1 else "00_0000"
    current_quarter = get_next_quarter(prev_quarter)
    sheet[f"{target_column}{row}"].value = current_quarter

    # Ввод квартального дохода
    try:
        quarterly_income = float(income_var.get())
    except ValueError:
        messagebox.showerror("Ошибка", "Некорректный доход")
        return

    # Заполняем ячейки
    income_col = get_column_letter(ord(target_column) + 1)
    total_inc_col = get_column_letter(ord(target_column) + 2)
    total_tax_col = get_column_letter(ord(target_column) + 3)
    prev_tax_col = get_column_letter(ord(target_column) + 4)
    tax_pay_col = get_column_letter(ord(target_column) + 5)

    print(f"Adding data to row {row}:")
    print(f"Quarter: {current_quarter}")
    print(f"Quarterly Income: {quarterly_income}")

    sheet[f"E{row}"].value = current_quarter
    sheet[f"F{row}"].value = quarterly_income
    print(f"F{row}: {sheet[f'F{row}'].value}")

    prev_total_income = sheet[f"G{row - 1}"].value if row > 1 else 0
    if prev_total_income is None:
        prev_total_income = 0

    total_income = quarterly_income + prev_total_income
    sheet[f"G{row}"].value = total_income
    print(f"G{row}: {sheet[f'G{row}'].value}")

    total_tax = total_income * 0.05
    sheet[f"H{row}"].value = total_tax
    print(f"H{row}: {sheet[f'H{row}'].value}")

    prev_total_tax = sheet[f"I{row - 1}"].value if row > 1 else 0
    if prev_total_tax is None:
        prev_total_tax = 0

    sheet[f"I{row}"].value = prev_total_tax
    print(f"I{row}: {sheet[f'I{row}'].value}")

    tax_pay = total_tax - prev_total_tax
    sheet[f"J{row}"].value = tax_pay
    print(f"J{row}: {sheet[f'J{row}'].value}")

    wb.save(file_path)
    print("Data saved to file.")
    messagebox.showinfo("Успех", "Данные успешно добавлены")
    income_var.set("")

def create_gui():
    root = Tk()
    root.title("Добавление квартальных данных")
    
    Label(root, text="Введите квартальный доход:").grid(row=0, column=0, padx=10, pady=10)
    
    global income_var
    income_var = StringVar()
    Entry(root, textvariable=income_var).grid(row=0, column=1, padx=10, pady=10)
    
    Button(root, text="Добавить данные", command=add_data).grid(row=1, column=0, columnspan=2, padx=10, pady=10)
    
    root.mainloop()

if __name__ == "__main__":
    create_gui()

"""
Желательно вынести куда то в заголовки важные константы : 
tax_rate = 0.05

неправильно отрабатывается логика обнуления нарастающего итога
должно быть так:
если 1 й квартал то 
 total_income = quarterly_income
 если нет то
  total_income = quarterly_income + prev_total_income

  точно та же логика и для   total_tax:
  если квартал 1
  total_tax = total_income * tax_rate
  prev_tax = 0
  иначе 
    total_tax = prev_total_tax +  quarterly_income * tax_rate

Нужно построить общую логику прогрммы так:
при вызове программа выводит  2 последних строки имеющихся записей
и запрашивает дальнейшее - либо выйти ничего не сделав либо запросить ввод 
данных по текущему кварталу для обработки и тут же показать бывшую последнюю строку
а под ней другим цветом например введенные и вычисленные данные по текущему кварталу. 
с акцентом на сумму должного к оплате налога за текущий квартал
"""