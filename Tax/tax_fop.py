import openpyxl
from openpyxl.utils import get_column_letter
from tkinter import Tk, Button, Label, Entry, StringVar, messagebox

# Путь к файлу Excel
file_path = r"G:\Programming\Py\Tax\data\L_02a.xlsx"

def load_workbook():
    wb = openpyxl.load_workbook(file_path)
    return wb

def get_next_quarter(current_quarter):
    quarter, year = map(int, current_quarter.split("_"))
    quarter += 1
    if quarter > 4:
        quarter = 1
        year += 1
    return f"{quarter:02d}_{year:04d}"

def add_data():
    wb = load_workbook()
    sheet = wb.active
    target_column = "E"

    # Найти первую пустую строку в столбце E
    row = 1
    while sheet[f"{target_column}{row}"].value:
        row += 1

    prev_quarter = sheet[f"{target_column}{row - 1}"].value if row > 1 else "00_0000"
    current_quarter = get_next_quarter(prev_quarter)
    sheet[f"{target_column}{row}"].value = current_quarter

    # Сбрасываем накопленные данные для первого квартала нового года
    if current_quarter.startswith("01"):
        if row > 1:
            sheet[f"{get_column_letter(ord(target_column) + 1)}{row - 1}"].value = 0
            sheet[f"{get_column_letter(ord(target_column) + 2)}{row - 1}"].value = 0
            sheet[f"{get_column_letter(ord(target_column) + 3)}{row - 1}"].value = 0
            sheet[f"{get_column_letter(ord(target_column) + 4)}{row - 1}"].value = 0

    # Ввод квартального дохода
    try:
        quarterly_income = float(income_var.get())
    except ValueError:
        messagebox.showerror("Ошибка", "Некорректный доход")
        return

    sheet[f"{get_column_letter(ord(target_column) + 1)}{row}"].value = quarterly_income
    prev_total_income = sheet[f"{get_column_letter(ord(target_column) + 2)}{row - 1}"].value or 0
    sheet[f"{get_column_letter(ord(target_column) + 2)}{row}"].value = (
        quarterly_income + prev_total_income
    )
    sheet[f"{get_column_letter(ord(target_column) + 3)}{row}"].value = (
        sheet[f"{get_column_letter(ord(target_column) + 2)}{row}"].value * 0.05
    )
    prev_total_tax = sheet[f"{get_column_letter(ord(target_column) + 3)}{row - 1}"].value or 0
    sheet[f"{get_column_letter(ord(target_column) + 4)}{row}"].value = prev_total_tax
    sheet[f"{get_column_letter(ord(target_column) + 5)}{row}"].value = (
        sheet[f"{get_column_letter(ord(target_column) + 3)}{row}"].value
        - sheet[f"{get_column_letter(ord(target_column) + 4)}{row}"].value
    )

    wb.save(file_path)
    messagebox.showinfo("Успех", "Данные успешно добавлены")
    income_var.set("")

def create_gui():
    root = Tk()
    root.title("Добавление квартальных данных")
    
    Label(root, text="Введите квартальный доход:").grid(row=0, column=0, padx=10, pady=10)
    
    global income_var
    income_var = StringVar()
    Entry(root, textvariable=income_var).grid(row=0, column=1, padx=10, pady=10)
    
    Button(root, text="Добавить данные", command=add_data).grid(row=1, column=0, columnspan=2, padx=10, pady=10)
    
    root.mainloop()

if __name__ == "__main__":
    create_gui()
