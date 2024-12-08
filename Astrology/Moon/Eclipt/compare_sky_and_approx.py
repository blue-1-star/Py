import os
from datetime import datetime, timezone
from math import degrees
from skyfield.api import load
import math
import csv

def approx_moon(date=None):
    import math
    from datetime import datetime

    if date is None:
        date = datetime.now(timezone.utc)
    base_date = datetime(2000, 1, 1, tzinfo=timezone.utc)
    delta = date - base_date  # Разница между датами, результат - timedelta
    julian_date = delta.total_seconds()/86400 + 2451545.0

    MEAN_LONGITUDE_MOON = 218.316
    MEAN_ELONGATION = 297.850
    MEAN_ANOMALY = 134.963
    MEAN_DISTANCE = 93.272
    ASCENDING_NODE = 125.044  # Средняя долгота восходящего узла
    INCLINATION = 5.145  # Средний наклон лунной орбиты

    days_since_epoch = julian_date - 2451545.0
    # Число 2451545.0 — это Юлианская дата эпохи J2000.0, стандартный момент отсчёта времени в астрономии. Этот момент соответствует:
    # 1 января 2000 года в 12:00 UTC.
    """
    Юлианская дата (JD) — это система счёта времени, используемая в астрономии. Она измеряет время в днях (и дробях дня)
    от полудня 1 января 4713 года до н. э. по Юлианскому календарю.
     Юлианская дата J2000.0:  JD=2451545.0
     Это фиксированная точка, с которой ведётся отсчёт дней в современных астрономических расчётах.
    """
    mean_longitude = MEAN_LONGITUDE_MOON + 13.176396 * days_since_epoch
    # 13.176396 - средняя угловая скорость движения Луны по её орбите вокруг Земли (синодическое движение).
    """
     Луна проходит полный оборот по своей орбите за 27.321661 дней (средний сидерический месяц). Угловая скорость:
    360/27.321661 дней ≈ 13.17639∘/день 
    Это скорость увеличения средней долготы Луны ( mean longitude  и её элонгации ( elongation ).

    13.064993 (в mean_anomaly) средняя угловая скорость изменения средней аномалии Луны. 
    360 / 27.55455  ~ 13.064993 / день
    27.554550 дней — это длина аномалистического месяца (время между двумя последовательными перигеями орбиты Луны).

    -0.0529539 (в ascending_node)  - скорость регрессии (обратного движения) восходящего узла Луны.
    Лунная орбита медленно прецессирует (вращается) под влиянием гравитации Земли и Солнца.
    Узлы орбиты (точки, где орбита Луны пересекает плоскость эклиптики) смещаются в обратном направлении с угловой скоростью:
    −360∘/6798.383 дней  ≈−0.0529539∘ /день 
    Это используется для определения положения узлов орбиты, что важно при расчётах, связанных с затмениями.
    """
    mean_anomaly = MEAN_ANOMALY + 13.064993 * days_since_epoch
    elongation = MEAN_ELONGATION + 13.176396 * days_since_epoch
    ascending_node = ASCENDING_NODE - 0.0529539 * days_since_epoch

    # Истинная долгота Луны
    longitude = mean_longitude + 6.289 * math.sin(math.radians(mean_anomaly))

    # Аргумент широты
    argument_of_latitude = longitude - ascending_node

    # Истинная широта Луны
    latitude = INCLINATION * math.sin(math.radians(argument_of_latitude))

    # Расстояние
    distance = 385000 * (1 - 0.0549**2) / (1 + 0.0549 * math.cos(math.radians(mean_anomaly)))

    # Приведение долготы в диапазон [0, 360)
    longitude %= 360

    # Долгота Солнца (упрощённо)
    # sun_longitude = (280.460 + 0.9856474 * days_since_epoch) % 360
    sun_longitude = (280.460 + (360 / 365.25) * days_since_epoch) % 360
    # 365.25 — это средняя продолжительность года с учётом високосных лет.
    # Число 280.460 в формуле для расчёта долготы Солнца представляет собой среднюю эклиптическую долготу Солнца
    # (в градусах) на момент Юлианской эпохи J2000.0,то есть 1 января 2000 года в 12:00 по всемирному времени (UTC).

    # Разница долготы Луны и Солнца
    phase_angle = (longitude - sun_longitude) % 360

    # Фаза Луны (с поправкой на широту)
    moon_phase = (1 - math.cos(math.radians(phase_angle))) / 2

    # Лунный день
    synodic_month = 29.53059  # Средняя длина синодического месяца
    lunar_day = (phase_angle / 360) * synodic_month

    return {
        "longitude": longitude,
        "latitude": latitude,  # Значение в пределах [-5.145, 5.145]
        "distance_km": distance,
        "moon_phase": moon_phase,
        "lunar_day": lunar_day  # Значение в диапазоне [0, 29.53]
    }


# --- Точные вычисления Skyfield ---
def skyfield_moon(date=None):
    if date is None:
        date = datetime.now()

    ts = load.timescale()
    t = ts.utc(date.year, date.month, date.day, date.hour, date.minute, date.second)

    eph = load('de421.bsp')
    earth = eph['earth']
    moon = eph['moon']

    astrometric = earth.at(t).observe(moon).apparent()
    ecliptic = astrometric.ecliptic_latlon()

    sun = eph['sun']
    # elongation = degrees(earth.at(t).observe(moon).apparent().separation_from(earth.at(t).observe(sun).apparent()))
    elongation = earth.at(t).observe(moon).apparent().separation_from(earth.at(t).observe(sun).apparent()).degrees

    moon_phase = (1 - math.cos(math.radians(elongation))) / 2
    lunar_day = (elongation % 360) / 12.2

    return {
        "longitude": ecliptic[1].degrees,
        "latitude": ecliptic[0].degrees,
        "distance_km": astrometric.distance().km,
        "moon_phase": moon_phase,
        "lunar_day": lunar_day
    }


# --- Извлечение дат из файла ---
def extract_timestamps(file_path):
    if not os.path.exists(file_path):
        return []

    dates = []
    with open(file_path, 'r') as file:
        for line in file:
            if line.startswith("Время расчёта:"):
                try:
                    timestamp = line.split(": ", 1)[1].strip()
                    date = datetime.strptime(timestamp, '%Y-%m-%d %H:%M:%S')
                    dates.append(date)
                except ValueError:
                    continue  # Пропускаем строки с некорректным форматом времени
    return dates

def get_full_file_path(filename):
    """
    Возвращает полный путь к файлу данных, находящемуся в той же директории,
    что и текущий скрипт, или в указанной поддиректории.
    """
    script_dir = os.path.dirname(os.path.abspath(__file__))  # Путь к директории скрипта
    return os.path.join(script_dir, filename)

import openpyxl
from openpyxl.styles import Alignment
from skyfield.api import load
from math import radians, degrees

# Предполагается, что функция approx_moon уже определена
# и возвращает данные в виде словаря с ключами:
# "longitude", "latitude", "distance_km", "moon_phase", "lunar_day"


import openpyxl
from openpyxl.styles import Alignment, PatternFill, Border, Side
from datetime import datetime
from skyfield.api import load
from datetime import timezone

def compare_results(dates, filename="moon_comparison.xlsx"):
    """
    Сравнивает результаты между approx_moon и skyfield_moon для указанных дат.
    Результаты сохраняются в Excel-файл.

    :param dates: Список дат (datetime), для которых выполняется сравнение.
    :param filename: Имя Excel-файла для сохранения результатов.
    """
    file_path = get_full_file_path(filename)

    # Проверяем, существует ли файл, и если да, то загружаем его
    if os.path.exists(file_path):
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
    else:
        # Создаем новый файл, если его нет
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Comparison Results"
        sheet.append([
            "Date", "Approx Longitude", "Skyfield Longitude",
            "Approx Latitude", "Skyfield Latitude",
            "Approx Distance (km)", "Skyfield Distance (km)",
            "Approx Moon Phase", "Skyfield Moon Phase",
            "Approx Lunar Day", "Skyfield Lunar Day"
        ])

    # Заполнение данных для каждой даты
    for date in dates:
        date_with_timezone = date.replace(tzinfo=timezone.utc)

        # Получение результатов из approx_moon
        approx = approx_moon(date_with_timezone)

        # Получение результатов из skyfield_moon
        skyfield = skyfield_moon(date_with_timezone)

        # Добавляем строку в файл
        # sheet.append([
        #     date.strftime("%Y-%m-%d %H:%M:%S"),
        #     approx["longitude"], skyfield["longitude"],
        #     approx["latitude"], skyfield["latitude"],
        #     approx["distance_km"], skyfield["distance_km"],
        #     approx["moon_phase"], skyfield["moon_phase"],
        #     approx["lunar_day"], skyfield["lunar_day"]
        # ])
        row = [
            date.strftime("%Y-%m-%d %H:%M:%S"),
            round(approx["longitude"], 3), round(skyfield["longitude"], 3),
            round(approx["latitude"], 3), round(skyfield["latitude"], 3),
            round(approx["distance_km"], 2), round(skyfield["distance_km"], 2),
            round(approx["moon_phase"], 4), round(skyfield["moon_phase"], 4),
            round(approx["lunar_day"], 3), round(skyfield["lunar_day"], 3)
        ]
        sheet.append(row)
           # Применение стилей
    border = Border(
        left=Side(border_style="thin"),
        right=Side(border_style="thin"),
        top=Side(border_style="thin"),
        bottom=Side(border_style="thin")
    )

    # Выравнивание текста в заголовках и настройка ширины столбцов
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, max_col=sheet.max_column):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for column in sheet.columns:
        max_length = 0
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 1)
        sheet.column_dimensions[column[0].column_letter].width = adjusted_width

    # Применение попарного выделения для столбцов
        colors = ["FFFFCC", "CCFFFF"]  # Цвета для выделения
              
    for col_idx in range(2, len(sheet[1]), 2):  # Попарно по 2 столбца (начиная со второго столбца)
        color_fill = PatternFill(start_color=colors[(col_idx // 2) % len(colors)],
                                 end_color=colors[(col_idx // 2) % len(colors)],
                                 fill_type="solid")
        for row in sheet.iter_rows(min_row=2, max_col=col_idx + 1, min_col=col_idx, max_row=sheet.max_row):
            for cell in row:
                cell.fill = color_fill

    # Сохранение файла
    workbook.save(file_path)
    print(f"Results saved to {file_path}")

# --- Основная логика ---
# file_path_ld = 'lunar_data.txt'
# dates = extract_timestamps(file_path_ld)
dates = []
if not dates:
    dates = [datetime.now()]

compare_results(dates)
print("Таблица сравнений сохранена в 'comparison_results.csv'.")
