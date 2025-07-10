import os
import pandas as pd
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import matplotlib.pyplot as plt
from datetime import datetime, timedelta

# Регистрация шрифтов
pdfmetrics.registerFont(TTFont('Arial', 'arial.ttf'))  # Обычный шрифт
pdfmetrics.registerFont(TTFont('Arial-Bold', 'arialbd.ttf'))  # Жирный шрифт

# Функция для перевода номера месяца в название с учетом сдвига на один месяц назад
def get_month_name(month_number):
    months = {
        1: "январь", 2: "февраль", 3: "март", 4: "апрель",
        5: "май", 6: "июнь", 7: "июль", 8: "август",
        9: "сентябрь", 10: "октябрь", 11: "ноябрь", 12: "декабрь"
    }
    # Сдвиг на один месяц назад
    previous_month = month_number - 1 if month_number > 1 else 12
    return months.get(previous_month, "неизвестный месяц")

# Функция для подготовки данных отчета
def prepare_report_data(file_path, sheet_name, val_cell, txt_list):
    # Чтение данных из Excel с принудительным закрытием файла
    with pd.ExcelFile(file_path) as excel_file:
        df = pd.read_excel(excel_file, sheet_name=sheet_name)
    
    # Преобразование столбца 'Date' в формат datetime
    df['Date'] = pd.to_datetime(df['Date'], format='%d.%m.%Y')
    print(df['heating'],df['Total'])
    # Получение последней строки данных
    last_row = df.iloc[-1]
    prev_row = df.iloc[-2]
    
    # Извлечение месяца и года из столбца 'Date'
    date = last_row['Date']
    month = date.month
    year = date.year
    month_name = get_month_name(month)  # Учитываем сдвиг на один месяц назад
    suffix = f"{month:02d}_{year}"
    
    # Вычисление расхода электроэнергии и воды
    el_count = last_row['El_count'] - prev_row['El_count']
    wat_count = last_row['Wat_count'] - prev_row['Wat_count']
    
    # Формирование данных для отчета
    report_data = [
        [txt_list[0], f"{last_row['El_bill']:.2f} грн", txt_list[1], f"{el_count:.2f} кВт·ч"],
        [txt_list[2], f"{last_row['water']:.2f} грн", txt_list[3], f"{wat_count:.2f} м³"],
        [txt_list[4], f"{last_row['utilities']:.2f} грн", "", ""],
        [txt_list[5], f"{last_row['TV & Internet']:.2f} грн", "", ""],
        [txt_list[6], f"{last_row['litter']:.2f} грн", "", ""],
        [txt_list[7], f"{last_row['heating']:.2f} грн", "", ""],
        [txt_list[8], f"{last_row['Total']:.2f} грн", "", ""]
    ]
    
    return report_data, month_name, year, suffix

# Функция для создания PDF-отчета
def create_report(inp_dir, excel_file, sheet_name, val_cell, txt_list, title):
    # Полный путь к файлу
    file_path = os.path.join(inp_dir, excel_file)
    
    # Подготовка данных
    report_data, month_name, year, suffix = prepare_report_data(file_path, sheet_name, val_cell, txt_list)
    
    # Создание текстового файла
    txt_report_path = os.path.join(inp_dir, f"report_{suffix}.txt")
    with open(txt_report_path, 'w', encoding='utf-8') as txt_file:
        txt_file.write(title.replace("__", month_name).replace("++++", str(year)) + "\n\n")
        txt_file.write("+----------------+----------+----------------+----------+\n")
        txt_file.write("| Показатель     | Сумма    | Показатель     | Сумма    |\n")
        txt_file.write("+----------------+----------+----------------+----------+\n")
        for row in report_data:
            line = f"| {row[0].ljust(14)} | {row[1].rjust(8)} | {row[2].ljust(14)} | {row[3].rjust(8)} |"
            txt_file.write(line + "\n")
        txt_file.write("+----------------+----------+----------------+----------+\n")
    
    # Создание PDF-файла
    pdf_report_path = os.path.join(inp_dir, f"report_{suffix}.pdf")
    
    doc = SimpleDocTemplate(pdf_report_path, pagesize=A4)
    elements = []
    
    # Заголовок
    styles = getSampleStyleSheet()
    styles['Title'].fontName = 'Arial'  # Используем шрифт с поддержкой кириллицы
    title_text = title.replace("__", month_name).replace("++++", str(year))
    elements.append(Paragraph(title_text, styles['Title']))
    
    # Таблица с данными
    table_data = [
        ["Показатель", "Сумма", "Показатель", "Сумма"]
    ] + report_data
    
    table = Table(table_data, colWidths=[100, 80, 100, 80])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),  # Заголовок таблицы
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('ALIGN', (3, 0), (3, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Arial'),  # Шрифт для заголовка
        ('FONTNAME', (0, 1), (-1, -1), 'Arial'),  # Шрифт для данных
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),  # Фон для данных
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightgreen),  # Фон для строки "ВСЕГО"
        ('FONTNAME', (0, -1), (-1, -1), 'Arial-Bold'),  # Жирный шрифт для строки "ВСЕГО"
    ]))
    
    elements.append(table)
    doc.build(elements)

# Функция для записи отчета в новый Excel-файл
def write_report_to_excel(inp_dir, report_data, suffix, title_text):
    # Создание нового файла Excel
    report_file_path = os.path.join(inp_dir, f"report_{suffix}.xlsx")
    book = Workbook()
    sheet = book.active
    
    # Запись заголовка
    sheet.append([title_text])
    sheet.append([])  # Пустая строка для разделения
    
    # Запись данных в новый лист
    headers = ["Показатель", "Сумма", "Показатель", "Сумма"]
    sheet.append(headers)
    for row in report_data:
        sheet.append(row)
    
    # Настройка ширины столбцов
    column_widths = [20, 15, 20, 15]  # Ширина столбцов
    for i, width in enumerate(column_widths, start=1):
        sheet.column_dimensions[get_column_letter(i)].width = width
    
    # Стили для заголовка
    header_font = Font(bold=True, color="FFFFFF")  # Белый жирный шрифт
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # Синий фон
    for cell in sheet[3]:  # Заголовок таблицы (строка 3)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Стили для данных
    data_font = Font(name='Arial', size=11)
    for row in sheet.iter_rows(min_row=4, max_row=sheet.max_row - 1):
        for cell in row:
            cell.font = data_font
            cell.alignment = Alignment(horizontal='left')
    
    # Стили для строки "ВСЕГО"
    total_row = sheet.max_row
    for cell in sheet[total_row]:
        cell.font = Font(bold=True, color="000000")  # Жирный черный шрифт
        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Зеленый фон
    
    # Выравнивание заголовка отчета по центру
    sheet['A1'].font = Font(bold=True, size=14)
    sheet['A1'].alignment = Alignment(horizontal='center')
    sheet.merge_cells('A1:D1')  # Объединение ячеек для заголовка
    
    # Сохранение изменений в файл
    book.save(report_file_path)

# Параметры
# inp_dir = r"D:\OneDrive\Документы"
inp_dir = r"G:\Flat"
excel_file = "Flat_Arn.xlsx"
# excel_file = "Flat_Arn3.xlsx"
sheet_name = "Push"
val_cell = ['El_bill', 'El_count', 'water', 'Wat_count', 'utilities', 'TV & Internet', 'litter', 'heating', 'Total']
txt_list = ['Свет', 'Расход эл.энергии', 'вода', 'Расход воды', 'квартплата', 'TV', 'мусор', 'отопление', 'ВСЕГО']
title = "Коммунальные расходы за __  ++++ года"

# Подготовка данных
file_path = os.path.join(inp_dir, excel_file)
report_data, month_name, year, suffix = prepare_report_data(file_path, sheet_name, val_cell, txt_list)

# Создание PDF-отчета
create_report(inp_dir, excel_file, sheet_name, val_cell, txt_list, title)

# Запись отчета в новый Excel-файл (опционально)
write_report_to_excel(inp_dir, report_data, suffix, title.replace("__", month_name).replace("++++", str(year)))