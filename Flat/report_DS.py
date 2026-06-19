# report_mac2.py - версия для macOS и Windows
import os
import sys
import platform
import pandas as pd
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
from datetime import datetime, timedelta
from pathlib import Path
import shutil

# Добавляем корень Py/ в путь для импорта config
PY_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PY_ROOT))

from config import paths

# ===== АВТОМАТИЧЕСКАЯ НАСТРОЙКА ШРИФТОВ =====
def setup_fonts():
    """Регистрирует ВСЕ шрифты"""
    system = platform.system()
    fonts_registered = False
    
    if system == "Darwin":  # macOS
        mac_fonts = [
            ('Arial', '/System/Library/Fonts/Supplemental/Arial.ttf'),
            ('Arial-Bold', '/System/Library/Fonts/Supplemental/Arial Bold.ttf'),
            ('Arial', '/Library/Fonts/Arial.ttf'),
            ('Arial-Bold', '/Library/Fonts/Arial Bold.ttf'),
        ]
        
        for name, path in mac_fonts:
            if os.path.exists(path):
                try:
                    pdfmetrics.registerFont(TTFont(name, path))
                    fonts_registered = True
                    print(f"✅ Шрифт {name} загружен из: {path}")
                except Exception as e:
                    print(f"⚠️ Ошибка загрузки {name}: {e}")
        
        if not fonts_registered:
            print("⚠️ Шрифты Arial не найдены")
        else:
            try:
                pdfmetrics.getFont('Arial-Bold')
                print("✅ Arial-Bold успешно зарегистрирован")
            except:
                print("⚠️ Arial-Bold не зарегистрирован, использую Arial для жирного")
                if 'Arial' in pdfmetrics._fonts:
                    pdfmetrics.registerFont(TTFont('Arial-Bold', 
                        pdfmetrics._fonts['Arial'].face.name))
    
    elif system == "Windows":
        windows_fonts = [
            ('Arial', 'C:\\Windows\\Fonts\\arial.ttf'),
            ('Arial-Bold', 'C:\\Windows\\Fonts\\arialbd.ttf'),
        ]
        for name, path in windows_fonts:
            if os.path.exists(path):
                try:
                    pdfmetrics.registerFont(TTFont(name, path))
                    fonts_registered = True
                    print(f"✅ Шрифт {name} загружен")
                except:
                    continue
    
    return fonts_registered

# Регистрируем шрифты
FONTS_LOADED = setup_fonts()

# Функция для перевода номера месяца в название с учетом сдвига на один месяц назад
def get_month_name(month_number):
    months = {
        1: "январь", 2: "февраль", 3: "март", 4: "апрель",
        5: "май", 6: "июнь", 7: "июль", 8: "август",
        9: "сентябрь", 10: "октябрь", 11: "ноябрь", 12: "декабрь"
    }
    previous_month = month_number - 1 if month_number > 1 else 12
    return months.get(previous_month, "неизвестный месяц")

# Функция для подготовки данных отчета
def prepare_report_data(file_path, sheet_name, val_cell, txt_list):
    with pd.ExcelFile(file_path) as excel_file:
        df = pd.read_excel(excel_file, sheet_name=sheet_name)
    
    print("📅 Преобразую Excel даты...")
    
    try:
        numeric_series = pd.to_numeric(df['Date'], errors='coerce')
        
        if numeric_series.between(1, 50000).any():
            print("   Обнаружены числа Excel (диапазон 1-50000)")
            df['Date'] = pd.to_datetime(numeric_series, unit='D', origin='1899-12-30')
        else:
            raise ValueError("Не похоже на Excel даты")
            
    except Exception as e:
        print(f"   Не удалось преобразовать как числа: {e}")
        first_val = str(df['Date'].iloc[0])
        print(f"   Пробуем как строки: {first_val}")
        
        if '.' in first_val:
            df['Date'] = pd.to_datetime(df['Date'], format='%d.%m.%Y', errors='coerce')
        elif '/' in first_val:
            df['Date'] = pd.to_datetime(df['Date'], format='%d/%m/%Y', errors='coerce')
        else:
            df['Date'] = pd.to_datetime(df['Date'], dayfirst=True, errors='coerce')
    
    print(f"   Первая дата после преобразования: {df['Date'].iloc[0]}")
    print(f"   Последняя дата: {df['Date'].iloc[-1]}")
    
    df = df.dropna(subset=['Date'])
    df = df.sort_values('Date')
    
    print(f"✅ Диапазон дат: {df['Date'].min().strftime('%d.%m.%Y')} -> {df['Date'].max().strftime('%d.%m.%Y')}")
    
    last_row = df.iloc[-1]
    prev_row = df.iloc[-2]
    
    date = last_row['Date']
    month = date.month
    year = date.year
    month_name = get_month_name(month)
    suffix = f"{month:02d}_{year}"
    
    print(f"📅 Отчет за {month_name} {year}")
    
    el_count = last_row['El_count'] - prev_row['El_count']
    wat_count = last_row['Wat_count'] - prev_row['Wat_count']
    
    report_data = [
        [txt_list[0], f"{last_row['El_bill']:.2f} грн", txt_list[1], f"{el_count:.2f} кВт·ч"],
        [txt_list[2], f"{last_row['water']:.2f} грн", txt_list[3], f"{wat_count:.2f} м³"],
        [txt_list[4], f"{last_row['utilities']:.2f} грн", "", ""],
        [txt_list[5], f"{last_row['TV & Internet']:.2f} грн", "", ""],
        [txt_list[6], f"{last_row['litter']:.2f} грн", "", ""],
        [txt_list[7], f"{last_row['heating']:.2f} грн", "", ""],
        [txt_list[8], f"{last_row['Total']:.2f} грн", "", ""]
    ]
    
    return report_data, month_name, year, suffix

# ===== ФУНКЦИЯ ДЛЯ ДОБАВЛЕНИЯ В EXCEL-ИСТОРИЮ =====
def append_to_excel_history(report_data, month_name, year, suffix, script_dir):
    """
    Добавляет новый месяц в Excel-файл истории
    """
    local_data_dir = script_dir / "Data"
    local_data_dir.mkdir(parents=True, exist_ok=True)
    
    history_file = local_data_dir / "utilities_history.xlsx"
    
    # Подготовка данных для нового месяца
    new_row = {
        "Месяц": f"{month_name} {year}",
        "Свет (грн)": float(report_data[0][1].split()[0]),
        "Расход эл.энергии (кВт·ч)": float(report_data[0][3].split()[0]),
        "Вода (грн)": float(report_data[1][1].split()[0]),
        "Расход воды (м³)": float(report_data[1][3].split()[0]),
        "Квартплата (грн)": float(report_data[2][1].split()[0]),
        "TV (грн)": float(report_data[3][1].split()[0]),
        "Мусор (грн)": float(report_data[4][1].split()[0]),
        "Отопление (грн)": float(report_data[5][1].split()[0]),
        "ВСЕГО (грн)": float(report_data[6][1].split()[0]),
    }
    
    # Проверяем, существует ли файл истории
    if history_file.exists():
        # Открываем существующий файл
        book = load_workbook(history_file)
        sheet = book.active
        
        # Находим первую пустую строку
        next_row = sheet.max_row + 1
    else:
        # Создаём новый файл
        book = Workbook()
        sheet = book.active
        
        # Заголовки
        headers = list(new_row.keys())
        for col, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col, value=header)
            sheet.cell(row=1, column=col).font = Font(bold=True)
            sheet.cell(row=1, column=col).fill = PatternFill(
                start_color="4F81BD", end_color="4F81BD", fill_type="solid"
            )
            sheet.cell(row=1, column=col).font = Font(bold=True, color="FFFFFF")
        
        next_row = 2
    
    # Добавляем новую строку
    for col, (key, value) in enumerate(new_row.items(), 1):
        sheet.cell(row=next_row, column=col, value=value)
    
    # Настройка ширины столбцов
    for col in range(1, len(new_row) + 1):
        column_letter = get_column_letter(col)
        sheet.column_dimensions[column_letter].width = 20
    
    # Сохраняем
    book.save(history_file)
    print(f"✅ Данные добавлены в историю: {history_file}")
    print(f"   Добавлен месяц: {month_name} {year}")

# Функция для создания PDF-отчета
def create_report(inp_dir, excel_file, sheet_name, val_cell, txt_list, title):
    file_path = os.path.join(inp_dir, excel_file)
    
    report_data, month_name, year, suffix = prepare_report_data(file_path, sheet_name, val_cell, txt_list)
    title_text = title.replace("__", month_name).replace("++++", str(year))
    
    # ===== 1. СОЗДАНИЕ PDF =====
    # PDF в OneDrive
    pdf_onedrive_path = os.path.join(inp_dir, f"report_{suffix}.pdf")
    
    doc = SimpleDocTemplate(pdf_onedrive_path, pagesize=A4)
    elements = []
    
    styles = getSampleStyleSheet()
    if FONTS_LOADED:
        styles['Title'].fontName = 'Arial'
    
    elements.append(Paragraph(title_text, styles['Title']))
    
    table_data = [["Показатель", "Сумма", "Показатель", "Сумма"]] + report_data
    
    table = Table(table_data, colWidths=[100, 80, 100, 80])
    
    table_style = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('ALIGN', (3, 0), (3, -1), 'RIGHT'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightgreen),
    ]
    
    if FONTS_LOADED:
        table_style.append(('FONTNAME', (0, 0), (-1, 0), 'Arial'))
        table_style.append(('FONTNAME', (0, 1), (-1, -1), 'Arial'))
        table_style.append(('FONTNAME', (0, -1), (-1, -1), 'Arial-Bold'))
    
    table.setStyle(TableStyle(table_style))
    elements.append(table)
    doc.build(elements)
    
    print(f"✅ PDF сохранён в OneDrive: {pdf_onedrive_path}")
    
    # ===== 2. КОПИРОВАНИЕ PDF В ЛОКАЛЬНУЮ ПАПКУ DATA (для Git) =====
    script_dir = Path(__file__).parent
    local_data_dir = script_dir / "Data"
    local_data_dir.mkdir(parents=True, exist_ok=True)
    
    pdf_local_path = local_data_dir / f"report_{suffix}.pdf"
    shutil.copy2(pdf_onedrive_path, pdf_local_path)
    print(f"✅ PDF скопирован в локальную папку: {pdf_local_path}")
    
    # ===== 3. ДОБАВЛЕНИЕ В EXCEL-ИСТОРИЮ =====
    append_to_excel_history(report_data, month_name, year, suffix, script_dir)

# ===== ОСНОВНАЯ ПРОГРАММА =====
if __name__ == "__main__":
    # Параметры (пути берутся из config.py)
    inp_dir = str(paths.ONEDRIVE)  # корень OneDrive
    excel_file = "Flat_Arn.xlsx"
    sheet_name = "Push"
    val_cell = ['El_bill', 'El_count', 'water', 'Wat_count', 'utilities', 'TV & Internet', 'litter', 'heating', 'Total']
    txt_list = ['Свет', 'Расход эл.энергии', 'вода', 'Расход воды', 'квартплата', 'TV', 'мусор', 'отопление', 'ВСЕГО']
    title = "Коммунальные расходы за __  ++++ года"
    
    # Проверяем, существует ли файл
    file_path = os.path.join(inp_dir, excel_file)
    if not os.path.exists(file_path):
        print(f"❌ Файл не найден: {file_path}")
        sys.exit(1)
    
    # Подготовка данных
    report_data, month_name, year, suffix = prepare_report_data(file_path, sheet_name, val_cell, txt_list)
    
    # Создание отчётов
    create_report(inp_dir, excel_file, sheet_name, val_cell, txt_list, title)
    
    print("\n🎉 Отчет успешно создан!")
    print(f"   📁 OneDrive (PDF): {inp_dir}")
    print(f"   📁 Локальная копия PDF: {Path(__file__).parent / 'Data'}")
    print(f"   📊 История расходов: {Path(__file__).parent / 'Data' / 'utilities_history.xlsx'}")