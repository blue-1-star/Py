# report_DS_5b.py - версия для macOS и Windows
import os
import sys
import platform
import pandas as pd
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
from datetime import datetime, timedelta
from pathlib import Path

# Добавляем корень Py/ в путь
PY_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PY_ROOT))

from config import paths
# ===== АВТОМАТИЧЕСКАЯ НАСТРОЙКА ШРИФТОВ =====
def setup_fonts():
    """Регистрирует ВСЕ шрифты"""
    
    system = platform.system()
    fonts_registered = False
    
    if system == "Darwin":  # macOS
        mac_fonts = [
            ('Arial', '/System/Library/Fonts/Supplemental/Arial.ttf'),
            ('Arial-Bold', '/System/Library/Fonts/Supplemental/Arial Bold.ttf'),
            # Альтернативные варианты
            ('Arial', '/Library/Fonts/Arial.ttf'),
            ('Arial-Bold', '/Library/Fonts/Arial Bold.ttf'),
        ]
        
        # Регистрируем ВСЕ найденные шрифты
        for name, path in mac_fonts:
            if os.path.exists(path):
                try:
                    pdfmetrics.registerFont(TTFont(name, path))
                    fonts_registered = True
                    print(f"✅ Шрифт {name} загружен из: {path}")
                    # УБРАЛИ break - продолжаем регистрировать остальные!
                except Exception as e:
                    print(f"⚠️ Ошибка загрузки {name}: {e}")
        
        # Проверяем, что оба шрифта загружены
        if not fonts_registered:
            print("⚠️ Шрифты Arial не найдены")
        else:
            # Проверяем, есть ли Arial-Bold
            try:
                pdfmetrics.getFont('Arial-Bold')
                print("✅ Arial-Bold успешно зарегистрирован")
            except:
                print("⚠️ Arial-Bold не зарегистрирован, использую Arial для жирного")
                # Если нет Bold, регистрируем Arial как Bold
                if 'Arial' in pdfmetrics._fonts:
                    pdfmetrics.registerFont(TTFont('Arial-Bold', 
                        pdfmetrics._fonts['Arial'].face.name))
    
    elif system == "Windows":
        # Аналогично для Windows
        windows_fonts = [
            ('Arial', 'C:\\Windows\\Fonts\\arial.ttf'),
            ('Arial-Bold', 'C:\\Windows\\Fonts\\arialbd.ttf'),
        ]
        for name, path in windows_fonts:
            if os.path.exists(path):
                try:
                    pdfmetrics.registerFont(TTFont(name, path))
                    fonts_registered = True
                    print(f"✅ Шрифт {name} загружен")
                except:
                    continue
    
    return fonts_registered

# Регистрируем шрифты
FONTS_LOADED = setup_fonts()

# Функция для перевода номера месяца в название с учетом сдвига на один месяц назад
def get_month_name(month_number):
    months = {
        1: "январь", 2: "февраль", 3: "март", 4: "апрель",
        5: "май", 6: "июнь", 7: "июль", 8: "август",
        9: "сентябрь", 10: "октябрь", 11: "ноябрь", 12: "декабрь"
    }
    previous_month = month_number - 1 if month_number > 1 else 12
    return months.get(previous_month, "неизвестный месяц")

# Функция для подготовки данных отчета
def prepare_report_data(file_path, sheet_name, val_cell, txt_list):
    # Чтение файла
    with pd.ExcelFile(file_path) as excel_file:
        df = pd.read_excel(excel_file, sheet_name=sheet_name)
    
    # ===== ПРИНУДИТЕЛЬНОЕ ПРЕОБРАЗОВАНИЕ EXCEL ДАТ =====
    print("📅 Преобразую Excel даты...")
    
    # Способ 1: Принудительно преобразуем в числа
    try:
        # Преобразуем столбец в числа (если это строки с числами)
        numeric_series = pd.to_numeric(df['Date'], errors='coerce')
        
        # Проверяем, есть ли числа в разумном диапазоне Excel дат (от 1 до 50000)
        if numeric_series.between(1, 50000).any():
            print("   Обнаружены числа Excel (диапазон 1-50000)")
            df['Date'] = pd.to_datetime(numeric_series, unit='D', origin='1899-12-30')
        else:
            # Если числа не в диапазоне Excel, пробуем как строки
            raise ValueError("Не похоже на Excel даты")
            
    except Exception as e:
        print(f"   Не удалось преобразовать как числа: {e}")
        # Пробуем как строки
        first_val = str(df['Date'].iloc[0])
        print(f"   Пробуем как строки: {first_val}")
        
        if '.' in first_val:
            df['Date'] = pd.to_datetime(df['Date'], format='%d.%m.%Y', errors='coerce')
        elif '/' in first_val:
            df['Date'] = pd.to_datetime(df['Date'], format='%d/%m/%Y', errors='coerce')
        else:
            df['Date'] = pd.to_datetime(df['Date'], dayfirst=True, errors='coerce')
    
    # Проверка результата
    print(f"   Первая дата после преобразования: {df['Date'].iloc[0]}")
    print(f"   Последняя дата: {df['Date'].iloc[-1]}")
    
    # Удаляем строки с неправильными датами
    df = df.dropna(subset=['Date'])
    
    # Сортируем по дате
    df = df.sort_values('Date')
    
    print(f"✅ Диапазон дат: {df['Date'].min().strftime('%d.%m.%Y')} -> {df['Date'].max().strftime('%d.%m.%Y')}")
    
    # Получение последней строки данных
    last_row = df.iloc[-1]
    prev_row = df.iloc[-2]
    
    date = last_row['Date']
    month = date.month
    year = date.year
    month_name = get_month_name(month)
    suffix = f"{month:02d}_{year}"
    
    print(f"📅 Отчет за {month_name} {year}")
    
    # Вычисление расхода
    el_count = last_row['El_count'] - prev_row['El_count']
    wat_count = last_row['Wat_count'] - prev_row['Wat_count']
    
    # Формирование данных для отчета
    report_data = [
        [txt_list[0], f"{last_row['El_bill']:.2f} грн", txt_list[1], f"{el_count:.2f} кВт·ч"],
        [txt_list[2], f"{last_row['water']:.2f} грн", txt_list[3], f"{wat_count:.2f} м³"],
        [txt_list[4], f"{last_row['utilities']:.2f} грн", "", ""],
        [txt_list[5], f"{last_row['TV & Internet']:.2f} грн", "", ""],
        [txt_list[6], f"{last_row['litter']:.2f} грн", "", ""],
        [txt_list[7], f"{last_row['heating']:.2f} грн", "", ""],
        [txt_list[8], f"{last_row['Total']:.2f} грн", "", ""]
    ]
    
    return report_data, month_name, year, suffix

# Функция для создания PDF-отчета
def create_report(inp_dir, excel_file, sheet_name, val_cell, txt_list, title):
    file_path = os.path.join(inp_dir, excel_file)
    
    report_data, month_name, year, suffix = prepare_report_data(file_path, sheet_name, val_cell, txt_list)
    
    # Создание текстового файла
    txt_report_path = os.path.join(inp_dir, f"report_{suffix}.txt")
    with open(txt_report_path, 'w', encoding='utf-8') as txt_file:
        txt_file.write(title.replace("__", month_name).replace("++++", str(year)) + "\n\n")
        txt_file.write("+----------------+----------+----------------+----------+\n")
        txt_file.write("| Показатель     | Сумма    | Показатель     | Сумма    |\n")
        txt_file.write("+----------------+----------+----------------+----------+\n")
        for row in report_data:
            line = f"| {row[0].ljust(14)} | {row[1].rjust(8)} | {row[2].ljust(14)} | {row[3].rjust(8)} |"
            txt_file.write(line + "\n")
        txt_file.write("+----------------+----------+----------------+----------+\n")
    
    # Создание PDF-файла
    pdf_report_path = os.path.join(inp_dir, f"report_{suffix}.pdf")
    
    doc = SimpleDocTemplate(pdf_report_path, pagesize=A4)
    elements = []
    
    # Заголовок
    styles = getSampleStyleSheet()
    
    # Устанавливаем шрифт для заголовка (если загружены)
    if FONTS_LOADED:
        styles['Title'].fontName = 'Arial'
    
    title_text = title.replace("__", month_name).replace("++++", str(year))
    elements.append(Paragraph(title_text, styles['Title']))
    
    # Таблица с данными
    table_data = [["Показатель", "Сумма", "Показатель", "Сумма"]] + report_data
    
    table = Table(table_data, colWidths=[100, 80, 100, 80])
    
    # Стили таблицы
    table_style = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('ALIGN', (3, 0), (3, -1), 'RIGHT'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightgreen),
    ]
    
    # Добавляем шрифты только если они загружены
    if FONTS_LOADED:
        table_style.append(('FONTNAME', (0, 0), (-1, 0), 'Arial'))
        table_style.append(('FONTNAME', (0, 1), (-1, -1), 'Arial'))
        # table_style.append(('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'))
        table_style.append(('FONTNAME', (0, -1), (-1, -1), 'Arial-Bold'))
    
    table.setStyle(TableStyle(table_style))
    
    elements.append(table)
    doc.build(elements)
    
    print(f"✅ PDF создан: {pdf_report_path}")

# Функция для записи отчета в новый Excel-файл
def write_report_to_excel(inp_dir, report_data, suffix, title_text):
    report_file_path = os.path.join(inp_dir, f"report_{suffix}.xlsx")
    book = Workbook()
    sheet = book.active
    
    sheet.append([title_text])
    sheet.append([])
    
    headers = ["Показатель", "Сумма", "Показатель", "Сумма"]
    sheet.append(headers)
    for row in report_data:
        sheet.append(row)
    
    column_widths = [20, 15, 20, 15]
    for i, width in enumerate(column_widths, start=1):
        sheet.column_dimensions[get_column_letter(i)].width = width
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for cell in sheet[3]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    data_font = Font(name='Arial', size=11)
    for row in sheet.iter_rows(min_row=4, max_row=sheet.max_row - 1):
        for cell in row:
            cell.font = data_font
            cell.alignment = Alignment(horizontal='left')
    
    total_row = sheet.max_row
    for cell in sheet[total_row]:
        cell.font = Font(bold=True, color="000000")
        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    
    sheet['A1'].font = Font(bold=True, size=14)
    sheet['A1'].alignment = Alignment(horizontal='center')
    sheet.merge_cells('A1:D1')
    
    book.save(report_file_path)
    print(f"✅ Excel создан: {report_file_path}")

# ===== ОСНОВНАЯ ПРОГРАММА =====
if __name__ == "__main__":
    # Параметры
    # inp_dir = "/Users/san/Library/CloudStorage/OneDrive-Personal/"
    inp_dir = str(paths.ONEDRIVE)
    excel_file = paths.FLAT_FILE.name   # 'Flat_Arn.xlsx'
    # excel_file = "Flat_Arn.xlsx"
    sheet_name = "Push"
    val_cell = ['El_bill', 'El_count', 'water', 'Wat_count', 'utilities', 'TV & Internet', 'litter', 'heating', 'Total']
    txt_list = ['Свет', 'Расход эл.энергии', 'вода', 'Расход воды', 'квартплата', 'TV', 'мусор', 'отопление', 'ВСЕГО']
    title = "Коммунальные расходы за __  ++++ года"
    
    # Проверяем, существует ли файл
    file_path = os.path.join(inp_dir, excel_file)
    if not os.path.exists(file_path):
        print(f"❌ Файл не найден: {file_path}")
        sys.exit(1)
    
    # Подготовка данных
    report_data, month_name, year, suffix = prepare_report_data(file_path, sheet_name, val_cell, txt_list)
    
    # Создание PDF-отчета
    create_report(inp_dir, excel_file, sheet_name, val_cell, txt_list, title)
    
    # Запись отчета в новый Excel-файл
    write_report_to_excel(inp_dir, report_data, suffix, title.replace("__", month_name).replace("++++", str(year)))
    
    print("🎉 Отчет успешно создан!")